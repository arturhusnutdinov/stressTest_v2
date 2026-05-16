#!/usr/bin/env python3
"""Sync RUSAL Excel BS sheet with cleaned DB state."""
import openpyxl, shutil
from pathlib import Path

SRC = Path(__file__).parent.parent / "companies/rusal/data/rusal_unified_complete.xlsx"

# Backup
shutil.copy2(SRC, SRC.with_suffix('.xlsx.bak_sync'))

wb = openpyxl.load_workbook(SRC)
ws = wb['history_bs']

# 1. Find and remove computed metric rows
COMPUTED = {'total_ca', 'total_cl', 'total_nca', 'total_ncl', 'total_liabilities'}
rows_to_delete = []
for r in range(4, ws.max_row + 1):
    metric = ws.cell(r, 1).value
    if metric and str(metric).strip() in COMPUTED:
        rows_to_delete.append(r)
        print(f"  DELETE row {r}: '{metric}' — computed by engine")

for r in sorted(rows_to_delete, reverse=True):
    ws.delete_rows(r)

# 2. Find last data row
last_row = 3
for r in range(3, ws.max_row + 1):
    if ws.cell(r, 1).value:
        last_row = r

# 3. Add missing canonical metrics
CANONICAL_TO_ADD = [
    'restricted_cash',
    'interest_payable',
    'payroll_payable',
    'treasury_stock',
    'total_liab_equity',
]

# Check what's already there
existing = set()
for r in range(4, ws.max_row + 1):
    m = ws.cell(r, 1).value
    if m: existing.add(str(m).strip())

insert_row = last_row + 1
for metric in CANONICAL_TO_ADD:
    if metric not in existing:
        ws.cell(insert_row, 1).value = metric
        print(f"  ADD row {insert_row}: '{metric}'")
        insert_row += 1
    else:
        print(f"  SKIP '{metric}' — already exists")

wb.save(SRC)
print(f"\nDone. BS sheet: {ws.max_row} rows.")
