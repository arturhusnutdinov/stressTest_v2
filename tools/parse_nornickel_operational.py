"""
parse_nornickel_operational.py — все операционные показатели Норникеля в Excel.

Источники: Databook (PRODUCTION DATA, ORE OUTPUT, RECOVERY RATES),
           Annual Report 2024, Factsheet 2024, Operational Results 2025.
"""
import openpyxl
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
COMPANY_DIR = ROOT / "companies" / "nornickel"
DATABOOK = COMPANY_DIR / "data" / "statements" / "Databook_12m_25_Final.xlsx"
EXCEL_OUT = COMPANY_DIR / "data" / "excel" / "nornickel_unified.xlsx"

TARGET_YEARS = list(range(2011, 2026))

def load_databook_sheet(name):
    wb = openpyxl.load_workbook(DATABOOK, data_only=True)
    ws = wb[name]
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append([str(c) if c is not None else None for c in row])
    wb.close()
    return rows

def get_yr_map(rows, header_row_idx):
    """Extract {year: col_idx} from header row."""
    hdr = rows[header_row_idx]
    yr_map = {}
    for i, val in enumerate(hdr):
        try:
            y = int(float(val))
            if 2009 <= y <= 2026:
                yr_map[y] = i
        except (ValueError, TypeError):
            pass
    return yr_map

def val_or_none(row, col):
    if col is None or col >= len(row) or row[col] is None:
        return None
    try:
        return float(row[col].replace(",", "").replace(" ", ""))
    except (ValueError, AttributeError):
        return None

print("Loading Databook...")
prod_rows = load_databook_sheet("PRODUCTION DATA")
prod_yr = get_yr_map(prod_rows, 0)

ore_rows = load_databook_sheet("ORE OUTPUT")
ore_yr = get_yr_map(ore_rows, 0)

rec_rows = load_databook_sheet("RECOVERY RATES")
rec_yr = get_yr_map(rec_rows, 0)

# Note: Databook has half-year columns interleaved. We only use full-year (integer) columns.
# But the header row has years like 2009, 1H 10, 2010, 1H 11, 2011...
# We need the full-year columns.
# The production header row 0 has: 2009, 1H 10, 2010, 1H 11, 2011...
# So full years are at even indices after 2009? Let me check.

# Actually prod_yr picks up ALL year-ish numbers including half-years.
# Let me filter to only full years (2009 < val < 2026 and val % 1 == 0)
# Wait, they're stored as strings -> float conversion -> int check already done.
# The problem is 1H 10 columns are skipped because they're strings like "1H 10" not ints.
# Let me verify...

print(f"Production years: {sorted(prod_yr.keys())}")
print(f"Ore years: {sorted(ore_yr.keys())}")
print(f"Recovery years: {sorted(rec_yr.keys())}")

# Open Excel
wb = openpyxl.load_workbook(EXCEL_OUT)

def find_or_create_sheet(name):
    if name in wb.sheetnames:
        ws = wb[name]
        # Clear old data except header
        for r in range(2, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                ws.cell(row=r, column=c).value = None
    else:
        ws = wb.create_sheet(name)
    # Header
    ws.cell(row=1, column=1, value="metric")
    ws.cell(row=1, column=2, value="unit")
    for i, yr in enumerate(TARGET_YEARS):
        ws.cell(row=1, column=i + 3, value=yr)
    return ws, {yr: i + 3 for i, yr in enumerate(TARGET_YEARS)}

def fill_sheet(ws, tgt_map, metrics, source_rows, source_map, divisor=1.0, row_offset=0):
    """Fill sheet with list of (label, unit, row_idx_in_source) tuples."""
    for idx, (label, unit, src_row_idx) in enumerate(metrics):
        r = idx + 2
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=unit)
        if src_row_idx is not None:
            src = source_rows[src_row_idx + row_offset] if row_offset else source_rows[src_row_idx]
            for yr in TARGET_YEARS:
                sc = source_map.get(yr)
                tc = tgt_map.get(yr)
                if sc is not None and tc is not None:
                    v = val_or_none(src, sc)
                    if v is not None:
                        v = v / divisor
                    from openpyxl.styles import Font
                    cell = ws.cell(row=r, column=tc)
                    if v is not None:
                        cell.value = round(v, 6)
                    cell.font = Font(color="0066CC")

# ── 1. Production by metal ──────────────────────────────────────────────

print("\n=== Production Data ===")
ws_prod, prod_tgt = find_or_create_sheet("production_metrics")

# Map Databook row indices (from earlier dump):
# Row 5: Nickel, tonnes
# Row 7: Copper, tonnes
# Row 9: Palladium, koz
# Row 11: Platinum, koz
# Row 6: Nickel from Russian feed
# Row 8: Copper from Russian feed
# Row 2: Trans-Baikal Cu concentrate
# Row 3: Trans-Baikal Iron ore concentrate

PROD_METRICS = [
    ("nickel_total_kt", "kt", 5, 1000),
    ("nickel_russian_feed_kt", "kt", 6, 1000),
    ("copper_total_kt", "kt", 7, 1000),
    ("copper_russian_feed_kt", "kt", 8, 1000),
    ("palladium_total_koz", "koz", 9, 1),
    ("palladium_russian_feed_koz", "koz", 10, 1),
    ("platinum_total_koz", "koz", 11, 1),
    ("platinum_russian_feed_koz", "koz", 12, 1),
    ("transbaikal_copper_concentrate_kt", "kt", 2, 1000),
    ("transbaikal_iron_ore_concentrate_kt", "kt", 3, 1),
]

for idx, (label, unit, src_row, divisor) in enumerate(PROD_METRICS):
    r = idx + 2
    ws_prod.cell(row=r, column=1, value=label)
    ws_prod.cell(row=r, column=2, value=unit)
    src = prod_rows[src_row]
    for yr in TARGET_YEARS:
        sc = prod_yr.get(yr)
        tc = prod_tgt.get(yr)
        if sc is not None and tc is not None:
            v = val_or_none(src, sc)
            if v is not None:
                v = v / divisor
            cell = ws_prod.cell(row=r, column=tc)
            if v is not None:
                cell.value = round(v, 6)
            from openpyxl.styles import Font
            cell.font = Font(color="0066CC")

# Add 2025 data from operational results PDF
DATA_2025 = {
    "nickel_total_kt": 198.521,
    "nickel_russian_feed_kt": 198.160,
    "copper_total_kt": 425.309,
    "copper_russian_feed_kt": 355.0,
    "palladium_total_koz": 2725,
    "platinum_total_koz": 667,
    "transbaikal_copper_concentrate_kt": 70.213,
}
tc_2025 = prod_tgt.get(2025)
if tc_2025:
    for idx, (label, unit, src_row, divisor) in enumerate(PROD_METRICS):
        if label in DATA_2025:
            r = idx + 2
            cell = ws_prod.cell(row=r, column=tc_2025)
            cell.value = DATA_2025[label]
            from openpyxl.styles import Font
            cell.font = Font(color="008800")

print(f"  ✓ production_metrics: {len(PROD_METRICS)} metrics")

# ── 2. Ore Output ───────────────────────────────────────────────────────

print("\n=== Ore Output ===")
ws_ore, ore_tgt = find_or_create_sheet("ore_output")

# Map key ore rows (from earlier dump):
ORE_METRICS = [
    # Total sulfide ore
    ("ore_sulfide_total_kt", "kt", 1),
    ("ore_rich_kt", "kt", 2),
    ("ore_cupreous_kt", "kt", 3),
    ("ore_disseminated_kt", "kt", 4),
    # Norilsk site
    ("ore_norilsk_site_total_kt", "kt", 8),
    ("ore_norilsk_rich_kt", "kt", 9),
    ("ore_norilsk_cupreous_kt", "kt", 10),
    ("ore_norilsk_disseminated_kt", "kt", 11),
    # Oktyabrsky mine
    ("ore_oktyabrsky_mine_kt", "kt", 13),
    # Taimyrsky mine (rich ore)
    ("ore_taimyrsky_mine_rich_kt", "kt", 17),
    # Komsomolsky mine
    ("ore_komsomolsky_mine_kt", "kt", 19),
    # Norilsk-1 (disseminated)
    ("ore_norilsk_1_disseminated_kt", "kt", 29),
    # Kola site
    ("ore_kola_site_total_kt", "kt", 31),
    # Severny underground
    ("ore_severny_underground_kt", "kt", 34),
]

for idx, (label, unit, src_row) in enumerate(ORE_METRICS):
    r = idx + 2
    ws_ore.cell(row=r, column=1, value=label)
    ws_ore.cell(row=r, column=2, value=unit)
    if src_row < len(ore_rows):
        src = ore_rows[src_row]
        for yr in TARGET_YEARS:
            sc = ore_yr.get(yr)
            tc = ore_tgt.get(yr)
            if sc is not None and tc is not None:
                v = val_or_none(src, sc)
                if v is not None:
                    v = v / 1000  # kt (already in kt from source)
                cell = ws_ore.cell(row=r, column=tc)
                if v is not None:
                    cell.value = round(v, 6)
                from openpyxl.styles import Font
                cell.font = Font(color="0066CC")

print(f"  ✓ ore_output: {len(ORE_METRICS)} metrics")

# ── 3. Recovery Rates ───────────────────────────────────────────────────

print("\n=== Recovery Rates ===")
ws_rec, rec_tgt = find_or_create_sheet("recovery_rates")

REC_METRICS = [
    # Concentration
    ("concentration_ni_norilsk_pct", "%", 3),
    ("concentration_ni_kola_pct", "%", 4),
    ("concentration_cu_norilsk_pct", "%", 6),
    ("concentration_cu_kola_pct", "%", 7),
    ("concentration_cu_transbaikal_pct", "%", 8),
    ("concentration_pgm_norilsk_pct", "%", 10),
    # Smelting
    ("smelting_ni_norilsk_pct", "%", 14),
    ("smelting_ni_kola_pct", "%", 15),
    ("smelting_ni_foreign_pct", "%", 16),
    ("smelting_cu_norilsk_pct", "%", 18),
    ("smelting_cu_kola_pct", "%", 19),
    ("smelting_cu_foreign_pct", "%", 20),
    ("smelting_pgm_norilsk_pct", "%", 22),
    ("smelting_pgm_kola_pct", "%", 23),
    ("smelting_pgm_foreign_pct", "%", 24),
]

for idx, (label, unit, src_row) in enumerate(REC_METRICS):
    r = idx + 2
    ws_rec.cell(row=r, column=1, value=label)
    ws_rec.cell(row=r, column=2, value=unit)
    if src_row < len(rec_rows):
        src = rec_rows[src_row]
        for yr in TARGET_YEARS:
            sc = rec_yr.get(yr)
            tc = rec_tgt.get(yr)
            if sc is not None and tc is not None:
                v = val_or_none(src, sc)
                cell = ws_rec.cell(row=r, column=tc)
                if v is not None:
                    cell.value = round(v, 4)
                from openpyxl.styles import Font
                cell.font = Font(color="0066CC")

print(f"  ✓ recovery_rates: {len(REC_METRICS)} metrics")

# ── 4. Average realized prices (Revenue / Volume) ────────────────────────

print("\n=== Average Realized Prices ===")
ws_price, price_tgt = find_or_create_sheet("realized_prices")

# Load IS data for revenue
wb2 = openpyxl.load_workbook(DATABOOK, data_only=True)
ws_is = wb2["INCOME STATEMENT"]
is_rows = []
for row in ws_is.iter_rows(values_only=True):
    is_rows.append([str(c) if c is not None else None for c in row])
is_hdr = is_rows[1]
is_yr_map = {}
for i, val in enumerate(is_hdr):
    try:
        y = int(float(val))
        if 2009 <= y <= 2026:
            is_yr_map[y] = i
    except:
        pass
wb2.close()

# Revenue rows: 2=Nickel, 3=Copper, 4=Palladium, 5=Platinum, 9=Other metals
# USD million
# Production rows: 5=Nickel (t), 7=Copper (t), 9=Palladium (koz), 11=Platinum (koz)

PRICE_METRICS = [
    ("nickel_realized_price_usd_per_t", "$/t", 2, 5, 1e6, 1),       # rev $M / prod t * 1e6
    ("copper_realized_price_usd_per_t", "$/t", 3, 7, 1e6, 1),
    ("palladium_realized_price_usd_per_oz", "$/oz", 4, 9, 1e6, 1000),  # rev $M / prod koz * 1e6 / 1000
    ("platinum_realized_price_usd_per_oz", "$/oz", 5, 11, 1e6, 1000),
]

for idx, (label, unit, rev_row, prod_row, rev_scale, prod_scale) in enumerate(PRICE_METRICS):
    r = idx + 2
    ws_price.cell(row=r, column=1, value=label)
    ws_price.cell(row=r, column=2, value=unit)
    for yr in TARGET_YEARS:
        sc_rev = is_yr_map.get(yr)
        sc_vol = prod_yr.get(yr)
        tc = price_tgt.get(yr)
        if sc_rev is not None and sc_vol is not None and tc is not None:
            rev = val_or_none(is_rows[rev_row], sc_rev)  # $M
            vol = val_or_none(prod_rows[prod_row], sc_vol)  # t or koz
            if rev and vol and vol != 0:
                price = (rev * rev_scale) / (vol * prod_scale)
                cell = ws_price.cell(row=r, column=tc)
                cell.value = round(price, 2)
                from openpyxl.styles import Font
                cell.font = Font(color="0066CC")

print(f"  ✓ realized_prices: {len(PRICE_METRICS)} metrics")

# ── 5. Mineral Reserves (static, as of 01.01.2026) ──────────────────────

print("\n=== Mineral Reserves ===")
ws_res, _ = find_or_create_sheet("mineral_reserves")

res_rows = load_databook_sheet("MINERAL RESOURCES_ORE RESERVES")

# Key rows: 5=P&P total, 6=M&I total, 7=Inferred total, 9=Taimyr P&P
RESERVE_DATA = [
    ("proven_probable_reserves_ore_mln_t", "mln t", 5, 0),
    ("proven_probable_reserves_ni_pct", "%", 5, 2),
    ("proven_probable_reserves_cu_pct", "%", 5, 3),
    ("proven_probable_reserves_pd_gt", "g/t", 5, 4),
    ("proven_probable_reserves_pt_gt", "g/t", 5, 5),
    ("proven_probable_reserves_contained_ni_kt", "kt", 5, 8),
    ("proven_probable_reserves_contained_cu_kt", "kt", 5, 9),
]

for idx, (label, unit, src_row, src_col) in enumerate(RESERVE_DATA):
    r = idx + 2
    ws_res.cell(row=r, column=1, value=label)
    ws_res.cell(row=r, column=2, value=unit)
    if src_row < len(res_rows):
        v = val_or_none(res_rows[src_row], src_col)
        if v is not None:
            ws_res.cell(row=r, column=3, value=round(v, 4))
            from openpyxl.styles import Font
            ws_res.cell(row=r, column=3).font = Font(color="0066CC")

print(f"  ✓ mineral_reserves: {len(RESERVE_DATA)} metrics")

# ── 6. Update operational_drivers with better data ──────────────────────

print("\n=== Operational Drivers ===")
if "operational_drivers" not in wb.sheetnames:
    ws_drv = wb.create_sheet("operational_drivers")
else:
    ws_drv = wb["operational_drivers"]
    # Clear old
    for r in range(2, ws_drv.max_row + 1):
        for c in range(1, ws_drv.max_column + 1):
            ws_drv.cell(row=r, column=c).value = None

ws_drv.cell(row=1, column=1, value="driver")
ws_drv.cell(row=1, column=2, value="unit")
for i, yr in enumerate(TARGET_YEARS):
    ws_drv.cell(row=1, column=i + 3, value=yr)
drv_tgt = {yr: i + 3 for i, yr in enumerate(TARGET_YEARS)}

DRIVERS = [
    # Production
    ("nickel_production_kt", "kt", prod_rows, prod_yr, 5, 1000),
    ("copper_production_kt", "kt", prod_rows, prod_yr, 7, 1000),
    ("palladium_production_koz", "koz", prod_rows, prod_yr, 9, 1),
    ("platinum_production_koz", "koz", prod_rows, prod_yr, 11, 1),
    # Ore
    ("total_ore_mined_mt", "mln t", ore_rows, ore_yr, 1, 1000),
    ("ore_rich_mt", "mln t", ore_rows, ore_yr, 2, 1000),
    ("ore_disseminated_mt", "mln t", ore_rows, ore_yr, 4, 1000),
    # Revenue (from IS)
    ("revenue_nickel_m_usd", "$M", is_rows, is_yr_map, 2, 1),
    ("revenue_copper_m_usd", "$M", is_rows, is_yr_map, 3, 1),
    ("revenue_palladium_m_usd", "$M", is_rows, is_yr_map, 4, 1),
    ("revenue_platinum_m_usd", "$M", is_rows, is_yr_map, 5, 1),
    ("revenue_other_metals_m_usd", "$M", is_rows, is_yr_map, 9, 1),
    ("revenue_total_metal_sales_m_usd", "$M", is_rows, is_yr_map, 10, 1),
]

for idx, (label, unit, src_rows, src_map, src_row_idx, divisor) in enumerate(DRIVERS):
    r = idx + 2
    ws_drv.cell(row=r, column=1, value=label)
    ws_drv.cell(row=r, column=2, value=unit)
    if src_row_idx < len(src_rows):
        src = src_rows[src_row_idx]
        for yr in TARGET_YEARS:
            sc = src_map.get(yr)
            tc = drv_tgt.get(yr)
            if sc is not None and tc is not None:
                v = val_or_none(src, sc)
                if v is not None:
                    v = v / divisor
                cell = ws_drv.cell(row=r, column=tc)
                if v is not None:
                    cell.value = round(v, 6)
                from openpyxl.styles import Font
                cell.font = Font(color="0066CC")

# Add 2025 production data
if 2025 in drv_tgt:
    tc = drv_tgt[2025]
    for idx, (label, unit, src_rows, src_map, src_row_idx, divisor) in enumerate(DRIVERS):
        if label in DATA_2025:
            r = idx + 2
            ws_drv.cell(row=r, column=tc).value = DATA_2025[label]
            from openpyxl.styles import Font
            ws_drv.cell(row=r, column=tc).font = Font(color="008800")

print(f"  ✓ operational_drivers: {len(DRIVERS)} drivers")

# ── Save ─────────────────────────────────────────────────────────────────

print(f"\nSaving to {EXCEL_OUT}...")
wb.save(EXCEL_OUT)
print("✅ All operational data parsed and saved!")
print("\nNew sheets added/updated:")
print("  - production_metrics (10 metals metrics)")
print("  - ore_output (15 ore metrics by site/mine)")
print("  - recovery_rates (15 concentration + smelting rates)")
print("  - realized_prices (4 metal prices $/t or $/oz)")
print("  - mineral_reserves (7 reserve metrics)")
print("  - operational_drivers (13 key drivers)")
