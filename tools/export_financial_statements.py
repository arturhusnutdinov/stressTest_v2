#!/usr/bin/env python3
"""
tools/export_financial_statements.py

Экспорт финансовых форм (IS + BS + CF) в формате раскрытия:
история (history_is/bs/cf) + прогноз (forecast_is/bs/cf) в одной таблице.

Формат совместим с листом "Financial Statements" из шаблона Rusal.
Каждая статья смаплена через dual-mapping (history_metric ≠ forecast_metric)
с явной обработкой знаков для обоих источников.

Usage:
    python3 tools/export_financial_statements.py --company rusal
    python3 tools/export_financial_statements.py --company rusal --scenario base --out financials.xlsx
"""

from __future__ import annotations

import argparse
import sqlite3
import sys
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Union

try:
    import openpyxl
    from openpyxl.styles import (Alignment, Border, Font, PatternFill, Side)
    from openpyxl.utils import get_column_letter
except ImportError:
    print("pip install openpyxl")
    sys.exit(1)

ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT))
DB_PATH = ROOT / "data_mart_v2.db"

# ─── row types ────────────────────────────────────────────────────────────────
SPACER   = "spacer"
HEADER   = "header"
DATA     = "data"
SUBTOTAL = "subtotal"
TOTAL    = "total"
DIVIDER  = "divider"
CHECK    = "check"

# Metric ref: str → single metric; list of (str, float) → weighted sum; None → skip
MetricRef = Union[str, List[Tuple[str, float]], None]


class Row:
    """Строка схемы финансовых форм."""
    __slots__ = ("label", "row_type", "h_stmt", "h_metric", "h_sign",
                 "f_stmt", "f_metric", "f_sign", "indent")

    def __init__(
        self,
        label:    str,
        row_type: str       = DATA,
        h_stmt:   str       = "IS",
        h_metric: MetricRef = None,
        h_sign:   float     = 1.0,
        f_stmt:   str       = "IS",
        f_metric: MetricRef = None,
        f_sign:   float     = 1.0,
        indent:   int       = 0,
    ):
        self.label    = label
        self.row_type = row_type
        self.h_stmt   = h_stmt
        self.h_metric = h_metric
        self.h_sign   = h_sign
        self.f_stmt   = f_stmt
        self.f_metric = f_metric
        self.f_sign   = f_sign
        self.indent   = indent


def _IS(label, rt=DATA, h=None, f=None, hs=1.0, fs=1.0, ind=0):
    return Row(label, rt, "IS", h, hs, "IS", f, fs, ind)

def _BS(label, rt=DATA, h=None, f=None, hs=1.0, fs=1.0, ind=0):
    return Row(label, rt, "BS", h, hs, "BS", f, fs, ind)

def _CF(label, rt=DATA, h=None, f=None, hs=1.0, fs=1.0, ind=0):
    return Row(label, rt, "CF", h, hs, "CF", f, fs, ind)


# ═══════════════════════════════════════════════════════════════
# INCOME STATEMENT
# history_is metric names  ↔  forecast_is metric names
# ═══════════════════════════════════════════════════════════════
IS_SCHEMA: List[Row] = [
    _IS("INCOME STATEMENT",                                  HEADER),
    _IS("All figures in USD millions",                       DIVIDER),
    _IS("", SPACER),

    _IS("Revenue",                          h="revenue",                    f="revenue",                 hs=1, fs=1,  ind=0),
    _IS("Cost of sales",                    h="cogs",                       f="cogs",                    hs=1, fs=1,  ind=0),
    _IS("Gross profit",            SUBTOTAL, h="gross_profit",              f="gross_profit",             hs=1, fs=1,  ind=0),
    _IS("", SPACER),

    # Opex — в history: sga = комбинированное (admin+ECL+other для 2011-2023, admin с 2024)
    # admin_implied = ebit - gross_profit - distribution - impairment - ecl - other_opex
    _IS("Distribution expenses",            h="distribution_expenses",      f="distribution_expenses",   hs=1, fs=1,  ind=1),
    _IS("Administrative expenses",
        h=[("ebit",1),("gross_profit",-1),("distribution_expenses",-1),
           ("asset_impairment",-1),("expected_credit_losses",-1),("other_operating_expenses",-1)],
        f="administrative_expenses",                                                                     hs=1, fs=1,  ind=1),
    _IS("Impairment of non-current assets", h="asset_impairment",           f="asset_impairment_charges",hs=1, fs=1,  ind=1),
    _IS("Expected credit losses",           h="expected_credit_losses",     f="expected_credit_losses",  hs=1, fs=1,  ind=1),
    _IS("Net other operating expenses",     h="other_operating_expenses",   f="other_operating_expenses",hs=1, fs=1,  ind=1),
    _IS("Results from operating activities (EBIT)", TOTAL,
                                            h="ebit",                       f="ebit",                    hs=1, fs=1,  ind=0),
    _IS("", SPACER),

    # Below EBIT
    # interest_expense: в history < 0 (expense), в forecast > 0 (absolute) → fs=-1
    _IS("Finance income",                   h="interest_income",            f="interest_income",         hs=1, fs=1,  ind=1),
    _IS("Finance expenses",                 h="interest_expense",           f="interest_expense",        hs=1, fs=-1, ind=1),
    _IS("Share of profits of associates",   h="earnings_from_investees",    f="earnings_from_investees", hs=1, fs=1,  ind=1),
    _IS("", SPACER),
    _IS("Profit before taxation (EBT)",     TOTAL,
                                            h="ebt",                        f="ebt",                     hs=1, fs=1,  ind=0),
    _IS("", SPACER),

    _IS("Current income tax expense",       h="current_tax",                f="current_tax_expense",     hs=1, fs=1,  ind=1),
    _IS("Deferred income tax credit/(charge)", h="deferred_tax",            f="deferred_tax_expense",    hs=1, fs=1,  ind=1),
    _IS("Income tax",                       SUBTOTAL,
                                            h="tax_expense",                f="tax_expense",             hs=1, fs=1,  ind=0),
    _IS("", SPACER),
    _IS("Profit for the year",              TOTAL,
                                            h="net_income",                 f="net_income",              hs=1, fs=1,  ind=0),
    _IS("", SPACER),

    _IS("─── KPIs ───",                     DIVIDER),
    _IS("EBITDA",                           h="ebitda",                     f="ebitda",                  hs=1, fs=1,  ind=0),
    _IS("D&A",                              h="total_da",                   f="total_da",                hs=1, fs=1,  ind=1),
]


# ═══════════════════════════════════════════════════════════════
# BALANCE SHEET
# BS sign issues:
#   accounts_payable:     history=+, forecast=- → fs=-1
#   taxes_payable:        history=-, forecast=0 → hs=-1
#   employee_benefits:    history=-, forecast=+ → hs=-1
#   lease_liab_*:         history=-, forecast=+ → hs=-1
# ═══════════════════════════════════════════════════════════════
BS_SCHEMA: List[Row] = [
    _BS("BALANCE SHEET",                                     HEADER),
    _BS("All figures in USD millions",                       DIVIDER),
    _BS("", SPACER),

    # Non-current assets
    _BS("ASSETS",                                            HEADER),
    _BS("Non-current assets",                                HEADER, ind=0),
    _BS("Property, plant and equipment, net",   h="ppe_net",            f="ppe_net",                          hs=1, fs=1, ind=1),
    _BS("Intangible assets",                    h="intangibles",        f="intangibles",                      hs=1, fs=1, ind=1),
    _BS("Goodwill",                             h="goodwill",           f="goodwill",                         hs=1, fs=1, ind=1),
    _BS("Interest in associates & JVs",
        h="investments_lt",                                             f="investments_and_long_term_receivables", hs=1, fs=1, ind=1),
    _BS("Deferred tax assets",                  h="dta",                f="dta",                              hs=1, fs=1, ind=1),
    _BS("ROU asset",                            h="rou_asset",          f="rou_asset",                        hs=1, fs=1, ind=1),
    _BS("Other non-current assets",             h="other_nca",          f="other_non_current_assets",         hs=1, fs=1, ind=1),
    _BS("Total non-current assets",             SUBTOTAL,
        h=[("ppe_net",1),("intangibles",1),("goodwill",1),("investments_lt",1),
           ("dta",1),("rou_asset",1),("other_nca",1)],
        f="total_non_current_assets",                                                                          hs=1, fs=1, ind=0),
    _BS("", SPACER),

    # Current assets
    _BS("Current assets",                                    HEADER, ind=0),
    _BS("Inventories",                          h="inventory",          f="inventory",                        hs=1, fs=1, ind=1),
    _BS("Trade and other receivables",          h="accounts_receivable",f="accounts_receivable",              hs=1, fs=1, ind=1),
    _BS("Cash and cash equivalents",            h="cash",               f="cash",                             hs=1, fs=1, ind=1),
    _BS("Other current assets",                 h="other_ca",           f="other_current_assets",             hs=1, fs=1, ind=1),
    _BS("Total current assets",                 SUBTOTAL,
        h=[("inventory",1),("accounts_receivable",1),("cash",1),("other_ca",1)],
        f="total_current_assets",                                                                              hs=1, fs=1, ind=0),
    _BS("", SPACER),
    _BS("Total assets",                         TOTAL,  h="total_assets", f="total_assets",                   hs=1, fs=1, ind=0),
    _BS("", SPACER),

    # Equity
    _BS("EQUITY AND LIABILITIES",                            HEADER),
    _BS("Equity",                                            HEADER, ind=0),
    _BS("Share capital",                        h="share_capital",      f="share_capital",                    hs=1, fs=1, ind=1),
    _BS("Share premium (APIC)",                 h="apic",               f="apic",                             hs=1, fs=1, ind=1),
    _BS("Retained earnings",                    h="retained_earnings",  f="retained_earnings",                hs=1, fs=1, ind=1),
    _BS("Other reserves / AOCI",                h="aoci",               f="aoci",                             hs=1, fs=1, ind=1),
    _BS("Non-controlling interest",             h="nci",                f="nci",                              hs=1, fs=1, ind=1),
    _BS("Total equity",                         TOTAL,  h="total_equity", f="total_equity",                   hs=1, fs=1, ind=0),
    _BS("", SPACER),

    # Non-current liabilities
    _BS("Non-current liabilities",                           HEADER, ind=0),
    _BS("Loans and borrowings (LT)",            h="long_term_debt",     f="long_term_debt",                   hs=1,  fs=1, ind=1),
    _BS("Lease liabilities (LT)",               h="lease_liab_noncurrent", f="lease_liab_noncurrent",         hs=-1, fs=1, ind=1),
    _BS("Deferred tax liabilities",             h="dtl",                f="dtl",                              hs=1,  fs=1, ind=1),
    # employee_benefits: history stores as negative → hs=-1
    _BS("Employee benefits",                    h="employee_benefits",  f="employee_benefits",                hs=-1, fs=1, ind=1),
    _BS("Other non-current liabilities",        h="other_ncl",          f="other_non_current_liabilities",    hs=1,  fs=1, ind=1),
    _BS("Total non-current liabilities",        SUBTOTAL,
        h=[("long_term_debt",1),("lease_liab_noncurrent",-1),
           ("dtl",1),("employee_benefits",-1),("other_ncl",1)],
        f="total_non_current_liabilities",                                                                     hs=1,  fs=1, ind=0),
    _BS("", SPACER),

    # Current liabilities
    _BS("Current liabilities",                               HEADER, ind=0),
    _BS("Loans and borrowings (ST)",            h="short_term_debt",    f="short_term_debt",                  hs=1,  fs=1, ind=1),
    _BS("Lease liabilities (current)",          h="lease_liab_current", f="lease_liab_current",               hs=-1, fs=1, ind=1),
    # accounts_payable: history=positive, forecast=negative → fs=-1
    _BS("Trade and other payables",             h="accounts_payable",   f="accounts_payable",                 hs=1,  fs=-1, ind=1),
    # taxes_payable: history=negative → hs=-1
    _BS("Taxes payable",                        h="taxes_payable",      f="taxes_payable",                    hs=-1, fs=1, ind=1),
    _BS("Other current liabilities",            h="other_cl",           f="other_current_liabilities",        hs=1,  fs=1, ind=1),
    _BS("Total current liabilities",            SUBTOTAL,
        h=[("short_term_debt",1),("lease_liab_current",-1),
           ("accounts_payable",1),("taxes_payable",-1),("other_cl",1)],
        f="total_current_liabilities",                                                                         hs=1,  fs=1, ind=0),
    _BS("", SPACER),
    _BS("Total liabilities",                    SUBTOTAL,
        h=[("long_term_debt",1),("lease_liab_noncurrent",-1),("dtl",1),("employee_benefits",-1),("other_ncl",1),
           ("short_term_debt",1),("lease_liab_current",-1),("accounts_payable",1),("taxes_payable",-1),("other_cl",1)],
        f="total_liabilities",                                                                                 hs=1,  fs=1, ind=0),
    _BS("Total equity and liabilities",         TOTAL,  h="total_liab_equity", f="total_liab_equity",         hs=1,  fs=1, ind=0),
    _BS("", SPACER),
    # BS check: total_assets - total_equity - total_liabilities (should be ~0)
    _BS("BS check (A – E – L)",                 CHECK,
        h=[("total_assets",1),("total_equity",-1),
           ("long_term_debt",-1),("lease_liab_noncurrent",1),("dtl",-1),("employee_benefits",1),("other_ncl",-1),
           ("short_term_debt",-1),("lease_liab_current",1),("accounts_payable",-1),("taxes_payable",1),("other_cl",-1)],
        f=[("total_assets",1),("total_equity",-1),("total_liabilities",-1)],                                  hs=1,  fs=1, ind=0),
]


# ═══════════════════════════════════════════════════════════════
# CASH FLOW STATEMENT
# history_cf metric names  ↔  forecast_cf metric names
# ═══════════════════════════════════════════════════════════════
CF_SCHEMA: List[Row] = [
    _CF("CASH FLOW STATEMENT",                               HEADER),
    _CF("All figures in USD millions (indirect method)",     DIVIDER),
    _CF("", SPACER),

    _CF("Operating activities",                              HEADER),
    _CF("Profit for the year",              h="net_income",                 f="net_income",              hs=1, fs=1, ind=1),
    _CF("Adjustments for non-cash items:",                   DIVIDER, ind=1),
    _CF("Depreciation & amortisation",      h="cfo_da",                     f="total_da",                hs=1, fs=1, ind=2),
    _CF("Impairment (non-cash)",            h="impairment_noncash",         f="impairment_noncash",      hs=1, fs=1, ind=2),
    _CF("Share of profits of associates",   h="share_associates_noncash",   f="associates_noncash",      hs=1, fs=1, ind=2),
    _CF("FX loss/(gain), non-cash",         h="fx_noncash",                 f="fx_noncash",              hs=1, fs=1, ind=2),
    _CF("Interest expense (non-cash)",      h="interest_noncash",           f=None,                      hs=1, fs=1, ind=2),
    _CF("Interest income (non-cash)",       h="interest_income_noncash",    f=None,                      hs=1, fs=1, ind=2),
    _CF("Deferred income tax",              h="deferred_income_taxes",      f="deferred_income_taxes",   hs=1, fs=1, ind=2),
    _CF("Changes in working capital:",                       DIVIDER, ind=1),
    _CF("Inventories",                      h="wc_inventory",               f="wc_inventory_change",     hs=1, fs=1, ind=2),
    _CF("Trade receivables",                h="wc_receivables",             f="wc_accounts_receivable_change", hs=1, fs=1, ind=2),
    _CF("Trade payables",                   h="wc_payables",                f="wc_accounts_payable_change",    hs=1, fs=1, ind=2),
    _CF("Other WC changes",                 h="wc_provisions",              f="change_other_wc",         hs=1, fs=1, ind=2),
    _CF("Interest paid",                    h="interest_paid",              f="interest_paid",           hs=1, fs=1, ind=1),
    _CF("Income tax paid",                  h="taxes_paid",                 f="taxes_paid",              hs=1, fs=1, ind=1),
    _CF("Net cash from operating activities", TOTAL,
                                            h="cfo_total",                  f="cfo_total",               hs=1, fs=1, ind=0),
    _CF("", SPACER),

    _CF("Investing activities",                              HEADER),
    _CF("Acquisition of PPE (capex)",       h="capex",                      f="capex",                   hs=1, fs=1, ind=1),
    _CF("Acquisition of intangibles",       h="capex_intangibles",          f=None,                      hs=1, fs=1, ind=1),
    _CF("Proceeds from disposal of PPE",    h="proceeds_ppe_disposal",      f="disposal_proceeds",       hs=1, fs=1, ind=1),
    _CF("Dividends from associates & JVs",  h="dividends_from_associates",  f=None,                      hs=1, fs=1, ind=1),
    _CF("Acquisitions of businesses",       h="acquisitions",               f="acquisitions",            hs=1, fs=1, ind=1),
    _CF("Proceeds from disposal of associates", h=None,                     f="associates_disposal_proceeds", hs=1, fs=1, ind=1),
    _CF("Net cash from investing activities", TOTAL,
                                            h="cfi_total",                  f="cfi_total",               hs=1, fs=1, ind=0),
    _CF("", SPACER),

    _CF("Financing activities",                              HEADER),
    _CF("Proceeds from borrowings",         h="proceeds_borrowings",        f="debt_issuance",           hs=1, fs=1, ind=1),
    _CF("Repayment of borrowings",          h="repayments_borrowings",      f="debt_repayments",         hs=1, fs=1, ind=1),
    _CF("Finance lease payments",           h="fin_lease_principal_cff",    f="fin_lease_principal_cff", hs=1, fs=1, ind=1),
    _CF("Dividends paid",                   h="cff_dividends",              f="dividends_paid",          hs=1, fs=1, ind=1),
    _CF("Net cash from financing activities", TOTAL,
                                            h="cff_total",                  f="cff_total",               hs=1, fs=1, ind=0),
    _CF("", SPACER),

    _CF("Net change in cash",               SUBTOTAL,
                                            h=[("cfo_total",1),("cfi_total",1),("cff_total",1)],
                                            f="net_change",                                               hs=1, fs=1, ind=0),
    _CF("Effect of exchange rates on cash", h="fx_effect_cash",             f="cf_fx_effect",            hs=1, fs=1, ind=1),
    _CF("Cash at beginning of year",        h="cash_opening",               f="cash_opening",            hs=1, fs=1, ind=0),
    _CF("Cash at end of year",              TOTAL,
                                            h="cash_ending",                f="cash_ending",             hs=1, fs=1, ind=0),
]


# ═══════════════════════════════════════════════════════════════
# DATA LAYER
# ═══════════════════════════════════════════════════════════════

def load_history(conn: sqlite3.Connection, company_id: str) -> Dict[str, Dict[int, Dict[str, float]]]:
    """
    Загружает историю IS/BS/CF.
    Возвращает {stmt: {year: {metric: value}}}
    """
    result: Dict[str, Dict[int, Dict[str, float]]] = {"IS": {}, "BS": {}, "CF": {}}
    table_map = {"IS": "history_is", "BS": "history_bs", "CF": "history_cf"}

    for stmt, table in table_map.items():
        rows = conn.execute(
            f"SELECT p.year, h.metric, h.value "
            f"FROM {table} h "
            f"JOIN periods p ON h.period_id = p.period_id "
            f"WHERE h.company_id = ? AND p.is_forecast = 0 "
            f"ORDER BY p.year",
            (company_id,),
        ).fetchall()
        for year, metric, value in rows:
            if year not in result[stmt]:
                result[stmt][year] = {}
            result[stmt][year][metric] = value

    return result


def load_forecast(
    conn: sqlite3.Connection,
    company_id: str,
    scenario_id: int,
) -> Dict[str, Dict[int, Dict[str, float]]]:
    """
    Загружает прогноз IS/BS/CF для сценария.
    Возвращает {stmt: {year: {metric: value}}}
    """
    result: Dict[str, Dict[int, Dict[str, float]]] = {"IS": {}, "BS": {}, "CF": {}}
    table_map = {"IS": "forecast_is", "BS": "forecast_bs", "CF": "forecast_cf"}

    for stmt, table in table_map.items():
        rows = conn.execute(
            f"SELECT p.year, f.metric, f.value "
            f"FROM {table} f "
            f"JOIN periods p ON f.period_id = p.period_id "
            f"WHERE f.company_id = ? AND f.scenario_id = ? "
            f"ORDER BY p.year",
            (company_id, scenario_id),
        ).fetchall()
        for year, metric, value in rows:
            if year not in result[stmt]:
                result[stmt][year] = {}
            result[stmt][year][metric] = value

    return result


def resolve_value(
    data: Dict[int, Dict[str, float]],
    year: int,
    metric: MetricRef,
    sign: float,
    display_unit: float,
) -> Optional[float]:
    """
    Получить значение для года из данных одного стейтмента.
    metric: str → прямой поиск; list[(str,float)] → взвешенная сумма.
    Возвращает None если данных нет.
    """
    if metric is None:
        return None
    year_data = data.get(year, {})
    if not year_data:
        return None

    if isinstance(metric, str):
        v = year_data.get(metric)
        if v is None:
            return None
        return v * sign / display_unit
    else:  # list of (metric_name, coeff)
        total = 0.0
        found_any = False
        for m, coeff in metric:
            v = year_data.get(m, 0.0)
            if v is not None:
                total += v * coeff
                found_any = True
        return (total * sign / display_unit) if found_any else None


# ═══════════════════════════════════════════════════════════════
# EXCEL FORMATTING
# ═══════════════════════════════════════════════════════════════

FONT_NAME = "Calibri"

# Цвета
C_HIST_HDR  = "D6E4F0"   # синеватый — шапка истории
C_FCAST_HDR = "E2EFDA"   # зеленоватый — шапка прогноза
C_HIST_ALT  = "EFF5FB"   # светло-синий — ячейки истории
C_FCAST_ALT = "F2F8EE"   # светло-зеленый — ячейки прогноза
C_HEADER    = "2F5496"   # тёмно-синий — строки-заголовки
C_SUBHEADER = "8EA9C1"   # средний синий — подзаголовки
C_SUBTOTAL  = "BDD7EE"   # светло-синий — подытоги
C_TOTAL     = "2F5496"   # тёмно-синий — итоги
C_CHECK_OK  = "C6EFCE"   # зелёный — ок
C_CHECK_ERR = "FFC7CE"   # красный — ошибка
C_DIVIDER   = "F5F5F5"   # очень светлый серый
C_WHITE     = "FFFFFF"

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, italic=False, size=10, color="000000", name=FONT_NAME) -> Font:
    return Font(name=name, bold=bold, italic=italic, size=size, color=color)

def _border_top() -> Border:
    s = Side(style="thin", color="2F5496")
    return Border(top=s)

def _border_bottom() -> Border:
    s = Side(style="thin", color="2F5496")
    return Border(bottom=s)

def _border_tb() -> Border:
    s = Side(style="thin", color="2F5496")
    return Border(top=s, bottom=s)

def _border_double_bottom() -> Border:
    s  = Side(style="thin",   color="2F5496")
    d  = Side(style="double", color="2F5496")
    return Border(top=s, bottom=d)

def _align(halign="left", valign="center", wrap=False) -> Alignment:
    return Alignment(horizontal=halign, vertical=valign, wrap_text=wrap)


def _num_format(v: Optional[float]) -> str:
    """Числовой формат."""
    return '#,##0.0;[Red]-#,##0.0;"-"'


class SheetBuilder:
    """Строит один лист Excel."""

    LABEL_COL    = 2       # B
    FIRST_DATA_COL = 3     # C
    ROW_HEIGHT_DEFAULT = 15
    ROW_HEIGHT_SPACER  = 6
    ROW_HEIGHT_HEADER  = 18
    INDENT_PX = 8          # пикселей на уровень отступа

    def __init__(self, ws, hist_years: List[int], fcast_years: List[int], display_unit: float):
        self.ws = ws
        self.hist_years  = hist_years
        self.fcast_years = fcast_years
        self.all_years   = hist_years + fcast_years
        self.display_unit = display_unit
        self._row = 1

        # настроить ширину столбцов
        ws.column_dimensions[get_column_letter(1)].width = 3   # A — indent marker
        ws.column_dimensions[get_column_letter(self.LABEL_COL)].width = 46
        for col_idx in range(self.FIRST_DATA_COL, self.FIRST_DATA_COL + len(self.all_years)):
            ws.column_dimensions[get_column_letter(col_idx)].width = 12

    def _col(self, year_idx: int) -> int:
        return self.FIRST_DATA_COL + year_idx

    def write_year_header(self) -> None:
        ws = self.ws
        r = self._row

        # Строка с категориями «History» / «Forecast»
        hist_start = self._col(0)
        hist_end   = self._col(len(self.hist_years) - 1)
        fcast_start = self._col(len(self.hist_years))
        fcast_end   = self._col(len(self.all_years) - 1)

        if len(self.hist_years) > 1:
            ws.merge_cells(
                start_row=r, start_column=hist_start,
                end_row=r, end_column=hist_end
            )
        c = ws.cell(r, hist_start, "History (actual)")
        c.font    = _font(bold=True, color="FFFFFF", size=10)
        c.fill    = _fill(C_HEADER)
        c.alignment = _align("center")
        for col in range(hist_start, hist_end + 1):
            ws.cell(r, col).fill = _fill(C_HEADER)

        if len(self.fcast_years) > 1:
            ws.merge_cells(
                start_row=r, start_column=fcast_start,
                end_row=r, end_column=fcast_end
            )
        c = ws.cell(r, fcast_start, "Forecast")
        c.font    = _font(bold=True, color="FFFFFF", size=10)
        c.fill    = _fill("375623")
        c.alignment = _align("center")
        for col in range(fcast_start, fcast_end + 1):
            ws.cell(r, col).fill = _fill("375623")

        ws.row_dimensions[r].height = 16
        self._row += 1

        # Строка с годами
        r = self._row
        ws.cell(r, self.LABEL_COL, "USD millions").font = _font(italic=True, size=9)
        for i, year in enumerate(self.all_years):
            col = self._col(i)
            c   = ws.cell(r, col, year)
            c.font      = _font(bold=True, size=10, color="FFFFFF",
                                name=FONT_NAME)
            c.alignment = _align("center")
            c.fill      = _fill(C_HIST_HDR if year in self.hist_years else "4E9A06"[:6])
            if year in self.hist_years:
                c.fill = _fill(C_HEADER)
            else:
                c.fill = _fill("375623")
        ws.row_dimensions[r].height = 16
        self._row += 1

    def write_row(
        self,
        schema_row: Row,
        hist_data:  Dict[str, Dict[int, Dict[str, float]]],
        fcast_data: Dict[str, Dict[int, Dict[str, float]]],
    ) -> None:
        rt = schema_row.row_type
        ws = self.ws
        r  = self._row

        if rt == SPACER:
            ws.row_dimensions[r].height = self.ROW_HEIGHT_SPACER
            self._row += 1
            return

        if rt == DIVIDER:
            c = ws.cell(r, self.LABEL_COL, schema_row.label)
            c.font = _font(italic=True, size=8, color="666666")
            c.fill = _fill(C_DIVIDER)
            for i in range(len(self.all_years)):
                ws.cell(r, self._col(i)).fill = _fill(C_DIVIDER)
            ws.row_dimensions[r].height = 12
            self._row += 1
            return

        if rt == HEADER:
            level = schema_row.indent
            c = ws.cell(r, self.LABEL_COL, ("    " * level) + schema_row.label)
            if level == 0:
                c.font  = _font(bold=True, size=10, color="FFFFFF")
                c.fill  = _fill(C_HEADER)
                for i in range(len(self.all_years)):
                    col = self._col(i)
                    ws.cell(r, col).fill = _fill(C_HEADER)
                ws.row_dimensions[r].height = self.ROW_HEIGHT_HEADER
            else:
                c.font  = _font(bold=True, size=9, color="FFFFFF")
                c.fill  = _fill(C_SUBHEADER)
                for i in range(len(self.all_years)):
                    ws.cell(r, self._col(i)).fill = _fill(C_SUBHEADER)
                ws.row_dimensions[r].height = 14
            c.alignment = _align()
            self._row += 1
            return

        # DATA / SUBTOTAL / TOTAL / CHECK
        indent_str = "   " * schema_row.indent
        label_cell = ws.cell(r, self.LABEL_COL, indent_str + schema_row.label)

        if rt == TOTAL:
            label_cell.font   = _font(bold=True, size=10)
            label_cell.border = _border_double_bottom()
            row_fill_h = C_SUBTOTAL
            row_fill_f = "C6E0B4"
            ws.row_dimensions[r].height = 16
        elif rt == SUBTOTAL:
            label_cell.font   = _font(bold=True, size=9)
            label_cell.border = _border_tb()
            row_fill_h = "DDEBF7"
            row_fill_f = "E2EFDA"
            ws.row_dimensions[r].height = 14
        elif rt == CHECK:
            label_cell.font   = _font(italic=True, size=9, color="444444")
            row_fill_h = C_DIVIDER
            row_fill_f = C_DIVIDER
            ws.row_dimensions[r].height = 13
        else:  # DATA
            label_cell.font = _font(size=10)
            row_fill_h = C_HIST_ALT
            row_fill_f = C_FCAST_ALT
            ws.row_dimensions[r].height = self.ROW_HEIGHT_DEFAULT

        label_cell.alignment = _align()

        for i, year in enumerate(self.all_years):
            col      = self._col(i)
            is_hist  = year in self.hist_years
            src_data = hist_data if is_hist else fcast_data
            metric   = schema_row.h_metric if is_hist else schema_row.f_metric
            sign     = schema_row.h_sign   if is_hist else schema_row.f_sign
            stmt     = schema_row.h_stmt   if is_hist else schema_row.f_stmt

            cell = ws.cell(r, col)
            cell.alignment = _align("right")
            cell.fill = _fill(row_fill_h if is_hist else row_fill_f)

            val = resolve_value(src_data.get(stmt, {}), year, metric, sign, self.display_unit)

            if val is None:
                cell.value = None
                if rt in (TOTAL, SUBTOTAL):
                    cell.border = _border_double_bottom() if rt == TOTAL else _border_tb()
            else:
                cell.value = round(val, 1)
                cell.number_format = '#,##0.0;[Red](-#,##0.0);"-"'
                if rt in (TOTAL, SUBTOTAL):
                    cell.font   = _font(bold=True, size=10 if rt == TOTAL else 9)
                    cell.border = _border_double_bottom() if rt == TOTAL else _border_tb()
                elif rt == CHECK:
                    # зелёный если ≈0, красный иначе
                    threshold = 1.0  # $1M threshold
                    cell.fill = _fill(C_CHECK_OK if abs(val) < threshold else C_CHECK_ERR)
                    cell.font = _font(size=9)

        self._row += 1

    def write_section(
        self,
        schema: List[Row],
        hist_data: Dict[str, Dict[int, Dict[str, float]]],
        fcast_data: Dict[str, Dict[int, Dict[str, float]]],
    ) -> None:
        for row in schema:
            self.write_row(row, hist_data, fcast_data)
        # пустая строка после секции
        self._row += 1


# ═══════════════════════════════════════════════════════════════
# MAIN EXPORT
# ═══════════════════════════════════════════════════════════════

def export_financial_statements(
    company_id: str,
    scenario_name: str = "base",
    out_path: Optional[Path] = None,
    db_path: Path = DB_PATH,
    display_unit: float = 1_000_000.0,   # USD → mUSD
) -> Path:
    """
    Строит xlsx с IS + BS + CF: история + прогноз бок о бок.
    Возвращает путь к файлу.
    """
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row

    # Scenario id
    row = conn.execute(
        "SELECT scenario_id FROM scenarios WHERE company_id=? AND name=?",
        (company_id, scenario_name),
    ).fetchone()
    if row is None:
        available = [r[0] for r in conn.execute(
            "SELECT name FROM scenarios WHERE company_id=?", (company_id,)
        ).fetchall()]
        raise ValueError(
            f"Сценарий '{scenario_name}' не найден для {company_id}. "
            f"Доступны: {available}"
        )
    scenario_id = row["scenario_id"]

    # Load data
    hist_data  = load_history(conn, company_id)
    fcast_data = load_forecast(conn, company_id, scenario_id)
    conn.close()

    # Years
    hist_years  = sorted(hist_data["IS"].keys() | hist_data["BS"].keys() | hist_data["CF"].keys())
    fcast_years = sorted(fcast_data["IS"].keys() | fcast_data["BS"].keys() | fcast_data["CF"].keys())

    if not hist_years:
        raise RuntimeError(f"Нет исторических данных для {company_id}")
    if not fcast_years:
        raise RuntimeError(
            f"Нет прогнозных данных для {company_id} сценарий={scenario_name}. "
            f"Запустите build_model() сначала."
        )

    print(f"  История:  {hist_years[0]}–{hist_years[-1]}  ({len(hist_years)} лет)")
    print(f"  Прогноз:  {fcast_years[0]}–{fcast_years[-1]}  ({len(fcast_years)} лет)")

    # Build workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # удалить Sheet по умолчанию

    for stmt_name, schema in [
        ("Income Statement", IS_SCHEMA),
        ("Balance Sheet",    BS_SCHEMA),
        ("Cash Flow",        CF_SCHEMA),
    ]:
        ws = wb.create_sheet(stmt_name)
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "C3"   # зафиксировать label + первые два ряда

        builder = SheetBuilder(ws, hist_years, fcast_years, display_unit)
        builder.write_year_header()
        builder.write_section(schema, hist_data, fcast_data)

    # Output path
    if out_path is None:
        out_dir = ROOT / "companies" / company_id / "outputs"
        out_dir.mkdir(parents=True, exist_ok=True)
        out_path = out_dir / f"financial_statements_{company_id}_{scenario_name}.xlsx"

    wb.save(out_path)
    print(f"  Сохранено: {out_path}")
    return out_path


# ═══════════════════════════════════════════════════════════════
# CLI
# ═══════════════════════════════════════════════════════════════

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Экспорт IS+BS+CF: история + прогноз в формате Financial Statements"
    )
    parser.add_argument("--company",  required=True, help="company_id (rusal, us_steel, nornickel)")
    parser.add_argument("--scenario", default="base", help="Сценарий (default: base)")
    parser.add_argument("--out",      default=None,   help="Путь к выходному xlsx")
    parser.add_argument("--unit",     default=1e6, type=float,
                        help="Делитель для отображения (default: 1e6 → mUSD)")
    args = parser.parse_args()

    out = Path(args.out) if args.out else None
    print(f"\nЭкспорт Financial Statements: {args.company} / {args.scenario}")
    result = export_financial_statements(
        company_id    = args.company,
        scenario_name = args.scenario,
        out_path      = out,
        display_unit  = args.unit,
    )
    print(f"Готово: {result}\n")


if __name__ == "__main__":
    main()
