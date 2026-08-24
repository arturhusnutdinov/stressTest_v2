"""
rebuild_nornickel_excel.py — перестройка Excel как сырой слепок Databook.

Принцип: каждый лист Databook → лист Excel с теми же строками.
Годы: только полные 2011-2025 (без полугодовых).
Данные: 1-в-1 из Databook, без агрегации.
"""
import openpyxl
from pathlib import Path
from copy import copy

ROOT = Path(__file__).resolve().parent.parent
COMPANY_DIR = ROOT / "companies" / "nornickel"
DATABOOK = COMPANY_DIR / "data" / "statements" / "Databook_12m_25_Final.xlsx"
EXCEL_OUT = COMPANY_DIR / "data" / "excel" / "nornickel_unified.xlsx"

TARGET_YEARS = list(range(2011, 2026))

def load_sheet(name):
    wb = openpyxl.load_workbook(DATABOOK, data_only=True)
    ws = wb[name]
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append([v for v in row])
    wb.close()
    return rows

def build_year_map(header_row):
    """Return {full_year: col_index} skipping half-years like '1H 10'."""
    yr_map = {}
    for i, val in enumerate(header_row):
        if val is None:
            continue
        try:
            y = int(float(str(val)))
            if 2009 <= y <= 2026:
                yr_map[y] = i
        except (ValueError, TypeError):
            pass
    return yr_map

# ── Load Databook sheets ─────────────────────────────────────────────────

print("Loading Databook...")

sheets_to_mirror = {
    "INCOME STATEMENT":        "db_income_statement",
    "BALANCE ":                "db_balance_sheet",
    "CASH FLOW STATEMENT":     "db_cash_flow",
    "EBITDA CALCULATION":      "db_ebitda",
    "COST BREAKDOWN":          "db_cost_breakdown",
    "REVENUE BREAKDOWN":       "db_revenue_breakdown",
    "CAPEX BREAKDOWN":         "db_capex_breakdown",
    "DEBT AND LIQUIDITY ":     "db_debt_liquidity",
    "PRODUCTION DATA":         "db_production",
    "ORE OUTPUT":              "db_ore_output",
    "RECOVERY RATES":          "db_recovery_rates",
    "WORKING CAPITAL":         "db_working_capital",
    "SELECTED FINANCIAL RATIOS": "db_financial_ratios",
    "MINERAL RESOURCES_ORE RESERVES": "db_mineral_reserves",
}

# Load all source data
source_data = {}
for src_name in sheets_to_mirror:
    try:
        source_data[src_name] = load_sheet(src_name)
        print(f"  ✓ {src_name} ({len(source_data[src_name])} rows)")
    except Exception as e:
        print(f"  ⚠ {src_name}: {e}")

# ── Open target Excel ────────────────────────────────────────────────────

wb = openpyxl.load_workbook(EXCEL_OUT)

# Remove old db_ sheets if they exist
for name in list(wb.sheetnames):
    if name.startswith('db_'):
        del wb[name]

# ── Mirror each Databook sheet ───────────────────────────────────────────

for src_name, excel_name in sheets_to_mirror.items():
    if src_name not in source_data:
        continue
    
    rows = source_data[src_name]
    
    # Try multiple rows for year header (some sheets have it at row 0, others row 1)
    yr_map = {}
    for header_candidate in [0, 1, 2]:
        candidate_map = build_year_map(rows[header_candidate])
        if len(candidate_map) >= len(yr_map):
            yr_map = candidate_map
    if not yr_map:
        print(f"  ⚠ No year columns found, skipping data fill")
        continue

    print(f"\nMirroring {src_name} → {excel_name}")
    print(f"  Years found: {sorted(yr_map.keys())}")

    # Create sheet
    ws = wb.create_sheet(excel_name)

    # Find which rows have data labels (row[0] is not None/empty)
    # and which rows are data vs section headers

    # Strategy: write ALL rows from Databook, but with only TARGET_YEARS columns
    # Column 1 = label (from Databook col A)
    # Columns 2-16 = 2011-2025

    excel_row = 1
    # Header row
    ws.cell(row=1, column=1, value="Line item")
    for i, yr in enumerate(TARGET_YEARS):
        ws.cell(row=1, column=i + 2, value=yr)

    for src_row_idx, src_row in enumerate(rows):
        label = src_row[0]
        if label is None:
            label = ""

        # Skip the Databook header rows (first 2: title + "USD mln")
        if src_row_idx <= 1:
            continue

        label_str = str(label).strip()
        
        # Write label
        excel_row += 1
        ws.cell(row=excel_row, column=1, value=label_str)

        # Write data for each target year
        for i, yr in enumerate(TARGET_YEARS):
            src_col = yr_map.get(yr)
            if src_col is not None and src_col < len(src_row):
                v = src_row[src_col]
                if v is not None and str(v) != 'NA' and str(v) != '-':
                    try:
                        # Try as number
                        num = float(str(v).replace(",", "").replace(" ", ""))
                        ws.cell(row=excel_row, column=i + 2, value=num)
                    except (ValueError, TypeError):
                        # Keep as string (e.g. section headers)
                        if label_str == '':
                            ws.cell(row=excel_row, column=i + 2, value=str(v))

    # Auto-fit column A width
    ws.column_dimensions['A'].width = 60

    print(f"  → {excel_row} rows written")

# ── METADATA sheet ────────────────────────────────────────────────────────

# Add db_metadata
if 'db_metadata' in wb.sheetnames:
    del wb['db_metadata']
ws_meta = wb.create_sheet('db_metadata')
ws_meta.cell(row=1, column=1, value="Source")
ws_meta.cell(row=1, column=2, value="Databook_12m_25_Final.xlsx")
ws_meta.cell(row=2, column=1, value="Company")
ws_meta.cell(row=2, column=2, value="PJSC MMC Norilsk Nickel")
ws_meta.cell(row=3, column=1, value="Standard")
ws_meta.cell(row=3, column=2, value="IFRS")
ws_meta.cell(row=4, column=1, value="Currency")
ws_meta.cell(row=4, column=2, value="USD million (unless noted)")
ws_meta.cell(row=5, column=1, value="Period")
ws_meta.cell(row=5, column=2, value="2011-2025 (full years only)")
ws_meta.cell(row=6, column=1, value="Note")
ws_meta.cell(row=6, column=2, value="Half-year columns excluded. Raw data from Databook, 'NA' = not disclosed.")
ws_meta.cell(row=7, column=1, value="Generated")
ws_meta.cell(row=7, column=2, value="2026-05-19")
ws_meta.column_dimensions['A'].width = 20
ws_meta.column_dimensions['B'].width = 80

# ── Save ──────────────────────────────────────────────────────────────────

# Move db_metadata to first position
sheet_names = wb.sheetnames
idx = sheet_names.index('db_metadata')
wb.move_sheet('db_metadata', offset=-idx)

print(f"\nSaving to {EXCEL_OUT}...")
wb.save(EXCEL_OUT)

# ── Summary ───────────────────────────────────────────────────────────────

print("\n" + "=" * 70)
print("FINAL STRUCTURE")
print("=" * 70)
wb2 = openpyxl.load_workbook(EXCEL_OUT, data_only=True)
for name in wb2.sheetnames:
    ws = wb2[name]
    # Count non-empty data cells
    filled = 0
    total = 0
    for r in range(2, ws.max_row + 1):
        for c in range(2, ws.max_column + 1):
            total += 1
            if ws.cell(row=r, column=c).value is not None:
                filled += 1
    pct = f"{filled/total*100:.0f}%" if total > 0 else "—"
    prefix = "📋" if name.startswith('db_') else "📝"
    print(f"  {prefix} {name:<35s} {ws.max_row}r × {ws.max_column}c | {filled}/{total} cells ({pct})")
wb2.close()
print(f"\n{'='*70}")
print("db_* листы = сырые данные из Databook (1-в-1)")
print("Остальные = для движка модели (агрегированные)")
print(f"{'='*70}")
