"""
export_to_excel.py — Export ALL company data from data_mart_v2.db to Excel.

Creates a fully-populated Excel workbook matching the 31-sheet template.
Analyst adds new year column, then load_unified_excel.py loads only new data.

Usage:
    python3 tools/export_to_excel.py --company rusal
    python3 tools/export_to_excel.py --company us_steel --output us_steel_full.xlsx
"""
from __future__ import annotations

import argparse
import sqlite3
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from engine import DB_PATH

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("pip install openpyxl")
    sys.exit(1)


# ── Styles ────────────────────────────────────────────────────────────────────

HEADER_FONT = Font(bold=True, size=10, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
METRIC_FONT = Font(size=10, name="Consolas")
NUM_FMT = '#,##0'
PCT_FMT = '0.00%'


def _style_header(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')


def _auto_width(ws, min_width=12, max_width=40):
    ws.column_dimensions['A'].width = max_width
    if ws.max_column and ws.max_column > 1:
        for col in range(2, ws.max_column + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = min_width


# ── Helpers ───────────────────────────────────────────────────────────────────

def _get_years(conn, company_id: str) -> list[int]:
    """Get all historical years for company."""
    rows = conn.execute("""
        SELECT DISTINCT p.year FROM periods p
        WHERE p.company_id=? AND p.is_forecast=0
        ORDER BY p.year
    """, (company_id,)).fetchall()
    return [r[0] for r in rows]


def _wide_sheet(ws, conn, table: str, company_id: str, years: list[int],
                to_musd: bool = True, metric_col: str = "metric"):
    """Export EAV table (history_is/bs/cf) as wide format: metric × years."""
    metrics = [r[0] for r in conn.execute(f"""
        SELECT DISTINCT h.{metric_col} FROM {table} h
        JOIN periods p ON h.period_id=p.period_id
        WHERE h.company_id=? ORDER BY h.{metric_col}
    """, (company_id,)).fetchall()]

    ws.cell(1, 1, metric_col)
    for j, yr in enumerate(years):
        ws.cell(1, j + 2, yr)
    ws.cell(1, len(years) + 2, "unit")
    _style_header(ws, 1, len(years) + 2)

    for i, metric in enumerate(metrics, 2):
        ws.cell(i, 1, metric)
        rows = conn.execute(f"""
            SELECT p.year, h.value FROM {table} h
            JOIN periods p ON h.period_id=p.period_id
            WHERE h.company_id=? AND h.{metric_col}=?
        """, (company_id, metric)).fetchall()
        val_by_yr = {r["year"]: r["value"] for r in rows}
        for j, yr in enumerate(years):
            v = val_by_yr.get(yr)
            if v is not None:
                ws.cell(i, j + 2, v / 1e6 if to_musd else v)
                ws.cell(i, j + 2).number_format = NUM_FMT
        ws.cell(i, len(years) + 2, "mUSD" if to_musd else "")
        ws.cell(i, 1).font = METRIC_FONT


# ── Main Export ───────────────────────────────────────────────────────────────

def export_company(company_id: str, output_path: Path):
    conn = sqlite3.connect(str(DB_PATH))
    conn.row_factory = sqlite3.Row

    co = conn.execute("SELECT * FROM companies WHERE company_id=?", (company_id,)).fetchone()
    if not co:
        print(f"Company '{company_id}' not found in DB")
        return

    years = _get_years(conn, company_id)
    if not years:
        print(f"No historical data for '{company_id}'")
        return

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ═══════════════════════════════════════════════════════════════
    # 1. META
    # ═══════════════════════════════════════════════════════════════
    ws = wb.create_sheet("meta")
    meta = [
        ("template_version", "2.1.0"),
        ("company_code", company_id),
        ("company_name", co["name"] or company_id),
        ("base_currency", co["currency"] or "USD"),
        ("input_unit", "mUSD"),
        ("accounting_standard", co["accounting_standard"] or "US_GAAP"),
        ("history_start_year", years[0]),
        ("history_end_year", years[-1]),
    ]
    ws.cell(1, 1, "field").font = Font(bold=True)
    ws.cell(1, 2, "value").font = Font(bold=True)
    for i, (f, v) in enumerate(meta, 2):
        ws.cell(i, 1, f)
        ws.cell(i, 2, v)
    _auto_width(ws)

    # ═══════════════════════════════════════════════════════════════
    # 2-4. HISTORY IS / BS / CF
    # ═══════════════════════════════════════════════════════════════
    for stmt in ["is", "bs", "cf"]:
        ws = wb.create_sheet(f"history_{stmt}")
        _wide_sheet(ws, conn, f"history_{stmt}", company_id, years, to_musd=True)
        _auto_width(ws)

    # ═══════════════════════════════════════════════════════════════
    # 5. SCHEDULE_TAX → tax_schedule
    # ═══════════════════════════════════════════════════════════════
    ws = wb.create_sheet("schedule_tax")
    tax_cols = ["year", "ebt_mUSD", "current_tax_mUSD", "deferred_tax_mUSD",
                "effective_rate", "dta_open_mUSD", "dta_close_mUSD",
                "dtl_open_mUSD", "dtl_close_mUSD",
                "nol_open_mUSD", "nol_close_mUSD"]
    for j, c in enumerate(tax_cols, 1):
        ws.cell(1, j, c)
    _style_header(ws, 1, len(tax_cols))

    tax_rows = conn.execute("""
        SELECT p.year, t.* FROM tax_schedule t
        JOIN periods p ON t.period_id=p.period_id
        WHERE t.company_id=? ORDER BY p.year
    """, (company_id,)).fetchall()
    for i, r in enumerate(tax_rows, 2):
        ws.cell(i, 1, r["year"])
        for j, col in enumerate(["ebt", "current_tax", "deferred_tax"], 2):
            v = r[col]
            ws.cell(i, j, v / 1e6 if v else None)
        ws.cell(i, 5, r["effective_rate"])
        for j, col in enumerate(["dta_open", "dta_close", "dtl_open", "dtl_close",
                                  "nol_open", "nol_close"], 6):
            v = r[col]
            ws.cell(i, j, v / 1e6 if v else None)

    # ═══════════════════════════════════════════════════════════════
    # 6. SCHEDULE_EQUITY → equity_schedule
    # ═══════════════════════════════════════════════════════════════
    ws = wb.create_sheet("schedule_equity")
    eq_cols = ["year", "re_open_mUSD", "net_income_mUSD", "dividends_mUSD",
               "buybacks_mUSD", "issuance_mUSD", "other_equity_changes_mUSD", "re_close_mUSD"]
    for j, c in enumerate(eq_cols, 1):
        ws.cell(1, j, c)
    _style_header(ws, 1, len(eq_cols))

    eq_rows = conn.execute("""
        SELECT p.year, e.* FROM equity_schedule e
        JOIN periods p ON e.period_id=p.period_id
        WHERE e.company_id=? ORDER BY p.year
    """, (company_id,)).fetchall()
    for i, r in enumerate(eq_rows, 2):
        ws.cell(i, 1, r["year"])
        for j, col in enumerate(["re_open", "net_income", "dividends", "buybacks",
                                  "issuance", "other_equity_changes", "re_close"], 2):
            v = r[col]
            ws.cell(i, j, v / 1e6 if v else None)

    # ═══════════════════════════════════════════════════════════════
    # 7. SCHEDULE_LEASES → lease_schedule
    # ═══════════════════════════════════════════════════════════════
    ws = wb.create_sheet("schedule_leases")
    ls_cols = ["year", "lease_type", "lease_id", "lease_name",
               "rou_open_mUSD", "rou_dep_mUSD", "rou_close_mUSD",
               "liab_open_mUSD", "interest_exp_mUSD", "payment_mUSD",
               "liab_close_mUSD", "discount_rate"]
    for j, c in enumerate(ls_cols, 1):
        ws.cell(1, j, c)
    _style_header(ws, 1, len(ls_cols))

    ls_rows = conn.execute("""
        SELECT p.year, l.* FROM lease_schedule l
        JOIN periods p ON l.period_id=p.period_id
        WHERE l.company_id=? ORDER BY p.year, l.lease_id
    """, (company_id,)).fetchall()
    for i, r in enumerate(ls_rows, 2):
        ws.cell(i, 1, r["year"])
        ws.cell(i, 2, r["lease_type"])
        ws.cell(i, 3, r["lease_id"])
        ws.cell(i, 4, r["lease_name"])
        for j, col in enumerate(["rou_open", "rou_dep", "rou_close",
                                  "liab_open", "interest_exp", "payment", "liab_close"], 5):
            v = r[col]
            ws.cell(i, j, v / 1e6 if v else None)
        ws.cell(i, 12, r["discount_rate"])

    # ═══════════════════════════════════════════════════════════════
    # 8-9. EMPTY SCHEDULE SHEETS (WC, interest) — headers only
    # ═══════════════════════════════════════════════════════════════
    for sheet_name, cols in [
        ("schedule_working_capital", ["year", "component", "opening_balance_mUSD",
                                       "closing_balance_mUSD", "delta_mUSD",
                                       "driver_value", "driver_metric"]),
        ("schedule_interest", ["year", "interest_paid_debt_mUSD", "interest_paid_leases_mUSD",
                                "interest_paid_total_mUSD",
                                "interest_payable_debt_open_mUSD", "interest_payable_debt_close_mUSD",
                                "interest_payable_leases_open_mUSD", "interest_payable_leases_close_mUSD"]),
    ]:
        ws = wb.create_sheet(sheet_name)
        for j, c in enumerate(cols, 1):
            ws.cell(1, j, c)
        _style_header(ws, 1, len(cols))

    # ═══════════════════════════════════════════════════════════════
    # 10. PPE_COMPONENTS
    # ═══════════════════════════════════════════════════════════════
    ws = wb.create_sheet("ppe_components")
    ppe_cols = ["year", "component_id", "component_name", "value_type", "value_mUSD", "useful_life"]
    for j, c in enumerate(ppe_cols, 1):
        ws.cell(1, j, c)
    _style_header(ws, 1, len(ppe_cols))

    ppe_rows = conn.execute("""
        SELECT p.year, c.* FROM ppe_components c
        JOIN periods p ON c.period_id=p.period_id
        WHERE c.company_id=? ORDER BY p.year, c.component_id, c.value_type
    """, (company_id,)).fetchall()
    for i, r in enumerate(ppe_rows, 2):
        ws.cell(i, 1, r["year"])
        ws.cell(i, 2, r["component_id"])
        ws.cell(i, 3, r["component_name"])
        ws.cell(i, 4, r["value_type"])
        v = r["value"]
        ws.cell(i, 5, v / 1e6 if v else None)
        ws.cell(i, 6, r["useful_life"])

    # ═══════════════════════════════════════════════════════════════
    # 11. INTANGIBLE_ASSETS
    # ═══════════════════════════════════════════════════════════════
    ws = wb.create_sheet("intangible_assets")
    ia_cols = ["year", "category", "gross_amount_mUSD", "accumulated_amortization_mUSD",
               "net_amount_mUSD", "useful_life"]
    for j, c in enumerate(ia_cols, 1):
        ws.cell(1, j, c)
    _style_header(ws, 1, len(ia_cols))

    ia_rows = conn.execute("""
        SELECT p.year, a.* FROM intangible_assets a
        JOIN periods p ON a.period_id=p.period_id
        WHERE a.company_id=? ORDER BY p.year, a.category
    """, (company_id,)).fetchall()
    for i, r in enumerate(ia_rows, 2):
        ws.cell(i, 1, r["year"])
        ws.cell(i, 2, r["category"])
        for j, col in enumerate(["gross_amount", "accumulated_amortization", "net_amount"], 3):
            v = r[col]
            ws.cell(i, j, v / 1e6 if v else None)
        ws.cell(i, 6, r["useful_life"])

    # ═══════════════════════════════════════════════════════════════
    # 12. PROVISIONS
    # ═══════════════════════════════════════════════════════════════
    ws = wb.create_sheet("provisions")
    prov_cols = ["year", "category", "closing_mUSD"]
    for j, c in enumerate(prov_cols, 1):
        ws.cell(1, j, c)
    _style_header(ws, 1, len(prov_cols))

    prov_rows = conn.execute("""
        SELECT p.year, pr.* FROM provisions_schedule pr
        JOIN periods p ON pr.period_id=p.period_id
        WHERE pr.company_id=? ORDER BY p.year, pr.category
    """, (company_id,)).fetchall()
    for i, r in enumerate(prov_rows, 2):
        ws.cell(i, 1, r["year"])
        ws.cell(i, 2, r["category"])
        v = r["closing"]
        ws.cell(i, 3, v / 1e6 if v else None)

    # ═══════════════════════════════════════════════════════════════
    # 13. ASSOCIATES
    # ═══════════════════════════════════════════════════════════════
    ws = wb.create_sheet("associates")
    assoc_cols = ["year", "category", "movement", "value_mUSD"]
    for j, c in enumerate(assoc_cols, 1):
        ws.cell(1, j, c)
    _style_header(ws, 1, len(assoc_cols))

    assoc_rows = conn.execute("""
        SELECT p.year, a.* FROM associates_schedule a
        JOIN periods p ON a.period_id=p.period_id
        WHERE a.company_id=? ORDER BY p.year, a.category, a.movement
    """, (company_id,)).fetchall()
    for i, r in enumerate(assoc_rows, 2):
        ws.cell(i, 1, r["year"])
        ws.cell(i, 2, r["category"])
        ws.cell(i, 3, r["movement"])
        v = r["value"]
        ws.cell(i, 4, v / 1e6 if v else None)

    # ═══════════════════════════════════════════════════════════════
    # 14. DEBT_INSTRUMENTS
    # ═══════════════════════════════════════════════════════════════
    ws = wb.create_sheet("debt_instruments")
    di_cols = ["instrument_id", "instrument_name", "db_type", "currency",
               "opening_balance_mUSD", "committed_amount_mUSD",
               "maturity_date", "interest_rate", "rate_type", "base_rate_factor",
               "payment_frequency", "amortization_profile", "callable_flag",
               "covenant_package"]
    for j, c in enumerate(di_cols, 1):
        ws.cell(1, j, c)
    _style_header(ws, 1, len(di_cols))

    di_rows = conn.execute("""
        SELECT * FROM debt_instruments WHERE company_id=?
        ORDER BY opening_balance DESC
    """, (company_id,)).fetchall()
    for i, r in enumerate(di_rows, 2):
        for j, col in enumerate(di_cols):
            v = r[col] if col in r.keys() else None
            if col in ("opening_balance_mUSD", "committed_amount_mUSD"):
                db_col = col.replace("_mUSD", "")
                v = r[db_col] if db_col in r.keys() else None
                if v:
                    v = v / 1e6
            ws.cell(i, j + 1, v)

    # ═══════════════════════════════════════════════════════════════
    # 15. DEBT_CASHFLOWS
    # ═══════════════════════════════════════════════════════════════
    ws = wb.create_sheet("debt_cashflows")
    dc_cols = ["instrument_id", "year", "period", "cashflow_type", "amount_mUSD", "currency", "note"]
    for j, c in enumerate(dc_cols, 1):
        ws.cell(1, j, c)
    _style_header(ws, 1, len(dc_cols))

    dc_rows = conn.execute("""
        SELECT * FROM debt_cashflows WHERE company_id=?
        ORDER BY instrument_id, year
    """, (company_id,)).fetchall()
    for i, r in enumerate(dc_rows, 2):
        ws.cell(i, 1, r["instrument_id"])
        ws.cell(i, 2, r["year"])
        ws.cell(i, 3, r["period"])
        ws.cell(i, 4, r["cashflow_type"])
        v = r["amount"]
        ws.cell(i, 5, v / 1e6 if v else None)
        ws.cell(i, 6, r["currency"])
        ws.cell(i, 7, r["note"])

    # ═══════════════════════════════════════════════════════════════
    # 16-17. SEGMENTS (financial + operational)
    # ═══════════════════════════════════════════════════════════════
    for sheet_name in ["segments_financial", "segments_operational"]:
        ws = wb.create_sheet(sheet_name)
        ws.cell(1, 1, "segment_name")
        ws.cell(1, 2, "metric")
        for j, yr in enumerate(years):
            ws.cell(1, j + 3, yr)
        _style_header(ws, 1, len(years) + 2)

        seg_rows = conn.execute("""
            SELECT sd.segment_name, sd.metric, p.year, sd.value
            FROM segment_data sd JOIN periods p ON sd.period_id=p.period_id
            WHERE sd.company_id=? ORDER BY sd.segment_name, sd.metric, p.year
        """, (company_id,)).fetchall()

        seg_data = {}
        for r in seg_rows:
            key = (r["segment_name"], r["metric"])
            seg_data.setdefault(key, {})[r["year"]] = r["value"]

        row_idx = 2
        for (seg, metric), yr_vals in sorted(seg_data.items()):
            ws.cell(row_idx, 1, seg)
            ws.cell(row_idx, 2, metric)
            for j, yr in enumerate(years):
                v = yr_vals.get(yr)
                if v is not None:
                    ws.cell(row_idx, j + 3, v / 1e6 if abs(v) > 1e6 else v)
            row_idx += 1
        _auto_width(ws)

    # ═══════════════════════════════════════════════════════════════
    # 18. MACRO_FACTORS (wide: factor × years)
    # ═══════════════════════════════════════════════════════════════
    ws = wb.create_sheet("macro_factors")
    # Get factors relevant to this company (global scope)
    factors = [r[0] for r in conn.execute("""
        SELECT DISTINCT factor_name FROM macro_factors
        WHERE scope='global' OR company_id=?
        ORDER BY factor_name
    """, (company_id,)).fetchall()]

    ws.cell(1, 1, "factor")
    for j, yr in enumerate(years):
        ws.cell(1, j + 2, yr)
    _style_header(ws, 1, len(years) + 1)

    for i, fname in enumerate(factors, 2):
        ws.cell(i, 1, fname)
        vals = conn.execute("""
            SELECT year, value FROM macro_factors
            WHERE factor_name=? AND (scope='global' OR company_id=?)
        """, (fname, company_id)).fetchall()
        val_by_yr = {r["year"]: r["value"] for r in vals}
        for j, yr in enumerate(years):
            v = val_by_yr.get(yr)
            if v is not None:
                ws.cell(i, j + 2, v)
    _auto_width(ws)

    # ═══════════════════════════════════════════════════════════════
    # 19. OPERATIONAL_DRIVERS (wide: driver × years)
    # ═══════════════════════════════════════════════════════════════
    ws = wb.create_sheet("operational_drivers")
    drivers = [r[0] for r in conn.execute("""
        SELECT DISTINCT metric FROM operational_drivers
        WHERE company_id=? ORDER BY metric
    """, (company_id,)).fetchall()]

    ws.cell(1, 1, "driver")
    ws.cell(1, 2, "unit")
    for j, yr in enumerate(years):
        ws.cell(1, j + 3, yr)
    _style_header(ws, 1, len(years) + 2)

    for i, drv in enumerate(drivers, 2):
        ws.cell(i, 1, drv)
        vals = conn.execute("""
            SELECT year, value, unit FROM operational_drivers
            WHERE company_id=? AND metric=?
        """, (company_id, drv)).fetchall()
        if vals:
            ws.cell(i, 2, vals[0]["unit"])
        val_by_yr = {r["year"]: r["value"] for r in vals}
        for j, yr in enumerate(years):
            v = val_by_yr.get(yr)
            if v is not None:
                ws.cell(i, j + 3, v)
    _auto_width(ws)

    # ═══════════════════════════════════════════════════════════════
    # 20-25. CORKSCREW DETAIL SHEETS (headers only if empty in DB)
    # ═══════════════════════════════════════════════════════════════
    corkscrew_sheets = [
        ("sched_lease_finance", ["year", "lease_id", "opening_mUSD", "additions_mUSD",
            "payments_principal_mUSD", "payments_interest_mUSD",
            "depreciation_is_mUSD", "interest_expense_is_mUSD", "closing_mUSD",
            "rou_asset_open_mUSD", "rou_asset_dep_mUSD", "rou_asset_close_mUSD",
            "liab_current_mUSD", "liab_noncurrent_mUSD", "mode"]),
        ("sched_lease_operating", ["year", "lease_id", "opening_mUSD", "additions_mUSD",
            "payments_mUSD", "lease_expense_is_mUSD", "closing_mUSD",
            "rou_asset_open_mUSD", "rou_asset_dep_mUSD", "rou_asset_close_mUSD",
            "liab_current_mUSD", "liab_noncurrent_mUSD", "mode"]),
        ("sched_tax_corkscrew", ["year", "temp_diff_type",
            "dta_opening_mUSD", "dta_created_mUSD", "dta_utilized_mUSD", "dta_closing_mUSD",
            "dtl_opening_mUSD", "dtl_created_mUSD", "dtl_reversed_mUSD", "dtl_closing_mUSD"]),
        ("sched_wc_corkscrew", ["year", "component",
            "opening_balance_mUSD", "closing_balance_mUSD", "delta_mUSD",
            "driver_value", "driver_metric"]),
        ("interest_paid_split", ["year",
            "interest_paid_debt_mUSD", "interest_paid_leases_mUSD", "interest_paid_total_mUSD",
            "interest_payable_debt_open_mUSD", "interest_payable_debt_close_mUSD",
            "interest_payable_leases_open_mUSD", "interest_payable_leases_close_mUSD",
            "change_debt_mUSD", "change_leases_mUSD"]),
        ("lease_maturity_ladder", ["year", "lease_id", "lease_type", "maturity_year",
            "principal_amount_mUSD", "interest_amount_mUSD", "total_payment_mUSD", "currency_code"]),
    ]
    for sheet_name, cols in corkscrew_sheets:
        ws = wb.create_sheet(sheet_name)
        for j, c in enumerate(cols, 1):
            ws.cell(1, j, c)
        _style_header(ws, 1, len(cols))

    # ═══════════════════════════════════════════════════════════════
    # 26. BALANCING_ADJUSTMENTS
    # ═══════════════════════════════════════════════════════════════
    ws = wb.create_sheet("balancing_adjustments")
    ba_cols = ["year", "statement_type", "metric", "adjustment_value_mUSD",
               "is_balancing", "balancing_reason", "balancing_category", "original_value_mUSD"]
    for j, c in enumerate(ba_cols, 1):
        ws.cell(1, j, c)
    _style_header(ws, 1, len(ba_cols))

    # ═══════════════════════════════════════════════════════════════
    # 27-30. DICTIONARIES
    # ═══════════════════════════════════════════════════════════════
    for sheet_name, cols in [
        ("dictionary_metrics", ["canonical_metric", "statement", "description", "accepted_aliases"]),
        ("dictionary_debt_types", ["instrument_type", "description", "amortization_default", "is_active"]),
        ("dictionary_segments", ["segment_name", "description", "is_active", "commodity"]),
        ("dictionary_units", ["unit", "description", "multiplier"]),
    ]:
        ws = wb.create_sheet(sheet_name)
        for j, c in enumerate(cols, 1):
            ws.cell(1, j, c)
        _style_header(ws, 1, len(cols))

    # ═══════════════════════════════════════════════════════════════
    # 31. SCHEDULE_PPE (header only — data in ppe_components)
    # ═══════════════════════════════════════════════════════════════
    ws = wb.create_sheet("schedule_ppe")
    sp_cols = ["year", "category", "value_type", "value_mUSD", "useful_life"]
    for j, c in enumerate(sp_cols, 1):
        ws.cell(1, j, c)
    _style_header(ws, 1, len(sp_cols))

    # ── Save ──────────────────────────────────────────────────────
    conn.close()
    wb.save(str(output_path))
    print(f"✅ Exported: {output_path}")
    print(f"   Sheets: {len(wb.sheetnames)}")
    for name in wb.sheetnames:
        ws = wb[name]
        rows = ws.max_row - 1 if ws.max_row > 1 else 0
        print(f"   {name:35s} {rows:>5} rows")
    print(f"   Size: {output_path.stat().st_size // 1024}KB")


def main():
    parser = argparse.ArgumentParser(description="Export ALL company data from DB to unified Excel")
    parser.add_argument("--company", required=True, help="Company ID")
    parser.add_argument("--output", default=None, help="Output xlsx path")
    args = parser.parse_args()

    if args.output:
        out = Path(args.output)
    else:
        out = ROOT / "companies" / args.company / "data" / "excel" / f"{args.company}_unified.xlsx"

    out.parent.mkdir(parents=True, exist_ok=True)
    export_company(args.company, out)


if __name__ == "__main__":
    main()
