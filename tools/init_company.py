"""
init_company.py — инициализация новой компании.

Создаёт структуру директорий, копирует шаблоны конфигов и ноутбуков,
подставляет company_id во всех шаблонах.

Использование:
    python3 tools/init_company.py test_corp --name "Test Corporation" --industry metals --standard IFRS
    python3 tools/init_company.py test_corp --force  # перезаписать
"""
from __future__ import annotations

import argparse
import json
import shutil
from pathlib import Path


# Корень проекта — один уровень выше tools/
ROOT = Path(__file__).resolve().parent.parent
TEMPLATES_DIR = ROOT / "templates"


# ── Структура директорий ─────────────────────────────────────────────────────

DIRS = [
    "configs",
    "configs/forecast",
    "data",
    "data/excel",
    "data/macro",
    "data/debt",
    "data/operational",
    "data/statements",
    "data/annual_reports",
    "notebooks",
    "outputs/model",
    "outputs/stress",
    "outputs/reports",
]

# Ноутбуки: шаблон → целевое имя
NOTEBOOKS = {
    "00_Build_Model_Main.ipynb":              "00_Build_Model_Main.ipynb",
    "01_Data_Loading.ipynb":                  "01_Data_Loading.ipynb",
    "01_Test_Macro_Module.ipynb":             "01_Test_Macro_Module.ipynb",
    "02_Test_Model_Module.ipynb":             "02_Test_Model_Module.ipynb",
    "03_Stress_Testing.ipynb":                "03_Stress_Testing.ipynb",
    "04_Rating.ipynb":                        "04_Rating.ipynb",
    "05_Covenants.ipynb":                     "05_Covenants.ipynb",
    "98_Generate_Model_Documentation.ipynb":  "98_Generate_Model_Documentation.ipynb",
    "99_Configure_Excel_Loader.ipynb":        "99_Configure_Excel_Loader.ipynb",
    "99_Configure_YAML.ipynb":                "99_Configure_YAML.ipynb",
}

# Placeholder ноутбуки (создаются если шаблон не найден)
PLACEHOLDER_NOTEBOOKS = []


def init_company(
    company_id: str,
    name: str,
    industry: str = "metals",
    currency: str = "USD",
    standard: str = "US_GAAP",
    force: bool = False,
) -> Path:
    """Инициализирует структуру новой компании."""
    company_dir = ROOT / "companies" / company_id

    if company_dir.exists() and not force:
        print(f"  ⚠ Директория уже существует: {company_dir}")
        print(f"    Используйте --force для перезаписи")
        return company_dir

    print(f"Инициализация компании: {name} ({company_id})")
    print(f"  Отрасль:   {industry}")
    print(f"  Валюта:    {currency}")
    print(f"  Стандарт:  {standard}")
    print(f"  Директория: {company_dir}")
    print()

    # Общие подстановки для всех файлов
    replacements = {
        "{company}": company_id,
        "{COMPANY_NAME}": name,
        "{INDUSTRY}": industry,
        "{CURRENCY}": currency,
        "{STANDARD}": standard,
        "us_steel": company_id,
        "US Steel": name,
        "_COMPANY_": company_id,
    }

    def _apply_replacements(content: str) -> str:
        for old, new in replacements.items():
            content = content.replace(old, new)
        return content

    # ── 1. Создаём директории ────────────────────────────────────────────────
    for d in DIRS:
        (company_dir / d).mkdir(parents=True, exist_ok=True)
    print(f"  ✓ Директории ({len(DIRS)})")

    # ── 2. project.yaml ──────────────────────────────────────────────────────
    src = TEMPLATES_DIR / "project_template.yaml"
    dst = company_dir / "configs" / "project.yaml"
    if src.exists():
        content = _apply_replacements(src.read_text())
        # Подстановка is_income_sign по стандарту
        if standard == "IFRS":
            content = content.replace("is_income_sign: credit_negative", "is_income_sign: natural")
            content = content.replace("da_in_cogs: true", "da_in_cogs: false")
        dst.write_text(content)
    else:
        dst.write_text(f"# project.yaml — {name}\nmodel:\n  mode: standard\n")
    print(f"  ✓ configs/project.yaml")

    # ── 3. accounting_conventions.yaml ──────────────────────────────────────
    src = TEMPLATES_DIR / "accounting_conventions_template.yaml"
    dst = company_dir / "configs" / "accounting_conventions.yaml"
    if src.exists():
        dst.write_text(_apply_replacements(src.read_text()))
    else:
        dst.write_text(f"company: {company_id}\ncash_in_cf:\n  type: cash_only\n")
    print(f"  ✓ configs/accounting_conventions.yaml")

    # ── 4. excel_loader.yaml ─────────────────────────────────────────────────
    src = TEMPLATES_DIR / "excel_loader_template.yaml"
    if not src.exists():
        src = TEMPLATES_DIR / "mapping_template.yaml"
    dst = company_dir / "configs" / "excel_loader.yaml"
    if src.exists():
        dst.write_text(_apply_replacements(src.read_text()))
    else:
        dst.write_text(f"# Excel loader config for {name}\nloader_settings:\n  company_id: {company_id}\n")
    print(f"  ✓ configs/excel_loader.yaml")

    # ── 5. macro_ecm.yaml ────────────────────────────────────────────────────
    src = TEMPLATES_DIR / "macro_ecm_full_template.yaml"
    dst = company_dir / "configs" / "forecast" / "macro_ecm.yaml"
    if src.exists():
        dst.write_text(_apply_replacements(src.read_text()))
    else:
        dst.write_text(f"# Macro ECM config for {name}\n")
    print(f"  ✓ configs/forecast/macro_ecm.yaml")

    # ── 6. stress_scenarios.yaml ─────────────────────────────────────────────
    src = TEMPLATES_DIR / "scenario_template.yaml"
    dst = company_dir / "configs" / "stress_scenarios.yaml"
    if src.exists():
        dst.write_text(_apply_replacements(src.read_text()))
    else:
        _write_stress_scenarios(dst, industry)
    print(f"  ✓ configs/stress_scenarios.yaml")

    # ── 7. Ноутбуки ──────────────────────────────────────────────────────────
    notebooks_src = TEMPLATES_DIR / "notebooks"
    notebooks_dst = company_dir / "notebooks"

    copied = 0
    for src_name, dst_name in NOTEBOOKS.items():
        src = notebooks_src / src_name
        dst = notebooks_dst / dst_name
        if src.exists():
            _copy_notebook(src, dst, company_id, name)
            copied += 1
        else:
            _create_placeholder_notebook(dst, dst_name, company_id)

    for nb_name in PLACEHOLDER_NOTEBOOKS:
        _create_placeholder_notebook(notebooks_dst / nb_name, nb_name, company_id)

    print(f"  ✓ notebooks ({copied} скопировано + {len(PLACEHOLDER_NOTEBOOKS)} placeholder)")

    # ── 8. Excel шаблон ───────────────────────────────────────────────────────
    excel_path = company_dir / "data" / "excel" / f"{company_id}_unified.xlsx"
    if not excel_path.exists():
        _create_excel_template(excel_path, company_id)
        print(f"  ✓ data/excel/{company_id}_unified.xlsx")

    # ── 9. README.md ─────────────────────────────────────────────────────────
    (company_dir / "README.md").write_text(f"""\
# {name}

**company_id**: `{company_id}`
**Отрасль**: {industry}
**Валюта**: {currency}
**Стандарт учёта**: {standard}

## Структура
```
companies/{company_id}/
  configs/
    project.yaml              # Главный конфиг модели
    excel_loader.yaml         # Маппинг Excel → DB
    accounting_conventions.yaml
    forecast/macro_ecm.yaml   # Настройки макро-прогноза
    stress_scenarios.yaml     # Стресс-сценарии
  data/
    excel/                    # UNIFIED Excel с данными (IS/BS/CF/debt/macro)
    statements/               # МСФО / US GAAP отчётность (PDF)
    annual_reports/           # Годовые отчёты для акционеров (PDF)
    macro/                    # Макро-факторы (CSV)
    debt/                     # Долговые расписания
    operational/              # Операционные KPI
  notebooks/
    00_Build_Model_Main.ipynb # Главный pipeline
    01_Data_Loading.ipynb     # Загрузка данных
    02_Test_Model_Module.ipynb # Тестирование модели
    03_Stress_Testing.ipynb   # Стресс-тесты
    04_Rating.ipynb           # Кредитный рейтинг
    05_Covenants.ipynb        # Ковенанты
  outputs/                    # Результаты модели
```

## Быстрый старт

1. Заполните Excel: `data/excel/{company_id}_unified.xlsx`
2. Настройте: `configs/project.yaml`
3. Загрузите данные: `notebooks/01_Data_Loading.ipynb`
4. Запустите модель: `notebooks/00_Build_Model_Main.ipynb`

## CLI
```bash
# Полный прогон
python3 -m engine.orchestrator {company_id} --stress --rating

# Только модель
python3 -m engine.orchestrator {company_id} --no-preprocess
```
""")
    print(f"  ✓ README.md")

    print()
    print(f"✅ Компания '{company_id}' инициализирована: {company_dir}")
    print()
    print("Следующие шаги:")
    print(f"  1. Подготовьте Excel: companies/{company_id}/data/excel/{company_id}_unified.xlsx")
    print(f"  2. Отредактируйте:   companies/{company_id}/configs/project.yaml")
    print(f"  3. Запустите:        companies/{company_id}/notebooks/01_Data_Loading.ipynb")

    return company_dir


def _create_excel_template(path: Path, company_id: str) -> None:
    """Создаёт полный Excel шаблон со всеми листами для загрузки в БД.

    Листы соответствуют таблицам schema.py и покрывают:
    - Финансовые отчёты (IS/BS/CF)
    - Schedule-расписания (tax, equity, leases, WC, interest, PPE)
    - Детальные корксkrews (lease finance/operating, tax, WC)
    - Долговые инструменты + cashflows
    - Сегменты (финансовые + операционные)
    - Макро-факторы, операционные драйверы
    - Справочники (метрики, типы долга, сегменты, единицы)
    - Метаданные и балансировочные корректировки
    """
    try:
        import openpyxl
    except ImportError:
        path.write_bytes(b"")  # fallback: пустой файл
        return

    wb = openpyxl.Workbook()
    years = list(range(2011, 2026))

    # ── 1. meta ──────────────────────────────────────────────────
    ws = wb.active
    ws.title = "meta"
    ws.append(["Parameter", "Value"])
    ws.append(["template_version", "3.0"])
    ws.append(["company_id", company_id])
    ws.append(["company_code", company_id])
    ws.append(["company_name", ""])
    ws.append(["base_currency", "USD"])
    ws.append(["db_unit", "USD"])
    ws.append(["reporting_standard", ""])
    ws.append(["fiscal_year_end", "December"])

    # ── 2. history_is (template v3 format) ────────────────────────
    ws_is = wb.create_sheet("history_is")
    ws_is.append(["INCOME STATEMENT — template v3 (label | db_metric | sign | unit | years)"])
    ws_is.append(["label", "db_metric", "sign", "unit"] + years)
    IS_ROWS = [
        ("Revenue", "revenue", 1, True),
        ("Cost of Sales", "cogs", -1, True),
        ("Gross Profit", "gross_profit", 1, False),
        ("SG&A", "sga", -1, False),
        ("Distribution Expenses", "distribution_expenses", -1, False),
        ("R&D", "rnd", -1, False),
        ("Other Operating Income", "other_operating_income", 1, False),
        ("Other Operating Expense", "other_operating_expense", -1, False),
        ("Asset Impairment", "asset_impairment", -1, False),
        ("Depreciation (PPE)", "depreciation_owned", 1, False),
        ("Depreciation (RoU)", "depreciation_rou", 1, False),
        ("Amortization", "amortization", 1, False),
        ("Total D&A", "total_da", 1, False),
        ("EBITDA", "ebitda", 1, False),
        ("EBIT", "ebit", 1, False),
        ("Interest Expense", "interest_expense", -1, True),
        ("Interest Income", "interest_income", 1, False),
        ("Lease Interest", "lease_interest", -1, False),
        ("Earnings from Investees", "earnings_from_investees", 1, False),
        ("Other Financial Costs", "other_losses_gains_net", 1, False),
        ("EBT", "ebt", 1, False),
        ("Current Tax", "current_tax_expense", -1, False),
        ("Deferred Tax", "deferred_tax_expense", -1, False),
        ("Tax Expense", "tax_expense", -1, True),
        ("Net Income", "net_income", 1, True),
    ]
    for label, metric, sign, _ in IS_ROWS:
        ws_is.append([label, metric, sign, "mUSD"] + [None] * len(years))

    # ── 3. history_bs (template v3 format) ────────────────────────
    ws_bs = wb.create_sheet("history_bs")
    ws_bs.append(["BALANCE SHEET — template v3 (label | db_metric | sign | unit | years)"])
    ws_bs.append(["label", "db_metric", "sign", "unit"] + years)
    BS_ROWS = [
        ("Cash & Equivalents", "cash", 1), ("Restricted Cash", "restricted_cash", 1),
        ("Accounts Receivable", "accounts_receivable", 1), ("Inventory", "inventory", 1),
        ("Other Current Assets", "other_current_assets", 1),
        ("Total Current Assets", "total_current_assets", 1),
        ("PPE Net", "ppe_net", 1), ("RoU Asset", "rou_asset", 1),
        ("Intangibles", "intangibles", 1), ("Goodwill", "goodwill", 1),
        ("Investments in Associates", "investments_in_associates", 1),
        ("LT Investments", "investments_lt", 1),
        ("DTA", "dta", 1), ("Other Non-Current Assets", "other_nca", 1),
        ("Total Non-Current Assets", "total_non_current_assets", 1),
        ("Total Assets", "total_assets", 1),
        ("Short-Term Debt", "short_term_debt", 1), ("Accounts Payable", "accounts_payable", 1),
        ("Taxes Payable", "taxes_payable", 1), ("Interest Payable", "interest_payable", 1),
        ("Payroll & Benefits", "payroll_payable", 1),
        ("Lease Liab Current", "lease_liab_current", 1),
        ("Other Current Liabilities", "other_cl", 1),
        ("Total Current Liabilities", "total_current_liabilities", 1),
        ("Long-Term Debt", "long_term_debt", 1), ("DTL", "dtl", 1),
        ("Employee Benefits", "employee_benefits", 1),
        ("Lease Liab Non-Current", "lease_liab_noncurrent", 1),
        ("Other Non-Current Liabilities", "other_ncl", 1),
        ("Total Non-Current Liabilities", "total_non_current_liabilities", 1),
        ("Total Liabilities", "total_liabilities", 1),
        ("Share Capital", "share_capital", 1), ("APIC", "apic", 1),
        ("Treasury Stock", "treasury_stock", -1),
        ("Retained Earnings", "retained_earnings", 1),
        ("AOCI", "aoci", 1), ("NCI", "nci", 1),
        ("Total Equity", "total_equity", 1),
    ]
    for label, metric, sign in BS_ROWS:
        ws_bs.append([label, metric, sign, "mUSD"] + [None] * len(years))

    # ── 4. history_cf ────────────────────────────────────────────
    ws_cf = wb.create_sheet("history_cf")
    ws_cf.append(["metric"] + years)
    for m in [
        # Operating
        "net_income", "total_da", "depreciation_owned", "depreciation_rou",
        "amortization", "deferred_tax", "cfo_stock_compensation",
        "asset_impairment_charges", "loss_on_debt_extinguishment",
        "gain_loss_on_disposal", "earnings_from_investees",
        "change_ar", "change_inventory", "change_ap",
        "change_other_wc", "change_taxes_payable", "change_interest_payable",
        "interest_paid", "taxes_paid",
        "lease_payments_cfo", "other_operating", "cfo_total",
        # Investing
        "capex", "acquisitions", "disposal_proceeds",
        "other_investing", "cfi_total",
        # Financing
        "debt_issuance", "debt_repayments", "cff_borrowings", "cff_repayments",
        "finance_lease_principal", "dividends_paid", "share_buyback",
        "equity_issuance", "other_financing", "cff_total",
        # Cash bridge
        "fx_effect", "net_change_in_cash", "cash_beginning", "cash_ending",
    ]:
        ws_cf.append([m] + [None] * len(years))

    # ── 5. schedule_leases → lease_schedule ──────────────────────
    ws_sl = wb.create_sheet("schedule_leases")
    ws_sl.append(["year", "lease_type", "lease_id", "lease_name",
                  "rou_open_mUSD", "rou_dep_mUSD", "rou_close_mUSD",
                  "liab_open_mUSD", "interest_exp_mUSD", "payment_mUSD",
                  "liab_close_mUSD", "discount_rate"])

    # ── 6. schedule_ppe → ppe_components ─────────────────────────
    ws_sppe = wb.create_sheet("schedule_ppe")
    ws_sppe.append(["year", "category", "value_type", "value_mUSD", "useful_life"])

    # ── 7. schedule_tax → tax_schedule ───────────────────────────
    ws_st = wb.create_sheet("schedule_tax")
    ws_st.append(["year", "ebt_mUSD", "current_tax_mUSD", "deferred_tax_mUSD",
                  "effective_rate",
                  "dta_open_mUSD", "dta_additions_mUSD", "dta_used_mUSD", "dta_close_mUSD",
                  "dtl_open_mUSD", "dtl_additions_mUSD", "dtl_reversal_mUSD", "dtl_close_mUSD",
                  "nol_open_mUSD", "nol_additions_mUSD", "nol_used_mUSD", "nol_close_mUSD"])

    # ── 8. schedule_working_capital → sched_wc_corkscrew ─────────
    ws_swc = wb.create_sheet("schedule_working_capital")
    ws_swc.append(["year", "component", "opening_balance_mUSD",
                   "closing_balance_mUSD", "delta_mUSD",
                   "driver_value", "driver_metric"])

    # ── 9. schedule_interest → interest_paid_split ───────────────
    ws_si = wb.create_sheet("schedule_interest")
    ws_si.append(["year",
                  "interest_paid_debt_mUSD", "interest_paid_leases_mUSD", "interest_paid_total_mUSD",
                  "interest_payable_debt_open_mUSD", "interest_payable_debt_close_mUSD",
                  "interest_payable_leases_open_mUSD", "interest_payable_leases_close_mUSD"])

    # ── 10. schedule_equity → equity_schedule ────────────────────
    ws_se = wb.create_sheet("schedule_equity")
    ws_se.append(["year", "re_open_mUSD", "net_income_mUSD",
                  "dividends_mUSD", "buybacks_mUSD", "issuance_mUSD",
                  "other_equity_changes_mUSD", "re_close_mUSD"])

    # ── 11. sched_lease_finance ──────────────────────────────────
    ws_lf = wb.create_sheet("sched_lease_finance")
    ws_lf.append(["year", "lease_id",
                  "opening_mUSD", "additions_mUSD",
                  "payments_principal_mUSD", "payments_interest_mUSD",
                  "depreciation_is_mUSD", "interest_expense_is_mUSD", "closing_mUSD",
                  "rou_asset_open_mUSD", "rou_asset_dep_mUSD", "rou_asset_close_mUSD",
                  "liab_current_mUSD", "liab_noncurrent_mUSD", "mode"])

    # ── 12. sched_lease_operating ────────────────────────────────
    ws_lo = wb.create_sheet("sched_lease_operating")
    ws_lo.append(["year", "lease_id",
                  "opening_mUSD", "additions_mUSD", "payments_mUSD",
                  "lease_expense_is_mUSD", "closing_mUSD",
                  "rou_asset_open_mUSD", "rou_asset_dep_mUSD", "rou_asset_close_mUSD",
                  "liab_current_mUSD", "liab_noncurrent_mUSD", "mode"])

    # ── 13. sched_tax_corkscrew ──────────────────────────────────
    ws_tc = wb.create_sheet("sched_tax_corkscrew")
    ws_tc.append(["year", "temp_diff_type",
                  "dta_opening_mUSD", "dta_created_mUSD", "dta_utilized_mUSD", "dta_closing_mUSD",
                  "dtl_opening_mUSD", "dtl_created_mUSD", "dtl_reversed_mUSD", "dtl_closing_mUSD"])

    # ── 14. sched_wc_corkscrew ───────────────────────────────────
    ws_wc = wb.create_sheet("sched_wc_corkscrew")
    ws_wc.append(["year", "component",
                  "opening_balance_mUSD", "closing_balance_mUSD", "delta_mUSD",
                  "driver_value", "driver_metric"])

    # ── 15. interest_paid_split ──────────────────────────────────
    ws_ips = wb.create_sheet("interest_paid_split")
    ws_ips.append(["year",
                   "interest_paid_debt_mUSD", "interest_paid_leases_mUSD",
                   "interest_paid_total_mUSD",
                   "interest_payable_debt_open_mUSD", "interest_payable_debt_close_mUSD",
                   "interest_payable_leases_open_mUSD", "interest_payable_leases_close_mUSD",
                   "change_debt_mUSD", "change_leases_mUSD"])

    # ── 16. lease_maturity_ladder ────────────────────────────────
    ws_lml = wb.create_sheet("lease_maturity_ladder")
    ws_lml.append(["year", "lease_id", "lease_type", "maturity_year",
                   "principal_amount_mUSD", "interest_amount_mUSD",
                   "total_payment_mUSD", "currency_code"])

    # ── 17. segments_financial → segment_data ────────────────────
    ws_sf = wb.create_sheet("segments_financial")
    ws_sf.append(["segment_name", "metric"] + years)

    # ── 18. segments_operational → segment_data ──────────────────
    ws_so = wb.create_sheet("segments_operational")
    ws_so.append(["segment_name", "metric"] + years)

    # ── 19. debt_instruments ─────────────────────────────────────
    ws_di = wb.create_sheet("debt_instruments")
    ws_di.append(["instrument_id", "instrument_name", "db_type", "currency",
                  "opening_balance_mUSD", "committed_amount_mUSD",
                  "maturity_date", "interest_rate", "rate_type", "base_rate_factor",
                  "payment_frequency", "amortization_profile", "callable_flag",
                  "covenant_package"])

    # ── 20. debt_cashflows ───────────────────────────────────────
    ws_dc = wb.create_sheet("debt_cashflows")
    ws_dc.append(["instrument_id", "year", "period", "cashflow_type",
                  "amount_mUSD", "currency", "note"])

    # ── 21. ppe_components → ppe_components ──────────────────────
    ws_pp = wb.create_sheet("ppe_components")
    ws_pp.append(["year", "component_id", "component_name", "value_type",
                  "value_mUSD", "useful_life"])

    # ── 22. intangible_assets → intangible_assets ────────────────
    ws_ia = wb.create_sheet("intangible_assets")
    ws_ia.append(["year", "category", "gross_amount_mUSD",
                  "accumulated_amortization_mUSD", "net_amount_mUSD", "useful_life"])

    # ── 23. provisions → provisions_schedule ─────────────────────
    ws_prov = wb.create_sheet("provisions")
    ws_prov.append(["year", "category", "closing_mUSD"])

    # ── 24. associates → associates_schedule ─────────────────────
    ws_assoc = wb.create_sheet("associates")
    ws_assoc.append(["year", "category", "movement", "value_mUSD"])

    # ── 25. macro_factors → macro_factors ────────────────────────
    ws_mf = wb.create_sheet("macro_factors")
    ws_mf.append(["factor"] + years)

    # ── 26. operational_drivers → operational_drivers ────────────
    ws_od = wb.create_sheet("operational_drivers")
    ws_od.append(["driver", "unit"] + years)

    # ── 27. balancing_adjustments ────────────────────────────────
    ws_ba = wb.create_sheet("balancing_adjustments")
    ws_ba.append(["year", "statement_type", "metric",
                  "adjustment_value_mUSD", "is_balancing",
                  "balancing_reason", "balancing_category", "original_value_mUSD"])

    # ── 28-31. Справочники ───────────────────────────────────────
    ws_dm = wb.create_sheet("dictionary_metrics")
    ws_dm.append(["canonical_metric", "statement", "description", "accepted_aliases"])

    ws_ddt = wb.create_sheet("dictionary_debt_types")
    ws_ddt.append(["instrument_type", "description", "amortization_default", "is_active"])

    ws_ds = wb.create_sheet("dictionary_segments")
    ws_ds.append(["segment_name", "description", "is_active", "commodity"])

    ws_du = wb.create_sheet("dictionary_units")
    ws_du.append(["unit", "description", "multiplier"])

    wb.save(path)


def _write_stress_scenarios(path: Path, industry: str) -> None:
    pack = "metals_mining" if industry == "metals" else "recession"
    path.write_text(f"""\
scenarios:
  base_stress:
    description: "Base stress scenario"
    macro_shocks: {{}}
    driver_shocks:
      cogs_pct:
        type: pp
        value: 3.0
  rate_spike:
    description: "Interest rate +200bp"
    driver_shocks:
      avg_rate:
        type: pp
        value: 2.0
""")


def _copy_notebook(src: Path, dst: Path, company_id: str, company_name: str) -> None:
    """Копирует ноутбук и подставляет company_id."""
    with open(src) as f:
        nb = json.load(f)

    for cell in nb.get("cells", []):
        new_src = []
        for line in cell.get("source", []):
            line = line.replace("us_steel", company_id)
            line = line.replace("US Steel", company_name)
            new_src.append(line)
        cell["source"] = new_src
        # Clear outputs
        if cell.get("cell_type") == "code":
            cell["outputs"] = []
            cell["execution_count"] = None

    with open(dst, "w") as f:
        json.dump(nb, f, ensure_ascii=False, indent=1)


def _create_placeholder_notebook(dst: Path, name: str, company_id: str) -> None:
    """Создаёт минимальный placeholder ноутбук."""
    nb = {
        "nbformat": 4, "nbformat_minor": 4,
        "metadata": {
            "kernelspec": {"display_name": "Python 3", "language": "python", "name": "python3"},
            "language_info": {"name": "python", "version": "3.11.0"},
        },
        "cells": [
            {
                "cell_type": "markdown", "metadata": {},
                "source": [f"# {name.replace('.ipynb', '')}\n\nКомпания: `{company_id}`"],
            },
            {
                "cell_type": "code", "execution_count": None, "metadata": {}, "outputs": [],
                "source": [
                    "import sys\n",
                    "from pathlib import Path\n",
                    "\n",
                    "# Auto-detect project root\n",
                    "_nb = Path('.').resolve()\n",
                    "for _p in [_nb] + list(_nb.parents):\n",
                    "    if (_p / 'engine').is_dir():\n",
                    "        ROOT = _p\n",
                    "        break\n",
                    "sys.path.insert(0, str(ROOT))\n",
                    f"COMPANY_ID = '{company_id}'\n",
                    "print(f'ROOT: {ROOT}')\n",
                    "print(f'Company: {COMPANY_ID}')\n",
                ],
            },
        ],
    }
    dst.write_text(json.dumps(nb, ensure_ascii=False, indent=1))


def main():
    parser = argparse.ArgumentParser(description="Инициализация новой компании")
    parser.add_argument("company",    help="ID компании (например: test_corp)")
    parser.add_argument("--name",     default=None, help="Полное название")
    parser.add_argument("--industry", default="metals",
                        choices=["metals", "steel", "energy", "consumer", "tech", "financial", "other"])
    parser.add_argument("--currency", default="USD")
    parser.add_argument("--standard", default="US_GAAP", choices=["US_GAAP", "IFRS"])
    parser.add_argument("--force",    action="store_true", help="Перезаписать если существует")
    args = parser.parse_args()

    name = args.name or args.company.replace("_", " ").title()
    init_company(
        company_id=args.company,
        name=name,
        industry=args.industry,
        currency=args.currency,
        standard=args.standard,
        force=args.force,
    )


if __name__ == "__main__":
    main()
