"""
populate_nornickel_excel.py — наполнение nornickel_unified.xlsx из Databook.

Источники:
  - Databook_12m_25_Final.xlsx (IFRS, USD mln, 2009-2025)
  - nornickel_production_results_2025 (операционные итоги 2025)
  - Годовой отчёт 2024 (выверка)
"""
import openpyxl
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
COMPANY_DIR = ROOT / "companies" / "nornickel"
DATABOOK = COMPANY_DIR / "data" / "statements" / "Databook_12m_25_Final.xlsx"
EXCEL_OUT = COMPANY_DIR / "data" / "excel" / "nornickel_unified.xlsx"

# ── Helpers ──────────────────────────────────────────────────────────────

def load_sheet(sheet_name):
    wb = openpyxl.load_workbook(DATABOOK, data_only=True)
    ws = wb[sheet_name]
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append([str(c) if c is not None else None for c in row])
    wb.close()
    return rows

def get_years(header_row):
    years = []
    for i, val in enumerate(header_row):
        try:
            v = int(float(val))
            if 2009 <= v <= 2026:
                years.append((i, v))
        except (ValueError, TypeError):
            pass
    return years

def val_or_none(row, col_idx):
    if col_idx is None or col_idx >= len(row) or row[col_idx] is None:
        return None
    try:
        return float(row[col_idx].replace(",", "").replace(" ", ""))
    except (ValueError, AttributeError):
        return None

def set_cell(ws, row_idx, col_idx, value, font_color=None):
    cell = ws.cell(row=row_idx, column=col_idx)
    if value is not None:
        cell.value = round(value, 12)
    if font_color:
        from openpyxl.styles import Font
        cell.font = Font(color=font_color)

# ── Load Databook ────────────────────────────────────────────────────────

print("Loading Databook...")
is_rows = load_sheet("INCOME STATEMENT")
is_header = is_rows[1]
is_years = get_years(is_header)
is_yr_map = {yr: col for col, yr in is_years}
print(f"  IS: {is_years[0][1]}-{is_years[-1][1]}")

bs_rows = load_sheet("BALANCE ")
bs_header = bs_rows[1]
bs_years = get_years(bs_header)
bs_yr_map = {yr: col for col, yr in bs_years}
print(f"  BS: {bs_years[0][1]}-{bs_years[-1][1]}")

cf_rows = load_sheet("CASH FLOW STATEMENT")
cf_header = cf_rows[1]
cf_years = get_years(cf_header)
cf_yr_map = {yr: col for col, yr in cf_years}
print(f"  CF: {cf_years[0][1]}-{cf_years[-1][1]}")

# EBITDA Calc — has D&A and Impairment separated
ebitda_rows = load_sheet("EBITDA CALCULATION")
ebitda_header = ebitda_rows[1]
ebitda_years = get_years(ebitda_header)
ebitda_yr_map = {yr: col for col, yr in ebitda_years}

# Cost Breakdown
cost_rows = load_sheet("COST BREAKDOWN")
cost_header = cost_rows[1]
cost_years = get_years(cost_header)
cost_yr_map = {yr: col for col, yr in cost_years}

# Production
prod_rows = load_sheet("PRODUCTION DATA")
prod_header = prod_rows[0]
prod_years = get_years(prod_header)
prod_yr_map = {yr: col for col, yr in prod_years}

# CAPEX
capex_rows = load_sheet("CAPEX BREAKDOWN")
capex_header = capex_rows[1]
capex_years = get_years(capex_header)
capex_yr_map = {yr: col for col, yr in capex_years}

# Debt
debt_rows = load_sheet("DEBT AND LIQUIDITY ")
debt_header = debt_rows[1]
debt_years = get_years(debt_header)
debt_yr_map = {yr: col for col, yr in debt_years}

# Ore
ore_rows = load_sheet("ORE OUTPUT")
ore_header = ore_rows[0]
ore_years = get_years(ore_header)
ore_yr_map = {yr: col for col, yr in ore_years}

# ── Open target Excel ────────────────────────────────────────────────────

wb = openpyxl.load_workbook(EXCEL_OUT)
TARGET_YEARS = list(range(2011, 2026))

def find_target_years(ws):
    yr_map = {}
    for col_idx in range(2, ws.max_column + 1):
        val = ws.cell(row=1, column=col_idx).value
        try:
            y = int(val)
            yr_map[y] = col_idx
        except (ValueError, TypeError):
            pass
    return yr_map

def find_metric_row(ws, metric_name):
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == metric_name:
            return r
    return None

def fill_metric(ws, metric_name, row_data, source_yr_map, target_map, font_color=None):
    metric_row = find_metric_row(ws, metric_name)
    if metric_row is None:
        return False
    for year in TARGET_YEARS:
        src_col = source_yr_map.get(year)
        tgt_col = target_map.get(year)
        if src_col is not None and tgt_col is not None:
            v = val_or_none(row_data, src_col)
            set_cell(ws, metric_row, tgt_col, v, font_color)
    return True

# ── 1. Fill history_is ───────────────────────────────────────────────────

print("\nFilling history_is...")
ws_is = wb["history_is"]
is_target_map = find_target_years(ws_is)

# Build label → row_index lookup from Databook IS
is_lookup = {}
for i, row in enumerate(is_rows):
    label = (row[0] or "").strip().lower()
    # Remove footnote numbers
    if label:
        is_lookup[label.rstrip(" 1234567890")] = i

# Build EBITDA lookup
ebitda_lookup = {}
for i, row in enumerate(ebitda_rows):
    label = (row[0] or "").strip().lower()
    if label:
        ebitda_lookup[label] = i

# Exact IS mappings from Databook → Excel metric
IS_MAP = {
    "revenue from metal sales": "metal_sales",
    "revenue from other sales": "other_revenue",
    "revenue": "revenue",
    "cost of metal sales": "cogs",
    "gross profit": "gross_profit",
    "general and administrative expenses": "general_and_admin",
    "selling and distribution expenses": "selling_distribution",
    "other operating expenses, net": "other_operating",
    "operating income": "ebit",
    "profit before tax": "earnings_before_tax",
    "income tax expense": "tax_expense",
    "profit for the period": "net_income",
    "shareholders of the parent company": "net_income_parent",
    "non-controlling interests": "net_income_nci",
}

# These come from EBITDA calc sheet
EBITDA_MAP = {
    "depreciation and amortisation": ("dep_ppe", ebitda_rows, ebitda_yr_map),
    "impairment of non-financial assets": ("asset_impairment", ebitda_rows, ebitda_yr_map),
}

# These need special handling
SPECIAL_IS = {
    "finance cost": ("finance_cost_net", is_lookup, is_rows, is_yr_map),
    "foreign exchange": ("fx_gain_loss", is_lookup, is_rows, is_yr_map),
    "income from investments": ("investment_income", is_lookup, is_rows, is_yr_map),
}

# Fill standard IS metrics
for db_key, excel_key in IS_MAP.items():
    if db_key in is_lookup:
        fill_metric(ws_is, excel_key, is_rows[is_lookup[db_key]], is_yr_map, is_target_map, "0066CC")
    else:
        # Partial match
        for rk, ri in is_lookup.items():
            if db_key in rk:
                fill_metric(ws_is, excel_key, is_rows[ri], is_yr_map, is_target_map, "0066CC")
                break
        else:
            print(f"  ⚠ IS not found: {db_key}")

# Fill EBITDA-sourced metrics
for db_key, (excel_key, src_rows, src_map) in EBITDA_MAP.items():
    if db_key in ebitda_lookup:
        fill_metric(ws_is, excel_key, src_rows[ebitda_lookup[db_key]], src_map, is_target_map, "0066CC")
    else:
        for rk, ri in ebitda_lookup.items():
            if db_key in rk:
                fill_metric(ws_is, excel_key, src_rows[ri], src_map, is_target_map, "0066CC")
                break
        else:
            print(f"  ⚠ EBITDA not found: {db_key}")

# Special: finance cost (contains "finance costs, net 3")
for rk, ri in is_lookup.items():
    if "finance cost" in rk:
        fill_metric(ws_is, "finance_cost_net", is_rows[ri], is_yr_map, is_target_map, "0066CC")
        break

# Special: FX
for rk, ri in is_lookup.items():
    if "foreign exchange" in rk:
        fill_metric(ws_is, "fx_gain_loss", is_rows[ri], is_yr_map, is_target_map, "0066CC")
        break

# Special: investment income
for rk, ri in is_lookup.items():
    if "income from investments" in rk:
        fill_metric(ws_is, "investment_income", is_rows[ri], is_yr_map, is_target_map, "0066CC")
        break

# EBITDA = OpIncome + D&A + Impairment (from EBITDA calc)
op_row = is_rows[is_lookup.get("operating income", 0)] if "operating income" in is_lookup else None
da_row = ebitda_rows[ebitda_lookup.get("depreciation and amortisation", 0)] if "depreciation and amortisation" in ebitda_lookup else None
imp_row = ebitda_rows[ebitda_lookup.get("impairment of non-financial assets", 0)] if "impairment of non-financial assets" in ebitda_lookup else None

ebitda_row = find_metric_row(ws_is, "ebitda")
if ebitda_row and op_row and da_row and imp_row:
    for year in TARGET_YEARS:
        src_col = is_yr_map.get(year)
        tgt_col = is_target_map.get(year)
        if src_col is not None and tgt_col is not None:
            op = val_or_none(op_row, src_col) or 0
            da = val_or_none(da_row, is_yr_map.get(year)) or 0  # EBITDA calc uses same column offsets
            imp = val_or_none(imp_row, is_yr_map.get(year)) or 0
            set_cell(ws_is, ebitda_row, tgt_col, op + da + imp, "0066CC")

print("  ✓ history_is filled")

# ── 2. Fill history_bs ───────────────────────────────────────────────────

print("\nFilling history_bs...")
ws_bs = wb["history_bs"]
bs_target_map = find_target_years(ws_bs)

bs_lookup = {}
for i, row in enumerate(bs_rows):
    label = (row[0] or "").strip().lower()
    if label:
        bs_lookup[label.rstrip(" 1234567890")] = i

BS_MAP = {
    "property, plant and equipment": "ppe_net",
    "intangible assets": "intangibles",
    "investments in associates and joint ventures": "investments_lt",
    "deferred tax assets": "dta",
    "other financial assets": "other_nca_financial",
    "other non-current assets": "other_nca_nonfinancial",
    "non-current assets": "total_nca",
    "inventories": "inventory",
    "trade and other receivables": "accounts_receivable",
    "advances paid and prepaid expenses": "prepaid_expenses",
    "income tax receivable": "current_tax_asset",
    "other taxes receivable": "other_tax_receivable",
    "other current assets": "other_current_assets",
    "cash and cash equivalents": "cash",
    "current assets": "total_ca",
    "total assets": "total_assets",
    "trade and other payables": "accounts_payable",
    "dividends payable": "dividend_payable",
    "income tax payable": "current_tax_liability",
    "other taxes payable": "social_taxes_payable",
    "employee benefit obligations": "employee_benefits",
    "current liabilities": "total_cl",
    "non-current liabilities": "total_ncl",
    "total liabilities": "total_liabilities",
    "share capital": "share_capital",
    "share premium": "apic",
    "translation and other reserves": "aoci",
    "retained earnings": "retained_earnings",
    "equity attributable to shareholders of the parent company": "equity_parent",
    "non-controlling interests": "nci",
    "total equity and liabilities": "total_liab_equity",
}

# Compute total_equity = equity_parent + nci
if "equity attributable to shareholders of the parent company" in bs_lookup and "non-controlling interests" in bs_lookup:
    eqp_row = bs_rows[bs_lookup["equity attributable to shareholders of the parent company"]]
    nci_row = bs_rows[bs_lookup["non-controlling interests"]]
    eq_total_row = find_metric_row(ws_bs, "total_equity")
    if eq_total_row:
        for year in TARGET_YEARS:
            sc = bs_yr_map.get(year)
            tc = bs_target_map.get(year)
            if sc is not None and tc is not None:
                eqp = val_or_none(eqp_row, sc) or 0
                nci = val_or_none(nci_row, sc) or 0
                set_cell(ws_bs, eq_total_row, tc, eqp + nci, "0066CC")

# Fill extra BS metrics not in main dict
for db_key, excel_key in {"provisions": "non_current_provisions", "deferred tax liabilities": "dtl"}.items():
    ri = bs_lookup.get(db_key)
    if ri is None:
        for rk, rv in bs_lookup.items():
            if db_key in rk:
                ri = rv
                break
    if ri is not None:
        fill_metric(ws_bs, excel_key, bs_rows[ri], bs_yr_map, bs_target_map, "0066CC")

for db_key, excel_key in BS_MAP.items():
    if db_key in bs_lookup:
        fill_metric(ws_bs, excel_key, bs_rows[bs_lookup[db_key]], bs_yr_map, bs_target_map, "0066CC")
    else:
        found = False
        for rk, ri in bs_lookup.items():
            if db_key in rk:
                fill_metric(ws_bs, excel_key, bs_rows[ri], bs_yr_map, bs_target_map, "0066CC")
                found = True
                break
        if not found:
            print(f"  ⚠ BS not found: {db_key}")

# ST/LT Debt — Databook lists Non-current first, then Current
st_found = False
for i, row in enumerate(bs_rows):
    label = (row[0] or "").strip().lower()
    if "loans and borrowings" in label:
        if not st_found:
            fill_metric(ws_bs, "long_term_debt", bs_rows[i], bs_yr_map, bs_target_map, "0066CC")
            st_found = True
        else:
            fill_metric(ws_bs, "short_term_debt", bs_rows[i], bs_yr_map, bs_target_map, "0066CC")
            break

# Social liabilities
for rk, ri in bs_lookup.items():
    if "social liabilities" in rk:
        if "non-current" in rk or "long" in rk:
            fill_metric(ws_bs, "social_liabilities_noncurrent", bs_rows[ri], bs_yr_map, bs_target_map, "0066CC")
        else:
            # Current social liabilities 
            fill_metric(ws_bs, "social_liabilities_current", bs_rows[ri], bs_yr_map, bs_target_map, "0066CC")

# Lease liabilities
for rk, ri in bs_lookup.items():
    if "lease liabilities" in rk:
        if "non-current" in rk or "long" in rk:
            fill_metric(ws_bs, "lease_liab_noncurrent", bs_rows[ri], bs_yr_map, bs_target_map, "0066CC")
        else:
            fill_metric(ws_bs, "lease_liab_current", bs_rows[ri], bs_yr_map, bs_target_map, "0066CC")

# Current provisions
for rk, ri in bs_lookup.items():
    if "provisions" in rk and "non-current" not in rk:
        fill_metric(ws_bs, "current_provisions", bs_rows[ri], bs_yr_map, bs_target_map, "0066CC")
        break

print("  ✓ history_bs filled")

# ── 3. Fill history_cf ───────────────────────────────────────────────────

print("\nFilling history_cf...")
ws_cf = wb["history_cf"]
cf_target_map = find_target_years(ws_cf)

cf_lookup = {}
for i, row in enumerate(cf_rows):
    label = (row[0] or "").strip().lower()
    if label:
        cf_lookup[label.rstrip(" 1234567890")] = i

CF_MAP = {
    "profit before tax": "net_income",
    "depreciation and amortisation": "dep_amort",
    "loss on disposal of property, plant and equipment": "disposal_nca_adj",
    "income tax paid": "income_tax_paid",
    "net cash generated from operating activities": "cfo_total",
    "purchase of property, plant and equipment": "capex",
    "purchase of intangible assets": "capex_intangibles",
    "net cash used in investing activities": "cfi_total",
    "proceeds from loans and borrowings": "debt_proceeds",
    "repayments of loans and borrowings": "debt_repayment",
    "interest paid": "interest_paid",
    "dividends paid to shareholders of the parent company": "dividends_paid",
    "payments of lease liabilities": "lease_payments",
    "net cash used in financing activities": "cff_total",
    "net change in cash and cash equivalents": "net_change_in_cash",
}

for db_key, excel_key in CF_MAP.items():
    if db_key in cf_lookup:
        fill_metric(ws_cf, excel_key, cf_rows[cf_lookup[db_key]], cf_yr_map, cf_target_map, "0066CC")
    else:
        found = False
        for rk, ri in cf_lookup.items():
            if db_key in rk:
                fill_metric(ws_cf, excel_key, cf_rows[ri], cf_yr_map, cf_target_map, "0066CC")
                found = True
                break
        if not found:
            print(f"  ⚠ CF not found: {db_key}")

# Cash opening/closing
for rk, ri in cf_lookup.items():
    if "cash and cash equivalents at the beginning" in rk:
        fill_metric(ws_cf, "cash_opening", cf_rows[ri], cf_yr_map, cf_target_map, "0066CC")
    if "cash and cash equivalents at the end" in rk:
        fill_metric(ws_cf, "cash_closing", cf_rows[ri], cf_yr_map, cf_target_map, "0066CC")
    if "effects of foreign exchange differences" in rk:
        fill_metric(ws_cf, "fx_effect_on_cash", cf_rows[ri], cf_yr_map, cf_target_map, "0066CC")
    if "finance costs, net" in rk:
        fill_metric(ws_cf, "finance_cost_adj", cf_rows[ri], cf_yr_map, cf_target_map, "0066CC")
    if "impairment of non-financial" in rk:
        fill_metric(ws_cf, "impairment_adj", cf_rows[ri], cf_yr_map, cf_target_map, "0066CC")
    if "income from investments" in rk:
        fill_metric(ws_cf, "investment_income_adj", cf_rows[ri], cf_yr_map, cf_target_map, "0066CC")
    if "change in provisions and allowances" in rk:
        fill_metric(ws_cf, "provisions_accrued", cf_rows[ri], cf_yr_map, cf_target_map, "0066CC")

# Also fill NCI dividends
for rk, ri in cf_lookup.items():
    if "dividends paid to non-controlling" in rk:
        fill_metric(ws_cf, "dividends_paid_nci", cf_rows[ri], cf_yr_map, cf_target_map, "0066CC")

print("  ✓ history_cf filled")

# ── 4. Fill segments ──────────────────────────────────────────────────────

print("\nFilling segments...")
ws_seg = wb["segments"]
seg_target_map = find_target_years(ws_seg)

# Build IS index
is_idx = {}
for i, row in enumerate(is_rows):
    label = (row[0] or "").strip().lower()
    if label:
        is_idx[label] = i

# Build prod index
prod_idx = {}
for i, row in enumerate(prod_rows):
    label = (row[0] or "").strip().lower()
    if label:
        prod_idx[label] = i

# Map segment rows
seg_rows = {}
for r in range(2, ws_seg.max_row + 1):
    seg = ws_seg.cell(row=r, column=1).value
    metric = ws_seg.cell(row=r, column=2).value
    if seg and metric:
        seg_rows[f"{seg}_{metric}"] = r

# Revenue per metal
METAL_REV = {
    "nickel 2": "nickel",
    "copper 2": "copper",
    "palladium 2": "palladium",
    "platinum 2": "platinum",
    "other metals 2": "other_metals",
}

for db_key, seg_name in METAL_REV.items():
    if db_key in is_idx:
        row_data = is_rows[is_idx[db_key]]
        for suffix in ["metal_sales_rev_m_usd", "revenue", "rev_m_usd"]:
            seg_key = f"{seg_name}_{suffix}"
            if seg_key in seg_rows:
                r = seg_rows[seg_key]
                for year in TARGET_YEARS:
                    sc = is_yr_map.get(year)
                    tc = seg_target_map.get(year)
                    if sc is not None and tc is not None:
                        set_cell(ws_seg, r, tc, val_or_none(row_data, sc), "0066CC")
                break

# Production volumes per metal
METAL_PROD = {
    "nickel, tonnes": "nickel",
    "copper, tonnes": "copper",
    "palladium, thousand troy ounces": "palladium",
    "platinum, thousand troy ounces": "platinum",
}

for db_key, seg_name in METAL_PROD.items():
    if db_key in prod_idx:
        row_data = prod_rows[prod_idx[db_key]]
        for suffix in ["metal_sales_vol_kt", "production_kt", "prod_kt"]:
            seg_key = f"{seg_name}_{suffix}"
            if seg_key in seg_rows:
                r = seg_rows[seg_key]
                for year in TARGET_YEARS:
                    sc = prod_yr_map.get(year)
                    tc = seg_target_map.get(year)
                    if sc is not None and tc is not None:
                        v = val_or_none(row_data, sc)
                        if v is not None and seg_name in ("nickel", "copper"):
                            v = v / 1000  # tonnes → kt
                        set_cell(ws_seg, r, tc, v, "0066CC")
                break

print("  ✓ segments filled")

# ── 5. Fill macro_factors (realized prices = revenue ÷ volume) ──────────

print("\nFilling macro_factors...")
ws_macro = wb["macro_factors"]
macro_target_map = find_target_years(ws_macro)

# Realized price from revenue/volume
for db_rev_key, db_vol_key, macro_name, to_unit in [
    ("nickel 2", "nickel, tonnes", "lme_nickel", 1),      # rev in $M, vol in t → $/t
    ("copper 2", "copper, tonnes", "lme_copper", 1),       # rev in $M, vol in t → $/t
    ("palladium 2", "palladium, thousand troy ounces", "lme_palladium", 1000),  # rev in $M, vol in koz → $/oz (M/1000 = k → *1000 to get $/oz then /1000?)
    ("platinum 2", "platinum, thousand troy ounces", "lme_platinum", 1000),
]:
    if db_rev_key in is_idx and db_vol_key in prod_idx:
        rev_data = is_rows[is_idx[db_rev_key]]
        vol_data = prod_rows[prod_idx[db_vol_key]]
        r = find_metric_row(ws_macro, macro_name)
        if r:
            for year in TARGET_YEARS:
                sc = is_yr_map.get(year)
                tc = macro_target_map.get(year)
                if sc is not None and tc is not None:
                    rev = val_or_none(rev_data, sc)  # $M
                    vol = val_or_none(vol_data, is_yr_map.get(year))  # tonnes or koz
                    if rev and vol and vol != 0:
                        price = (rev * 1_000_000) / (vol * to_unit)  # → $/t or $/oz
                        set_cell(ws_macro, r, tc, price, "0066CC")

print("  ✓ macro_factors: realized prices")

# ── 6. Fill operational_drivers ──────────────────────────────────────────

print("\nFilling operational_drivers...")
ws_drv = wb["operational_drivers"]
drv_target_map = find_target_years(ws_drv)

# Clear existing placeholder rows and rebuild
# Keep header, add new rows
DRIVERS = [
    ("nickel_production_kt", prod_idx.get("nickel, tonnes"), prod_rows, prod_yr_map, 1000),
    ("copper_production_kt", prod_idx.get("copper, tonnes"), prod_rows, prod_yr_map, 1000),
    ("palladium_production_koz", prod_idx.get("palladium, thousand troy ounces"), prod_rows, prod_yr_map, 1),
    ("platinum_production_koz", prod_idx.get("platinum, thousand troy ounces"), prod_rows, prod_yr_map, 1),
    ("ore_output_polar_kt", 1, ore_rows, ore_yr_map, 1),  # NORILSK SITE row 9 (0-based: 8)
    ("total_ore_output_kt", 1, ore_rows, ore_yr_map, 1),   # row 1 (0-based: 1)
    ("total_employees", None, None, None, 1),
    ("capex_total_m_usd", cf_lookup.get("purchase of property, plant and equipment"), cf_rows, cf_yr_map, 1),
]

# Find ore rows
ore_lookup = {}
for i, row in enumerate(ore_rows):
    label = (row[0] or "").strip().lower()
    if label:
        ore_lookup[label] = i

# Use first TOTAL row
for key in ore_lookup:
    if "total ore output" in key and "sulfide" in key:
        DRIVERS[4] = ("ore_output_polar_kt", ore_lookup[key], ore_rows, ore_yr_map, 1)
        break

# Setup operational_drivers sheet
# Clear old data
for r in range(2, ws_drv.max_row + 1):
    for c in range(1, ws_drv.max_column + 1):
        ws_drv.cell(row=r, column=c).value = None

for idx, (name, src_idx, src_rows, src_map, divisor) in enumerate(DRIVERS):
    r = idx + 2
    ws_drv.cell(row=r, column=1, value=name)
    ws_drv.cell(row=r, column=2, value="")  # unit
    if src_idx is not None and src_rows is not None:
        row_data = src_rows[src_idx]
        for year in TARGET_YEARS:
            sc = src_map.get(year)
            tc = drv_target_map.get(year)
            if sc is not None and tc is not None:
                v = val_or_none(row_data, sc)
                if v is not None and divisor != 1:
                    v = v / divisor
                set_cell(ws_drv, r, tc, v, "0066CC")

print("  ✓ operational_drivers filled")

# ── 7. Add cost_breakdown sheet ─────────────────────────────────────────

print("\nAdding cost_breakdown sheet...")
if "cost_breakdown" in wb.sheetnames:
    del wb["cost_breakdown"]
ws_cost = wb.create_sheet("cost_breakdown")
ws_cost.cell(row=1, column=1, value="metric")
for i, yr in enumerate(TARGET_YEARS):
    ws_cost.cell(row=1, column=i+2, value=yr)

cost_tgt = {yr: i+2 for i, yr in enumerate(TARGET_YEARS)}

cost_lookup = {}
for i, row in enumerate(cost_rows):
    label = (row[0] or "").strip().lower()
    if label:
        cost_lookup[label] = i

COST_KEYS = [
    "labour",
    "purchases of metals for resale, raw materials and semi-products",
    "materials and supplies",
    "third party services",
    "fuel",
    "electricity and heat energy",
    "mineral extraction tax and other levies",
    "export customs duties",
    "transportation expenses",
    "other costs",
]

for idx, key in enumerate(COST_KEYS):
    r = idx + 2
    ws_cost.cell(row=r, column=1, value=key.title())
    ri = cost_lookup.get(key)
    if ri is None:
        for rk, rv in cost_lookup.items():
            if key in rk:
                ri = rv
                break
    if ri is not None:
        row_data = cost_rows[ri]
        for year in TARGET_YEARS:
            sc = cost_yr_map.get(year)
            tc = cost_tgt.get(year)
            if sc is not None and tc is not None:
                set_cell(ws_cost, r, tc, val_or_none(row_data, sc), "0066CC")

# Add total row and D&A
total_r = len(COST_KEYS) + 2
ws_cost.cell(row=total_r, column=1, value="Total Cash Operating Costs")
if "total cash operating costs" in cost_lookup:
    ri = cost_lookup["total cash operating costs"]
    for year in TARGET_YEARS:
        sc = cost_yr_map.get(year)
        tc = cost_tgt.get(year)
        if sc is not None and tc is not None:
            set_cell(ws_cost, total_r, tc, val_or_none(cost_rows[ri], sc), "0066CC")

da_r = total_r + 1
ws_cost.cell(row=da_r, column=1, value="Depreciation And Amortisation")
if "depreciation and amortisation" in ebitda_lookup:
    for year in TARGET_YEARS:
        sc = ebitda_yr_map.get(year)
        tc = cost_tgt.get(year)
        if sc is not None and tc is not None:
            set_cell(ws_cost, da_r, tc, val_or_none(ebitda_rows[ebitda_lookup["depreciation and amortisation"]], sc), "0066CC")

print("  ✓ cost_breakdown added")

# ── 8. Add capex_breakdown sheet ─────────────────────────────────────────

print("\nAdding capex_breakdown sheet...")
if "capex_breakdown" in wb.sheetnames:
    del wb["capex_breakdown"]
ws_cpx = wb.create_sheet("capex_breakdown")
ws_cpx.cell(row=1, column=1, value="metric")
for i, yr in enumerate(TARGET_YEARS):
    ws_cpx.cell(row=1, column=i+2, value=yr)

cpx_tgt = {yr: i+2 for i, yr in enumerate(TARGET_YEARS)}

cpx_lookup = {}
for i, row in enumerate(capex_rows):
    label = (row[0] or "").strip().lower()
    if label:
        cpx_lookup[label] = i

CPX_KEYS = [
    "polar division",
    "trans-baikal division",
    "energy division",
    "other",
]

for idx, key in enumerate(CPX_KEYS):
    r = idx + 2
    ws_cpx.cell(row=r, column=1, value=key.title())
    ri = cpx_lookup.get(key)
    if ri is None:
        for rk, rv in cpx_lookup.items():
            if key in rk:
                ri = rv
                break
    if ri is not None:
        row_data = capex_rows[ri]
        for year in TARGET_YEARS:
            sc = capex_yr_map.get(year)
            tc = cpx_tgt.get(year)
            if sc is not None and tc is not None:
                set_cell(ws_cpx, r, tc, val_or_none(row_data, sc), "0066CC")

# Total CAPEX from CF
total_cpx_r = len(CPX_KEYS) + 2
ws_cpx.cell(row=total_cpx_r, column=1, value="Total Capex")
if "purchase of property, plant and equipment" in cf_lookup:
    for year in TARGET_YEARS:
        sc = cf_yr_map.get(year)
        tc = cpx_tgt.get(year)
        if sc is not None and tc is not None:
            set_cell(ws_cpx, total_cpx_r, tc, val_or_none(cf_rows[cf_lookup["purchase of property, plant and equipment"]], sc), "0066CC")

print("  ✓ capex_breakdown added")

# ── 9. Add production_data sheet ─────────────────────────────────────────

print("\nAdding production_data sheet...")
if "production_data" in wb.sheetnames:
    del wb["production_data"]
ws_prod = wb.create_sheet("production_data")
ws_prod.cell(row=1, column=1, value="metric")
for i, yr in enumerate(TARGET_YEARS):
    ws_prod.cell(row=1, column=i+2, value=yr)

prod_tgt = {yr: i+2 for i, yr in enumerate(TARGET_YEARS)}

PROD_DATA_KEYS = [
    "nickel, tonnes",
    "copper, tonnes",
    "palladium, thousand troy ounces",
    "platinum, thousand troy ounces",
]

for idx, key in enumerate(PROD_DATA_KEYS):
    r = idx + 2
    ws_prod.cell(row=r, column=1, value=key.title())
    if key in prod_idx:
        row_data = prod_rows[prod_idx[key]]
        for year in TARGET_YEARS:
            sc = prod_yr_map.get(year)
            tc = prod_tgt.get(year)
            if sc is not None and tc is not None:
                set_cell(ws_prod, r, tc, val_or_none(row_data, sc), "0066CC")

# Add 2025 production data from PDF (operational results)
# From PDF: Ni 199kt, Cu 425kt, Pd 2725koz, Pt 667koz
prod_2025 = {"nickel, tonnes": 198521, "copper, tonnes": 425309, "palladium, thousand troy ounces": 2725, "platinum, thousand troy ounces": 667}
for idx, key in enumerate(PROD_DATA_KEYS):
    r = idx + 2
    tc = prod_tgt.get(2025)
    if tc and key in prod_2025:
        set_cell(ws_prod, r, tc, prod_2025[key], "008800")

print("  ✓ production_data added")

# ── Save ──────────────────────────────────────────────────────────────────

print(f"\nSaving to {EXCEL_OUT}...")
wb.save(EXCEL_OUT)
print("✅ Done! Excel populated with Databook data (2011-2025).")
