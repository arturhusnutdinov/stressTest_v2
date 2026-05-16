"""
Excel-загрузчик для шаблона template_UNIFIED_All_Data.xlsx.
Читает все листы, применяет маппинг, записывает в БД через Repository.
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, Tuple

import openpyxl

from ..database.repository import Repository
from .base import (
    BaseLoader, FormulaEngine, LoadResult, MappingConfig,
    RowMapping, SheetMapping, unit_scale_factor,
)

logger = logging.getLogger(__name__)


class ExcelLoader(BaseLoader):
    """
    Загружает данные из Excel-шаблона в БД.

    Поддерживаемые листы:
    - Statement (history_is/bs/cf): wide format metric × years
    - Canonical (debt_instruments, ppe_components, lease_schedule, ...): 1-к-1 с БД
    - Служебные (meta, dictionary_*): читаются, но не загружаются в данные
    """

    # Листы с данными типа Statement (wide: metric × years)
    STATEMENT_SHEETS = {
        "history_is": "IS",
        "history_bs": "BS",
        "history_cf": "CF",
    }

    # Canonical листы и соответствующие методы загрузки
    CANONICAL_SHEETS = {
        "debt_instruments",
        "debt_cashflows",
        "ppe_components",
        "intangible_assets",
        "lease_schedule",
        "sched_lease_finance",
        "sched_lease_operating",
        "schedule_equity",
        "schedule_tax",
        "macro_factors",
        "segments_financial",
        "segments_operational",
        "Production_KPI",
    }

    def __init__(
        self,
        company_id: str,
        repo: Repository,
        mapping_config: Optional[MappingConfig] = None,
        db_unit: str = "tUSD",
        input_default_unit: str = "tUSD",
    ) -> None:
        super().__init__(company_id, db_unit, input_default_unit)
        self._repo = repo
        self._mapping = mapping_config
        self._formula_engine = FormulaEngine()

    # ── публичный API ──────────────────────────────────────────────────────────

    def load(self, excel_path: Path) -> LoadResult:
        """
        Загрузить Excel-файл в БД.
        Возвращает LoadResult с детальным отчётом.
        """
        result = LoadResult(
            company_id=self.company_id,
            source_file=str(excel_path),
        )

        if not excel_path.exists():
            result.errors.append(f"Файл не найден: {excel_path}")
            return result

        try:
            wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
        except Exception as e:
            result.errors.append(f"Ошибка открытия файла: {e}")
            return result

        try:
            # 1. Читаем meta
            meta = self._read_meta(wb)
            if meta:
                self._apply_meta(meta, result)

            # 2. Statement листы (IS / BS / CF)
            for sheet_name, statement in self.STATEMENT_SHEETS.items():
                if sheet_name in wb.sheetnames:
                    self._load_statement_sheet(wb[sheet_name], statement, result)

            # 3. Canonical листы
            for sheet_name in self.CANONICAL_SHEETS:
                if sheet_name in wb.sheetnames:
                    self._load_canonical_sheet(wb[sheet_name], sheet_name, result)

        except Exception as e:
            result.errors.append(f"Критическая ошибка при загрузке: {e}")
            logger.exception(f"Ошибка загрузки {excel_path}")
        finally:
            wb.close()

        # Аудит
        self._repo.log(
            operation="LOAD",
            table_name="excel",
            company_id=self.company_id,
            record_id=excel_path.name,
            details={
                "rows_written": result.rows_written,
                "rows_skipped": result.rows_skipped,
                "errors": len(result.errors),
                "warnings": len(result.warnings),
            },
        )

        return result

    # ── meta ───────────────────────────────────────────────────────────────────

    def _read_meta(self, wb: openpyxl.Workbook) -> Dict[str, Any]:
        if "meta" not in wb.sheetnames:
            return {}
        ws = wb["meta"]
        meta = {}
        for row in ws.iter_rows(values_only=True):
            if row[0] and row[1]:
                meta[str(row[0]).strip()] = row[1]
        return meta

    def _apply_meta(self, meta: Dict[str, Any], result: LoadResult) -> None:
        """Обновить company из meta-листа."""
        company_id = str(meta.get("company_code", self.company_id)).strip()
        name       = str(meta.get("company_name", company_id)).strip()
        currency   = str(meta.get("base_currency", "USD")).strip()
        db_unit    = str(meta.get("db_unit", self.db_unit)).strip()

        self._repo.upsert_company(
            company_id=company_id,
            name=name,
            currency=currency,
            db_unit=db_unit,
        )
        # Обновить db_unit если он задан в meta
        if db_unit:
            self.db_unit = db_unit

    # ── statement sheets (IS / BS / CF) ───────────────────────────────────────

    def _load_statement_sheet(
        self,
        ws,
        statement: str,
        result: LoadResult,
    ) -> None:
        """
        Загружает лист wide-format: metric | 2010 | 2011 | ... | unit | source
        или новый формат: label | db_metric | formula | sign | unit | 2010 | ...
        """
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return

        # Пропускаем строки-заголовки (комментарий + пустая + заголовок)
        header_row = None
        for i, row in enumerate(rows):
            if row and row[0] and str(row[0]).strip().lower() in ("metric", "label"):
                header_row = i
                break

        if header_row is None:
            result.warnings.append(f"Не найден заголовок в листе {statement}")
            return

        header = [str(c).strip() if c is not None else "" for c in rows[header_row]]
        data_rows = rows[header_row + 1:]

        # Определяем формат листа
        if "db_metric" in header:
            self._load_statement_new_format(data_rows, header, statement, result)
        else:
            self._load_statement_legacy_format(data_rows, header, statement, result)

    def _load_statement_new_format(
        self,
        rows: List[Tuple],
        header: List[str],
        statement: str,
        result: LoadResult,
    ) -> None:
        """
        Новый формат: label | db_metric | formula | sign | unit | 2010..N
        """
        def col(name: str) -> Optional[int]:
            try:
                return header.index(name)
            except ValueError:
                return None

        i_label     = col("label")
        i_db_metric = col("db_metric")
        i_formula   = col("formula")
        i_sign      = col("sign")
        i_unit      = col("unit")

        # Найти колонки с годами
        year_cols: List[Tuple[int, int]] = []
        for idx, h in enumerate(header):
            try:
                year = int(h)
                if 1990 <= year <= 2100:
                    year_cols.append((idx, year))
            except (ValueError, TypeError):
                pass

        if not year_cols:
            result.warnings.append(f"Нет колонок с годами в {statement}-листе")
            return

        # Парсим строки маппинга
        raw_mappings: List[RowMapping] = []
        for row in rows:
            if not row or all(v is None for v in row):
                continue
            label     = str(row[i_label]).strip() if i_label is not None and i_label < len(row) and row[i_label] else ""
            db_metric = str(row[i_db_metric]).strip() if i_db_metric is not None and i_db_metric < len(row) and row[i_db_metric] else None
            formula   = str(row[i_formula]).strip() if i_formula is not None and i_formula < len(row) and row[i_formula] else None
            sign      = float(row[i_sign]) if i_sign is not None and i_sign < len(row) and row[i_sign] else 1.0
            unit      = str(row[i_unit]).strip() if i_unit is not None and i_unit < len(row) and row[i_unit] else None

            if not db_metric:
                if label:
                    result.unmapped_labels.append(label)
                continue

            raw_mappings.append(RowMapping(
                label=label,
                db_metric=db_metric,
                formula=formula or None,
                sign=sign,
                unit=unit,
            ))

        if self._mapping and statement in (sm.statement for sm in self._mapping.sheets.values()):
            raw_mappings = self._apply_config_overrides(raw_mappings, statement)

        try:
            sorted_mappings = self.sort_mappings(raw_mappings)
        except ValueError as e:
            result.errors.append(f"Цикл в формулах {statement}: {e}")
            return

        # Загружаем данные по годам
        year_data: Dict[int, Dict[str, float]] = {}

        for row in rows:
            if not row or all(v is None for v in row):
                continue
            db_metric = str(row[i_db_metric]).strip() if i_db_metric is not None and i_db_metric < len(row) and row[i_db_metric] else None
            if not db_metric:
                continue
            sign = float(row[i_sign]) if i_sign is not None and i_sign < len(row) and row[i_sign] else 1.0
            unit = str(row[i_unit]).strip() if i_unit is not None and i_unit < len(row) and row[i_unit] else None

            for col_idx, year in year_cols:
                raw_val = row[col_idx] if col_idx < len(row) else None
                val = self.convert_value(raw_val, unit)
                if val is None:
                    continue
                if year not in year_data:
                    year_data[year] = {}
                year_data[year][db_metric] = val * sign

        # Вычисляем формульные метрики
        formula_mappings = [m for m in sorted_mappings if m.formula]
        for m in formula_mappings:
            for year, ctx in year_data.items():
                val = self._formula_engine.evaluate(m.formula, ctx, year)
                if val is not None:
                    year_data[year][m.db_metric] = val * m.sign

        self._write_history(year_data, statement, result)

    def _load_statement_legacy_format(
        self,
        rows: List[Tuple],
        header: List[str],
        statement: str,
        result: LoadResult,
    ) -> None:
        """
        Старый wide-формат: metric | 2010 | 2011 | ... | unit | source
        Поддерживается для обратной совместимости.
        """
        year_cols: List[Tuple[int, int]] = []
        for idx, h in enumerate(header):
            try:
                year = int(h)
                if 1990 <= year <= 2100:
                    year_cols.append((idx, year))
            except (ValueError, TypeError):
                pass

        i_unit = header.index("unit") if "unit" in header else None

        year_data: Dict[int, Dict[str, float]] = {}

        for row in rows:
            if not row or row[0] is None:
                continue
            db_metric = str(row[0]).strip()
            if not db_metric or db_metric.startswith("#"):
                continue

            mapped_metric = self._resolve_metric(db_metric, statement)
            if not mapped_metric:
                result.unmapped_labels.append(db_metric)
                continue

            unit = str(row[i_unit]).strip() if i_unit is not None and i_unit < len(row) and row[i_unit] else None

            for col_idx, year in year_cols:
                raw_val = row[col_idx] if col_idx < len(row) else None
                val = self.convert_value(raw_val, unit)
                if val is None:
                    continue
                if year not in year_data:
                    year_data[year] = {}
                year_data[year][mapped_metric] = val

        self._write_history(year_data, statement, result)

    def _write_history(
        self,
        year_data: Dict[int, Dict[str, float]],
        statement: str,
        result: LoadResult,
    ) -> None:
        """Записать year_data в history_* через Repository."""
        loaded_metrics: Set[str] = set()

        for year, metrics in sorted(year_data.items()):
            if not metrics:
                continue
            try:
                n = self._repo.upsert_history(
                    self.company_id, statement, year, metrics,
                    source="excel_loader",
                )
                result.rows_written += n
                loaded_metrics.update(metrics.keys())
            except Exception as e:
                result.errors.append(f"Ошибка записи {statement} {year}: {e}")

        missing = self.check_coverage(loaded_metrics, statement)
        if missing:
            result.missing_required.extend(missing)
            result.warnings.append(
                f"{statement}: отсутствуют обязательные метрики: {missing}"
            )

    # ── canonical sheets ───────────────────────────────────────────────────────

    def _load_canonical_sheet(
        self,
        ws,
        sheet_name: str,
        result: LoadResult,
    ) -> None:
        """Диспетчер загрузки canonical-листов."""
        handlers = {
            "debt_instruments":      self._load_debt_instruments,
            "debt_cashflows":        self._load_debt_cashflows,
            "ppe_components":        self._load_ppe_components,
            "macro_factors":         self._load_macro_factors,
            "segments_financial":    self._load_segments,
            "segments_operational":  self._load_segments,
            "Production_KPI":        self._load_production_kpi,
        }
        handler = handlers.get(sheet_name)
        if handler:
            try:
                handler(ws, result)
            except Exception as e:
                result.warnings.append(f"Ошибка загрузки листа {sheet_name}: {e}")
                logger.warning(f"Ошибка {sheet_name}: {e}", exc_info=True)
        else:
            logger.debug(f"Лист {sheet_name}: нет обработчика, пропускаем")

    def _load_debt_instruments(self, ws, result: LoadResult) -> None:
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return
        # Найти строку заголовка
        header_row = None
        for i, row in enumerate(rows):
            if row and row[0] and str(row[0]).strip().lower() == "instrument_id":
                header_row = i
                break
        if header_row is None:
            return
        header = [str(c).strip().lower() if c else "" for c in rows[header_row]]

        for row in rows[header_row + 1:]:
            if not row or row[0] is None:
                continue
            rec = {header[i]: row[i] for i in range(min(len(header), len(row)))}
            instrument_id = str(rec.get("instrument_id", "")).strip()
            if not instrument_id or instrument_id.startswith("#"):
                continue
            # Normalize ID: lowercase + spaces→underscores (matches DB convention)
            instrument_id = instrument_id.lower().replace(" ", "_")
            try:
                self._repo.upsert_debt_instrument(
                    company_id=self.company_id,
                    instrument_id=instrument_id,
                    instrument_name=str(rec.get("instrument_name", rec.get("name", ""))),
                    db_type=str(rec.get("instrument_type", rec.get("type", "other"))),
                    currency=str(rec.get("currency", "USD")),
                    opening_balance=self.convert_value(rec.get("opening_balance")) or 0,
                    maturity_date=str(rec.get("maturity_date", "")) or None,
                    interest_rate=_to_float(rec.get("interest_rate", rec.get("rate"))),
                    rate_type=str(rec.get("rate_type", "fixed")),
                    payment_frequency=str(rec.get("payment_frequency", rec.get("payment_freq", "semi_annual"))),
                    amortization_profile=str(rec.get("amortization_profile", rec.get("amortization", "bullet"))),
                    callable_flag=1 if rec.get("callable_flag", rec.get("callable")) else 0,
                )
                result.rows_written += 1
            except Exception as e:
                result.warnings.append(f"debt_instruments строка {instrument_id}: {e}")

    def _load_debt_cashflows(self, ws, result: LoadResult) -> None:
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return
        header_row = None
        for i, row in enumerate(rows):
            if row and row[0] and str(row[0]).strip().lower() == "instrument_id":
                header_row = i
                break
        if header_row is None:
            return
        header = [str(c).strip().lower() if c else "" for c in rows[header_row]]

        batch: List[tuple] = []
        for row in rows[header_row + 1:]:
            if not row or row[0] is None:
                continue
            rec = {header[i]: row[i] for i in range(min(len(header), len(row)))}
            instrument_id  = str(rec.get("instrument_id", "")).strip()
            cashflow_type  = str(rec.get("cashflow_type", "")).strip()
            year           = _to_int(rec.get("year"))
            amount         = self.convert_value(rec.get("amount"))
            if not instrument_id or not cashflow_type or year is None or amount is None:
                continue
            batch.append((
                self.company_id, instrument_id, year,
                rec.get("period"), cashflow_type, amount,
                str(rec.get("currency", "USD")),
                str(rec.get("note", "") or ""),
            ))

        if batch:
            self._repo.conn.executemany(
                """
                INSERT INTO debt_cashflows
                    (company_id, instrument_id, year, period, cashflow_type, amount, currency, note, updated_at)
                VALUES (?,?,?,?,?,?,?,?,CURRENT_TIMESTAMP)
                ON CONFLICT(company_id, instrument_id, year, cashflow_type) DO UPDATE SET
                    amount = excluded.amount, updated_at = CURRENT_TIMESTAMP
                """,
                batch,
            )
            result.rows_written += len(batch)

    def _load_ppe_components(self, ws, result: LoadResult) -> None:
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return
        header_row = None
        for i, row in enumerate(rows):
            if row and row[0] and str(row[0]).strip().lower() == "category":
                header_row = i
                break
        if header_row is None:
            return
        header = [str(c).strip().lower() if c else "" for c in rows[header_row]]

        for row in rows[header_row + 1:]:
            if not row or row[0] is None:
                continue
            rec = {header[i]: row[i] for i in range(min(len(header), len(row)))}
            category   = str(rec.get("category", "")).strip()
            value_type = str(rec.get("value_type", "")).strip()
            year       = _to_int(rec.get("year"))
            value      = self.convert_value(rec.get("value"))
            if not category or not value_type or year is None:
                continue

            component_id = category.lower().replace(" ", "_")
            period_id = self._repo.ensure_period(self.company_id, year)

            self._repo.conn.execute(
                """
                INSERT INTO ppe_components
                    (company_id, period_id, component_id, component_name, value_type, value, useful_life, source, updated_at)
                VALUES (?,?,?,?,?,?,?,?,CURRENT_TIMESTAMP)
                ON CONFLICT(company_id, period_id, component_id, value_type) DO UPDATE SET
                    value = excluded.value, updated_at = CURRENT_TIMESTAMP
                """,
                (self.company_id, period_id, component_id, category,
                 value_type, value, rec.get("useful_life"), "excel_loader"),
            )
            result.rows_written += 1

    def _load_macro_factors(self, ws, result: LoadResult) -> None:
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return
        header_row = None
        for i, row in enumerate(rows):
            if row and row[0] and str(row[0]).strip().lower() == "factor_name":
                header_row = i
                break
        if header_row is None:
            return
        header = [str(c).strip().lower() if c else "" for c in rows[header_row]]

        year_cols = []
        for idx, h in enumerate(header):
            try:
                year = int(h)
                if 1990 <= year <= 2100:
                    year_cols.append((idx, year))
            except (ValueError, TypeError):
                pass

        i_factor = header.index("factor_name") if "factor_name" in header else 0

        macro_data: Dict[str, Dict[int, float]] = {}

        for row in rows[header_row + 1:]:
            if not row or row[0] is None:
                continue
            factor_name = str(row[i_factor]).strip() if i_factor < len(row) else ""
            if not factor_name or factor_name.startswith("#"):
                continue

            for col_idx, year in year_cols:
                raw_val = row[col_idx] if col_idx < len(row) else None
                if raw_val is None or raw_val == "":
                    continue
                try:
                    val = float(raw_val)
                except (TypeError, ValueError):
                    continue
                if factor_name not in macro_data:
                    macro_data[factor_name] = {}
                macro_data[factor_name][year] = val

        if macro_data:
            n = self._repo.upsert_macro_factors(
                data=macro_data,
                scope="global",
                source="excel_loader",
            )
            result.rows_written += n

    # Денежные единицы, для которых нужна конвертация source → db_unit
    _MONETARY_UNITS = {"usd", "tusd", "musd", "busd", "rub", "trub", "mrub",
                       "eur", "teur", "meur", "cny", "tcny", "mcny"}

    def _load_segments(self, ws, result: LoadResult) -> None:
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return
        header_row = None
        for i, row in enumerate(rows):
            if row and row[0] and str(row[0]).strip().lower() == "segment_name":
                header_row = i
                break
        if header_row is None:
            return
        header = [str(c).strip().lower() if c else "" for c in rows[header_row]]

        year_cols = []
        for idx, h in enumerate(header):
            try:
                year = int(h)
                if 1990 <= year <= 2100:
                    year_cols.append((idx, year))
            except (ValueError, TypeError):
                pass

        i_segment = header.index("segment_name") if "segment_name" in header else 0
        i_metric  = header.index("metric") if "metric" in header else 1
        i_unit    = header.index("unit") if "unit" in header else None

        for row in rows[header_row + 1:]:
            if not row or row[0] is None:
                continue
            segment_name = str(row[i_segment]).strip() if i_segment < len(row) else ""
            metric       = str(row[i_metric]).strip() if i_metric < len(row) else ""
            if not segment_name or not metric:
                continue
            segment_id = segment_name.lower().replace(" ", "_")

            # Определяем единицу: денежная → convert_value, иначе → raw float
            row_unit = str(row[i_unit]).strip() if i_unit is not None and i_unit < len(row) and row[i_unit] else None
            is_monetary = row_unit and row_unit.lower() in self._MONETARY_UNITS

            for col_idx, year in year_cols:
                raw_val = row[col_idx] if col_idx < len(row) else None
                if is_monetary:
                    val = self.convert_value(raw_val, row_unit)
                else:
                    # Non-monetary: хранить как есть (kt, USD/t, percent, ...)
                    if raw_val is None or raw_val == "":
                        continue
                    try:
                        val = float(raw_val)
                    except (TypeError, ValueError):
                        continue
                if val is None:
                    continue
                period_id = self._repo.ensure_period(self.company_id, year)
                self._repo.conn.execute(
                    """
                    INSERT INTO segment_data
                        (company_id, period_id, segment_id, segment_name, metric, value, source, updated_at)
                    VALUES (?,?,?,?,?,?,?,CURRENT_TIMESTAMP)
                    ON CONFLICT(company_id, period_id, segment_id, metric) DO UPDATE SET
                        value = excluded.value, updated_at = CURRENT_TIMESTAMP
                    """,
                    (self.company_id, period_id, segment_id, segment_name,
                     metric, val, "excel_loader"),
                )
                result.rows_written += 1

    def _load_production_kpi(self, ws, result: LoadResult) -> None:
        """
        Загружает Production_KPI: wide формат metric | unit | year_cols.
        Пишет в preprocess_metrics с metric_group='production_kpi'.
        """
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return
        header_row = None
        for i, row in enumerate(rows):
            if row and row[0] and str(row[0]).strip().lower() == "metric":
                header_row = i
                break
        if header_row is None:
            return
        header = [str(c).strip().lower() if c else "" for c in rows[header_row]]

        year_cols = []
        for idx, h in enumerate(header):
            try:
                year = int(h)
                if 1990 <= year <= 2100:
                    year_cols.append((idx, year))
            except (ValueError, TypeError):
                pass

        i_metric = 0
        metrics: Dict[str, Dict[int, float]] = {}

        for row in rows[header_row + 1:]:
            if not row or row[0] is None:
                continue
            metric_name = str(row[i_metric]).strip()
            if not metric_name or metric_name.startswith("#"):
                continue

            for col_idx, year in year_cols:
                raw_val = row[col_idx] if col_idx < len(row) else None
                if raw_val is None or raw_val == "":
                    continue
                try:
                    val = float(raw_val)
                except (TypeError, ValueError):
                    continue
                if metric_name not in metrics:
                    metrics[metric_name] = {}
                metrics[metric_name][year] = val

        if metrics:
            n = self._repo.upsert_preprocess(
                company_id=self.company_id,
                metric_group="production_kpi",
                metrics=metrics,
                source="excel_loader",
            )
            result.rows_written += n

    # ── helpers ────────────────────────────────────────────────────────────────

    def _resolve_metric(self, raw_name: str, statement: str) -> Optional[str]:
        """
        Разрешить имя метрики через конфиг маппинга.
        Если конфига нет — возвращает имя как есть (предполагаем что уже канон).
        """
        if not self._mapping:
            return raw_name

        for sm in self._mapping.sheets.values():
            if sm.statement != statement:
                continue
            for m in sm.mappings:
                if m.label.lower() == raw_name.lower():
                    return m.db_metric

        return raw_name

    def _apply_config_overrides(
        self, mappings: List[RowMapping], statement: str
    ) -> List[RowMapping]:
        """Применить override'ы из MappingConfig поверх строк шаблона."""
        if not self._mapping:
            return mappings

        config_by_label: Dict[str, RowMapping] = {}
        for sm in self._mapping.sheets.values():
            if sm.statement == statement:
                for m in sm.mappings:
                    config_by_label[m.label.lower()] = m

        result = []
        for m in mappings:
            override = config_by_label.get(m.label.lower())
            if override:
                result.append(override)
            else:
                result.append(m)
        return result


# ── helpers ────────────────────────────────────────────────────────────────────

def _to_float(val: Any) -> Optional[float]:
    if val is None or val == "":
        return None
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


def _to_int(val: Any) -> Optional[int]:
    f = _to_float(val)
    return int(f) if f is not None else None
