#!/usr/bin/env python3
"""
Parse ALL Nornickel IFRS Excel files and load into PostgreSQL.
Extracts IS, BS, CF from each file, maps to canonical metrics, upserts into stress_v2.historical_data.
"""

import os
import re
import psycopg2
import openpyxl
from datetime import datetime as dt
from collections import defaultdict

BASE = "/Users/arturhusnutdinov/Documents/IT Development/Docker/stressTest_v2/companies/nornickel/data/statements/"
VID = "8b278956-898c-450d-b5a5-b717e72b1774"

FILES = [
    ("2200_nn_ifrs_consolidated_fs_2013_eng_usd_04_04_2014_final.xlsx", [2013, 2012]),
    ("2200_nn_ifrs_consolidated_fs_2014_eng_usd.xlsx", [2014, 2013]),
    ("nn_ifrs_consolidated_fs_2015_eng_usd.xlsx", [2015, 2014]),
    ("ifrs_eng_usd_consolidation_reporting_2016.xlsx", [2016, 2015]),
    ("ifrs_eng_usd_consolidation_reporting_12m2017.xlsx", [2017, 2016]),
    ("12M_2018_IFRS_Consolidation_FS_Eng_USD.xlsx", [2018, 2017]),
    ("12M_2019_IFRS_Consolidation_Eng_USD.xlsx", [2019, 2018, 2017]),
    ("ifrs_eng_usd_consolidation_reporting_12m_2021_final.xlsx", [2021, 2020, 2019]),
    ("rus_usd_ifrs_consolidation_reporting_en_2022.xlsx", [2022, 2021, 2020]),
    ("ifrs_eng_usd_consolidation_reporting_2023.xlsx", [2023, 2022, 2021]),
    ("ifrs_eng_usd_consolidation_reporting_2024.xlsx", [2024, 2023, 2022]),
    ("ifrs_rus_usd_consolidation_reporting_2025.xlsx", [2025, 2024, 2023]),
]


def parse_value(v):
    """Parse a cell value to float."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip().replace('\xa0', ' ').replace('\u2009', '').replace(',', '')
        s = s.strip()
        if s in ('', '–', '-', '—', 'n/a', 'N/A', '−', '–'):
            return None
        is_negative = False
        if s.startswith('(') and s.endswith(')'):
            s = s[1:-1].strip()
            is_negative = True
        elif s.startswith('-'):
            s = s[1:].strip()
            is_negative = True
        s = re.sub(r'(\d)\s+(\d)', r'\1\2', s)
        try:
            val = float(s)
            return -val if is_negative else val
        except ValueError:
            pass
        cleaned = re.sub(r'[^\d.]', '', s)
        if cleaned:
            try:
                val = float(cleaned)
                return -val if is_negative else val
            except:
                pass
    return None


def find_data_columns(rows, expected_years):
    """Find which columns contain data for which years.
    Handles: int years, string years, datetime years, merged header cells."""
    year_cols = {}

    for i, row in enumerate(rows[:8]):
        for j, cell in enumerate(row):
            if cell is None:
                continue

            # datetime cells (e.g. 2013-12-31)
            if isinstance(cell, dt):
                yr = cell.year
                if yr in expected_years and yr not in year_cols:
                    year_cols[yr] = j
                continue

            # int cells matching a year exactly
            if isinstance(cell, int) and cell in expected_years:
                if cell not in year_cols:
                    year_cols[cell] = j
                continue

            # String cells
            if isinstance(cell, str):
                s = cell.replace('\xa0', ' ').strip()
                # Skip long strings (title/description cells that contain years incidentally)
                if len(s) > 40:
                    continue

                # Check if this cell contains multiple years like "2023                    2022"
                # In that case, this column gets the FIRST year, next column gets second
                found_years_in_cell = []
                for yr in expected_years:
                    if str(yr) in s:
                        found_years_in_cell.append(yr)

                if len(found_years_in_cell) == 1:
                    yr = found_years_in_cell[0]
                    if yr not in year_cols:
                        year_cols[yr] = j
                elif len(found_years_in_cell) >= 2:
                    # Multiple years in one header cell — they map to consecutive columns
                    # Sort descending (newest first, which typically is the leftmost)
                    found_years_in_cell.sort(reverse=True)
                    for offset, yr in enumerate(found_years_in_cell):
                        if yr not in year_cols:
                            year_cols[yr] = j + offset

    return year_cols


def find_data_columns_flexible(rows, expected_years):
    """Enhanced version: also scan data rows to verify/correct column mapping.
    Uses first few data rows to find where numeric values actually live."""
    year_cols = find_data_columns(rows, expected_years)

    if len(year_cols) < len(expected_years):
        # Try scanning data rows for numeric patterns
        # Find rows with numeric data
        for i, row in enumerate(rows[3:12], start=3):
            numeric_cols = []
            for j, cell in enumerate(row):
                if isinstance(cell, (int, float)) and abs(cell) > 1:
                    numeric_cols.append(j)
            if len(numeric_cols) >= 2 and len(numeric_cols) <= len(expected_years):
                # Map expected years to these columns (newest year = leftmost numeric col)
                sorted_years = sorted(expected_years, reverse=True)
                for k, yr in enumerate(sorted_years):
                    if k < len(numeric_cols) and yr not in year_cols:
                        year_cols[yr] = numeric_cols[k]
                if len(year_cols) >= len(expected_years):
                    break

    return year_cols


def find_statement_sheets(wb, sheet_names):
    """Scan sheets to find IS, BS, CF sheets by content keywords."""
    is_sheets = []
    bs_sheets = []
    cf_op_sheets = []
    cf_fin_sheets = []

    for sname in sheet_names:
        ws = wb[sname]
        texts = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= 6:
                break
            for v in row:
                if v and isinstance(v, str):
                    texts.append(v[:200].lower())
        combined = ' '.join(texts)

        is_income = any(kw in combined for kw in [
            'income statement', 'profit or loss', 'profit and loss',
            'раскрываемый консолидированный отчет о прибыл',
            'консолидированный отчет о прибыл'
        ])
        is_comprehensive = any(kw in combined for kw in ['comprehensive', 'совокупн'])
        is_bs = any(kw in combined for kw in [
            'financial position', 'balance sheet',
            'финансового положения', 'финансовом положении'
        ])
        is_cf = any(kw in combined for kw in ['cash flow', 'движени денежных', 'денежных средств'])
        is_operating = any(kw in combined for kw in ['operating', 'операцион'])
        is_investing = any(kw in combined for kw in ['investing', 'инвестицион'])
        is_financing = any(kw in combined for kw in ['financing', 'финансов'])

        # Also check: does this sheet have actual data rows (numeric cells)?
        has_data = False
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i < 2 or i >= 10:
                continue
            for cell in row:
                if isinstance(cell, (int, float)) and abs(cell) > 10:
                    has_data = True
                    break
            if has_data:
                break

        sheet_num = int(sname.replace('Table ', ''))

        if is_income and not is_comprehensive and sheet_num <= 15 and has_data:
            is_sheets.append(sname)
        elif is_bs and sheet_num <= 15 and has_data:
            bs_sheets.append(sname)
        elif is_cf and sheet_num <= 20 and has_data:
            if is_financing and not is_operating and not is_investing:
                cf_fin_sheets.append(sname)
            else:
                cf_op_sheets.append(sname)

    return is_sheets, bs_sheets, cf_op_sheets, cf_fin_sheets


# ── IS metric mappings ──
IS_MAP_ORDERED = [
    ('total revenue', 'total_revenue'),
    ('metal sales', 'revenue_metal'),
    ('revenue metal sales', 'revenue_metal'),
    ('other sales', 'revenue_other'),
    ('cost of metal sales', 'cogs_metal'),
    ('cost of other sales', 'cogs_other'),
    ('gross profit', 'gross_profit'),
    ('selling and distribution', 'distribution_expenses'),
    ('selling, general', 'sga'),
    ('general and administrative', 'admin_expenses'),
    ('impairment of non-financial', 'asset_impairment_nonfin'),
    ('impairment losses', 'asset_impairment'),
    ('other operating expenses', 'other_operating_expenses'),
    ('other operating income', 'other_operating_income'),
    ('operating profit', 'ebit'),
    ('operating income', 'ebit'),
    ('profit from operations', 'ebit'),
    ('profit from operating activities', 'ebit'),
    ('foreign exchange', 'fx_gain_loss'),
    ('finance costs', 'interest_expense'),
    ('finance cost', 'interest_expense'),
    ('finance income', 'interest_income'),
    ('finance expenses', 'interest_expense'),
    ('income from investments', 'investment_income'),
    ('income from disposal', 'disposal_gain'),
    ('gain on disposal of subsidiaries', 'disposal_gain'),
    ('loss on disposal of subsidiaries', 'disposal_gain'),
    ('gain/loss from disposal of subsidiaries', 'disposal_gain'),
    ('profit before tax', 'ebt'),
    ('profit before income tax', 'ebt'),
    ('income tax', 'tax_expense'),
    ('profit for the period', 'net_income'),
    ('profit for the year', 'net_income'),
    ('net profit', 'net_income'),
    ('attributable to shareholders of the parent', 'net_income_parent'),
    ('attributable to non-controlling', 'net_income_nci'),
    ('итого выручка', 'total_revenue'),
    ('выручка от реализации металлов', 'revenue_metal'),
    ('выручка от прочей реализации', 'revenue_other'),
    ('себестоимость реализованных металлов', 'cogs_metal'),
    ('себестоимость прочей реализации', 'cogs_other'),
    ('валовая прибыль', 'gross_profit'),
    ('административные расходы', 'admin_expenses'),
    ('коммерческие расходы', 'distribution_expenses'),
    ('убыток от обесценения', 'asset_impairment_nonfin'),
    ('прочие операционные расходы', 'other_operating_expenses'),
    ('прибыль от операционной деятельности', 'ebit'),
    ('курсовые разницы', 'fx_gain_loss'),
    ('финансовые расходы', 'interest_expense'),
    ('доходы от инвестиционной деятельности', 'investment_income'),
    ('прибыль от выбытия дочерних', 'disposal_gain'),
    ('прибыль до налогообложения', 'ebt'),
    ('расходы по налогу на прибыль', 'tax_expense'),
    ('прибыль за год', 'net_income'),
    ('прибыль за период', 'net_income'),
    ('приходящаяся на акционеров материнской', 'net_income_parent'),
    ('приходящаяся на неконтролирующие', 'net_income_nci'),
    ('акционерам материнской компании', 'net_income_parent'),
    ('держателям неконтролирующих долей', 'net_income_nci'),
    ('revenue', 'total_revenue'),
]

BS_MAP_ORDERED = [
    ('property, plant and equipment', 'ppe_net'),
    ('property plant and equipment', 'ppe_net'),
    ('right-of-use assets', 'rou_asset'),
    ('investment property', 'investment_property'),
    ('intangible assets', 'intangibles'),
    ('goodwill', 'goodwill'),
    ('investments in associates and joint ventures', 'investments_in_associates'),
    ('investments in associates', 'investments_in_associates'),
    ('investments in joint ventures', 'investments_in_jv'),
    ('other financial assets', None),
    ('deferred tax assets', 'dta'),
    ('deferred tax asset', 'dta'),
    ('other non-current assets', 'other_nca'),
    ('inventories', 'inventory'),
    ('trade and other receivables', 'accounts_receivable'),
    ('trade receivables', 'accounts_receivable'),
    ('advances paid and prepaid expenses', 'prepaid_expenses'),
    ('advances paid', 'prepaid_expenses'),
    ('prepaid expenses', 'prepaid_expenses'),
    ('income tax prepaid', 'income_tax_prepaid'),
    ('income tax prepayments', 'income_tax_prepaid'),
    ('other taxes receivable', 'other_taxes_receivable'),
    ('cash and cash equivalents', 'cash'),
    ('other current assets', 'other_current_assets'),
    ('total assets', 'total_assets'),
    ('share capital', 'share_capital'),
    ('share premium', 'apic'),
    ('treasury shares', 'treasury_stock'),
    ('translation and other reserves', 'aoci'),
    ('translation reserve', 'aoci'),
    ('retained earnings', 'retained_earnings'),
    ('non-controlling interests', 'nci'),
    ('non-controlling interest', 'nci'),
    ('equity attributable to shareholders', 'equity_parent'),
    ('total equity', 'total_equity'),
    ('loans and borrowings', None),
    ('lease liabilities', None),
    ('provisions', None),
    ('social liabilities', None),
    ('employee benefits', 'employee_benefits'),
    ('employee benefit obligations', 'employee_benefits'),
    ('trade and other payables', 'accounts_payable'),
    ('trade payables', 'accounts_payable'),
    ('dividends payable', 'dividends_payable'),
    ('deferred tax liabilities', 'dtl'),
    ('deferred tax liability', 'dtl'),
    ('income tax payable', 'income_tax_payable'),
    ('other taxes payable', 'other_taxes_payable'),
    ('other current liabilities', 'other_current_liabilities'),
    ('other non-current liabilities', 'other_non_current_liabilities'),
    ('derivative financial instruments', 'derivatives'),
    ('total liabilities', 'total_liabilities'),
    ('total equity and liabilities', 'total_liab_equity'),
    ('total current liabilities', 'total_current_liabilities'),
    ('total non-current liabilities', 'total_non_current_liabilities'),
    ('total current assets', 'total_current_assets'),
    ('total non-current assets', 'total_non_current_assets'),
    # Russian
    ('основные средства', 'ppe_net'),
    ('активы в форме права пользования', 'rou_asset'),
    ('нематериальные активы', 'intangibles'),
    ('гудвилл', 'goodwill'),
    ('инвестиции в ассоциированные', 'investments_in_associates'),
    ('прочие финансовые активы', None),
    ('отложенные налоговые активы', 'dta'),
    ('прочие внеоборотные активы', 'other_nca'),
    ('запасы', 'inventory'),
    ('торговая и прочая дебиторская', 'accounts_receivable'),
    ('авансы выданные и расходы будущих', 'prepaid_expenses'),
    ('авансовые платежи по налогу на прибыль', 'income_tax_prepaid'),
    ('прочие налоги к возмещению', 'other_taxes_receivable'),
    ('денежные средства и их эквиваленты', 'cash'),
    ('прочие оборотные активы', 'other_current_assets'),
    ('итого активы', 'total_assets'),
    ('уставный капитал', 'share_capital'),
    ('эмиссионный доход', 'apic'),
    ('собственные акции', 'treasury_stock'),
    ('резерв накопленных курсовых', 'aoci'),
    ('нераспределенная прибыль', 'retained_earnings'),
    ('неконтролирующие доли', 'nci'),
    ('капитал, причитающийся акционерам', 'equity_parent'),
    ('кредиты и займы', None),
    ('обязательства по аренде', None),
    ('оценочные обязательства', None),
    ('социальные обязательства', None),
    ('обязательства по вознаграждениям', 'employee_benefits'),
    ('торговая и прочая долгосрочная кредиторская', 'lt_accounts_payable'),
    ('торговая и прочая кредиторская', 'accounts_payable'),
    ('дивиденды к уплате', 'dividends_payable'),
    ('отложенные налоговые обязательства', 'dtl'),
    ('обязательства по налогу на прибыль', 'income_tax_payable'),
    ('прочие налоговые обязательства', 'other_taxes_payable'),
    ('прочие долгосрочные обязательства', 'other_non_current_liabilities'),
    ('производные финансовые инструменты', 'derivatives'),
    ('итого обязательства', 'total_liabilities'),
    ('итого капитал и обязательства', 'total_liab_equity'),
]

CF_MAP_ORDERED = [
    ('profit before tax', 'ebt'),
    ('profit before income tax', 'ebt'),
    ('depreciation and amortisation', 'total_da'),
    ('depreciation and amortization', 'total_da'),
    ('impairment of non-financial', 'impairment'),
    ('impairment losses', 'impairment'),
    ('loss on disposal of property', 'loss_on_disposal'),
    ('gain on disposal of subsidiaries', 'disposal_gain_cf'),
    ('loss from disposal of subsidiaries', 'disposal_gain_cf'),
    ('(gain)/loss from disposal of subsidiaries', 'disposal_gain_cf'),
    ('gain/loss from disposal of subsidiaries', 'disposal_gain_cf'),
    ('change in provisions and employee', 'change_provisions'),
    ('change in provisions', 'change_provisions'),
    ('finance costs, net', 'finance_costs_net'),
    ('finance cost', 'finance_costs_net'),
    ('finance expenses', 'finance_costs_net'),
    ('investment income', 'investment_income_cf'),
    ('income from investments', 'investment_income_cf'),
    ('foreign exchange', 'fx_gain_loss'),
    ('share of results of associates', 'share_of_associates'),
    ('share of losses of associates', 'share_of_associates'),
    ('inventories', 'change_inventory'),
    ('trade and other receivables', 'change_ar'),
    ('advances paid', 'change_prepaid'),
    ('other taxes receivable', 'change_other_taxes'),
    ('employee benefit', 'change_employee_benefits'),
    ('trade and other payables', 'change_ap'),
    ('provisions and social', 'change_provisions_social'),
    ('other taxes payable', 'change_taxes_payable'),
    ('cash generated from operations', 'cfo_before_tax'),
    ('income tax paid', 'taxes_paid'),
    ('net cash from operating', 'cfo_total'),
    ('net cash generated from operating', 'cfo_total'),
    ('cash provided by operating', 'cfo_total'),
    ('acquisition of property', 'capex'),
    ('purchase of property', 'capex'),
    ('investments in associates', 'investments_in_associates_cf'),
    ('acquisition of intangible', 'capex_intangibles'),
    ('purchase of intangible', 'capex_intangibles'),
    ('loans granted', 'loans_granted'),
    ('proceeds from repayment of loans', 'loans_repaid'),
    ('change in deposits', 'change_deposits'),
    ('disposal of subsidiaries', 'disposal_subs'),
    ('acquisition of subsidiaries', 'acquisition_subs'),
    ('interest received', 'interest_received'),
    ('other investing', 'other_investing'),
    ('net cash used in investing', 'cfi_total'),
    ('net cash from investing', 'cfi_total'),
    ('proceeds from loans and borrowings', 'debt_issuance'),
    ('proceeds from borrowings', 'debt_issuance'),
    ('repayments of loans and borrowings', 'debt_repayment'),
    ('repayment of borrowings', 'debt_repayment'),
    ('repayment of loans and borrowings', 'debt_repayment'),
    ('payments of lease liabilities', 'lease_payments'),
    ('lease payments', 'lease_payments'),
    ('dividends paid to shareholders of the parent', 'dividends_paid'),
    ('dividends paid to shareholders', 'dividends_paid'),
    ('dividends paid to non-controlling', 'dividends_paid_nci'),
    ('dividends received', 'dividends_received'),
    ('unclaimed dividends received', 'dividends_received_unclaimed'),
    ('cross-currency interest rate swap', 'swap_payments'),
    ('interest paid', 'interest_paid'),
    ('net cash used in financing', 'cff_total'),
    ('net cash from financing', 'cff_total'),
    ('net increase in cash', 'net_change_in_cash'),
    ('net decrease in cash', 'net_change_in_cash'),
    ('net change in cash', 'net_change_in_cash'),
    ('cash and cash equivalents at the beginning', 'cash_opening'),
    ('cash and cash equivalents at beginning', 'cash_opening'),
    ('cash at beginning', 'cash_opening'),
    ('effect of exchange rate', 'fx_effect_on_cash'),
    ('effect of foreign exchange', 'fx_effect_on_cash'),
    ('cash and cash equivalents at the end', 'cash_closing'),
    ('cash and cash equivalents at end', 'cash_closing'),
    ('cash at end', 'cash_closing'),
    # Russian
    ('прибыль до налогообложения', 'ebt'),
    ('износ и амортизация', 'total_da'),
    ('убыток от обесценения нефинансовых', 'impairment'),
    ('убыток от обесценения', 'impairment'),
    ('убыток от выбытия основных средств', 'loss_on_disposal'),
    ('прибыль от выбытия дочерних', 'disposal_gain_cf'),
    ('изменение оценочных', 'change_provisions'),
    ('финансовые расходы', 'finance_costs_net'),
    ('доходы от инвестиционной деятельности', 'investment_income_cf'),
    ('курсовые разницы', 'fx_gain_loss'),
    ('доля в убытках ассоциированных', 'share_of_associates'),
    ('прочие', 'other_operating_cf'),
    ('запасы', 'change_inventory'),
    ('торговая и прочая дебиторская задолженность', 'change_ar'),
    ('авансы выданные и расходы будущих', 'change_prepaid'),
    ('прочие налоги к возмещению', 'change_other_taxes'),
    ('обязательства по вознаграждениям', 'change_employee_benefits'),
    ('торговая и прочая кредиторская задолженность', 'change_ap'),
    ('оценочные и социальные обязательства', 'change_provisions_social'),
    ('прочие налоговые обязательства', 'change_taxes_payable'),
    ('операционной деятельности, нетто', 'cfo_total'),
    ('операционной деятельности', 'cfo_before_tax'),
    ('налог на прибыль уплаченный', 'taxes_paid'),
    ('приобретение основных средств', 'capex'),
    ('инвестиции в ассоциированные организации', 'investments_in_associates_cf'),
    ('приобретение нематериальных активов', 'capex_intangibles'),
    ('займы выданные', 'loans_granted'),
    ('поступления от погашения займов', 'loans_repaid'),
    ('изменение величины размещенных депозитов', 'change_deposits'),
    ('выбытие дочерних', 'disposal_subs'),
    ('приобретение дочерних', 'acquisition_subs'),
    ('проценты полученные', 'interest_received'),
    ('прочие инвестиционные денежные потоки', 'other_investing'),
    ('инвестиционную деятельность, нетто', 'cfi_total'),
    ('привлечение кредитов и займов', 'debt_issuance'),
    ('погашение кредитов и займов', 'debt_repayment'),
    ('погашение обязательств по аренде', 'lease_payments'),
    ('дивиденды, выплаченные акционерам', 'dividends_paid'),
    ('дивиденды, выплаченные держателям', 'dividends_paid_nci'),
    ('поступление дивидендов', 'dividends_received'),
    ('восстановление невостребованных дивидендов', 'dividends_received_unclaimed'),
    ('платежи по обмену потоками', 'swap_payments'),
    ('поступления по обмену потоками', 'swap_payments'),
    ('проценты уплаченные', 'interest_paid'),
    ('финансовую деятельность, нетто', 'cff_total'),
    ('изменение денежных средств', 'net_change_in_cash'),
    ('на начало периода', 'cash_opening'),
    ('эффект от курсовых разниц', 'fx_effect_on_cash'),
    ('на конец периода', 'cash_closing'),
]


def match_metric_ordered(label, metric_list):
    """Match label to metric using ordered (pattern, metric) pairs."""
    if not label:
        return None
    label_lower = re.sub(r'\s+', ' ', label.lower().strip())
    for pattern, metric in metric_list:
        if pattern in label_lower:
            return metric
    return None


def read_all_rows(ws):
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append(list(row))
    return rows


def extract_label(row, max_cols=4):
    """Extract text label from row. Concatenate text from first few columns."""
    parts = []
    for j, cell in enumerate(row):
        if j > max_cols:
            break
        if cell and isinstance(cell, str):
            s = cell.strip().replace('\xa0', ' ').replace('\n', ' ')
            s = re.sub(r'\s+', ' ', s)
            if re.match(r'^[\d,.\s()\-–—]+$', s):
                continue
            if len(s) > 3:
                parts.append(s)
    return ' '.join(parts) if parts else None


def get_row_values(row, year_cols):
    """Extract values for each year from a row.
    Smart: try the exact column, but also scan nearby for data if the exact col is empty."""
    result = {}
    for yr, col in year_cols.items():
        val = parse_value(row[col]) if col < len(row) else None
        if val is not None:
            result[yr] = val
        else:
            # For BS with shifted columns: check col-1 and col+1
            for alt in [col - 1, col + 1]:
                if 0 <= alt < len(row):
                    alt_val = parse_value(row[alt])
                    if alt_val is not None:
                        # Only use alt if the exact col truly has nothing and alt isn't already claimed
                        if alt not in year_cols.values():
                            result[yr] = alt_val
                            break
    return result


def parse_is(ws, expected_years, filename):
    """Parse Income Statement."""
    rows = read_all_rows(ws)
    year_cols = find_data_columns_flexible(rows, expected_years)
    if not year_cols:
        print(f"  WARNING: No year columns found in IS for {filename}")
        return {}
    print(f"  IS year_cols: {year_cols}")
    results = {}
    prev_label = None

    for i, row in enumerate(rows):
        if i < 2:
            continue
        label = extract_label(row)
        if not label:
            continue
        label_lower = label.lower()
        if any(kw in label_lower for kw in ['mining and metallurgical', 'горно-металлургич',
                                              'for the year ended', 'за год, закончивш',
                                              'notes', 'примечан', 'приме- чания',
                                              'us dollars', 'долларов сша', 'disclosed']):
            continue

        metric = match_metric_ordered(label, IS_MAP_ORDERED)

        # Handle 'Revenue' as section header
        if metric == 'total_revenue' and label_lower.strip() in ('revenue', 'выручка'):
            if i + 1 < len(rows):
                next_label = extract_label(rows[i + 1])
                if next_label and any(kw in next_label.lower() for kw in ['metal sales', 'реализации металл']):
                    continue

        if metric is None:
            if prev_label:
                combined = prev_label + ' ' + label
                metric = match_metric_ordered(combined, IS_MAP_ORDERED)
            if metric is None:
                prev_label = label
                continue

        prev_label = label
        for yr, col in year_cols.items():
            val = parse_value(row[col]) if col < len(row) else None
            if val is not None:
                results[(yr, metric)] = val

    # Derived
    for yr in year_cols:
        cm = results.get((yr, 'cogs_metal'))
        co = results.get((yr, 'cogs_other'))
        if cm is not None and co is not None:
            results[(yr, 'cogs')] = cm + co
        elif cm is not None:
            results[(yr, 'cogs')] = cm
        tr = results.get((yr, 'total_revenue'))
        if tr is not None:
            results[(yr, 'revenue')] = tr

    return results


def parse_bs(ws, expected_years, filename):
    """Parse Balance Sheet with smart column detection per row."""
    rows = read_all_rows(ws)
    year_cols = find_data_columns_flexible(rows, expected_years)
    if not year_cols:
        print(f"  WARNING: No year columns found in BS for {filename}")
        return {}
    print(f"  BS year_cols: {year_cols}")

    results = {}
    section = None
    prev_label = None

    for i, row in enumerate(rows):
        if i < 2:
            continue
        label = extract_label(row)
        if not label:
            continue
        label_lower = label.lower()
        if any(kw in label_lower for kw in ['mining and metallurgical', 'горно-металлургич',
                                              'at 31 december', 'на 31 декабря', 'notes', 'примечан',
                                              'приме- чания', 'us dollars', 'page', 'disclosed',
                                              'раскрываемый']):
            continue

        # Detect section
        if any(kw in label_lower for kw in ['non-current assets', 'внеоборотные активы']):
            section = 'nca'
            if 'property' not in label_lower and 'основные' not in label_lower:
                continue
        if any(kw in label_lower for kw in ['current assets', 'оборотные активы']) and 'non' not in label_lower and 'внеоборотн' not in label_lower:
            section = 'ca'
            if 'inventories' not in label_lower and 'запасы' not in label_lower:
                continue
        if any(kw in label_lower for kw in ['equity and liabilities', 'капитал и обязательства']):
            section = 'equity'
            continue
        if any(kw in label_lower for kw in ['equity', 'капитал и резервы']) and 'total' not in label_lower and 'итого' not in label_lower:
            section = 'equity'
            continue
        if any(kw in label_lower for kw in ['non-current liabilities', 'долгосрочные обязательства']):
            section = 'ncl'
            continue
        if any(kw in label_lower for kw in ['current liabilities', 'краткосрочные обязательства']) and 'non' not in label_lower and 'долгосрочн' not in label_lower:
            section = 'cl'
            continue
        if label_lower.strip() in ('assets', 'активы'):
            continue

        metric = match_metric_ordered(label, BS_MAP_ORDERED)

        # Context-dependent
        if metric is None:
            if any(kw in label_lower for kw in ['loans and borrowings', 'кредиты и займы']):
                metric = 'long_term_debt' if section == 'ncl' else 'short_term_debt'
            elif any(kw in label_lower for kw in ['lease liabilit', 'обязательства по аренде']):
                metric = 'lease_liab_noncurrent' if section == 'ncl' else 'lease_liab_current'
            elif any(kw in label_lower for kw in ['provisions', 'оценочные обязательства']) and 'change' not in label_lower:
                metric = 'provisions' if section == 'ncl' else 'provisions_current'
            elif any(kw in label_lower for kw in ['social liabilit', 'социальные обязательства']):
                metric = 'social_liabilities' if section == 'ncl' else 'social_liabilities_current'
            elif any(kw in label_lower for kw in ['other financial assets', 'прочие финансовые активы']):
                metric = 'other_financial_assets_lt' if section == 'nca' else 'other_financial_assets_st'

        if metric is None:
            if prev_label:
                combined = prev_label + ' ' + label
                metric = match_metric_ordered(combined, BS_MAP_ORDERED)
                if metric is None and any(kw in combined.lower() for kw in ['investments in associates', 'инвестиции в ассоциированные']):
                    metric = 'investments_in_associates'
            if metric is None:
                prev_label = label
                continue

        prev_label = label

        # Use smart value extraction
        vals = get_row_values(row, year_cols)
        for yr, val in vals.items():
            results[(yr, metric)] = val

    # Derived
    for yr in year_cols:
        ltd = results.get((yr, 'long_term_debt'), 0)
        std = results.get((yr, 'short_term_debt'), 0)
        if ltd or std:
            results[(yr, 'total_debt')] = (ltd or 0) + (std or 0)
        llnc = results.get((yr, 'lease_liab_noncurrent'), 0)
        llc = results.get((yr, 'lease_liab_current'), 0)
        if llnc or llc:
            results[(yr, 'lease_liab_total')] = (llnc or 0) + (llc or 0)
        td = results.get((yr, 'total_debt'), 0)
        lt = results.get((yr, 'lease_liab_total'), 0)
        if td or lt:
            results[(yr, 'total_debt_gross')] = (td or 0) + (lt or 0)
        cash = results.get((yr, 'cash'), 0)
        tdg = results.get((yr, 'total_debt_gross'), results.get((yr, 'total_debt'), 0))
        if tdg:
            results[(yr, 'net_debt')] = (tdg or 0) - (cash or 0)

    return results


def parse_cf(ws_list, expected_years, filename):
    """Parse Cash Flow statement."""
    results = {}
    for ws in ws_list:
        rows = read_all_rows(ws)
        year_cols = find_data_columns_flexible(rows, expected_years)
        if not year_cols:
            continue
        print(f"  CF year_cols ({ws.title}): {year_cols}")

        section = None
        prev_label = None

        for i, row in enumerate(rows):
            if i < 2:
                continue
            label = extract_label(row)
            if not label:
                continue
            label_lower = label.lower()
            if any(kw in label_lower for kw in ['mining and metallurgical', 'горно-металлургич',
                                                  'for the year ended', 'за год, закончивш',
                                                  'us dollars', 'примечания являются',
                                                  'прилагаемые', 'disclosed']):
                continue

            if any(kw in label_lower for kw in ['operating activities', 'операционная деятельность']):
                section = 'operating'
            if any(kw in label_lower for kw in ['investing activities', 'инвестиционная деятельность']):
                section = 'investing'
            if any(kw in label_lower for kw in ['financing activities', 'финансовая деятельность']):
                section = 'financing'

            metric = match_metric_ordered(label, CF_MAP_ORDERED)
            if metric is None:
                if prev_label:
                    combined = prev_label + ' ' + label
                    metric = match_metric_ordered(combined, CF_MAP_ORDERED)
                if metric is None:
                    prev_label = label
                    continue

            prev_label = label

            vals = get_row_values(row, year_cols)
            for yr, val in vals.items():
                if (yr, metric) not in results:
                    results[(yr, metric)] = val

    # Derived
    for yr in expected_years:
        capex = results.get((yr, 'capex'))
        if capex is not None and capex > 0:
            results[(yr, 'capex')] = -capex
        cfo = results.get((yr, 'cfo_total'))
        capex = results.get((yr, 'capex'))
        if cfo is not None and capex is not None:
            results[(yr, 'fcf')] = cfo + capex

    return results


def parse_file(filename, expected_years):
    """Parse a single Excel file."""
    filepath = os.path.join(BASE, filename)
    if not os.path.exists(filepath):
        print(f"  FILE NOT FOUND: {filepath}")
        return {}, {}, {}

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    sheet_names = wb.sheetnames
    is_sheets, bs_sheets, cf_op_sheets, cf_fin_sheets = find_statement_sheets(wb, sheet_names)
    print(f"  Detected sheets: IS={is_sheets}, BS={bs_sheets}, CF_op={cf_op_sheets}, CF_fin={cf_fin_sheets}")

    is_data = {}
    bs_data = {}
    cf_data = {}

    if is_sheets:
        # Use the LAST detected IS sheet (more likely the actual statement, not auditor's page)
        # But prefer lower table numbers if they have data
        is_data = parse_is(wb[is_sheets[-1]], expected_years, filename)
        # If first sheet got more data, use that
        if len(is_sheets) > 1:
            is_data_first = parse_is(wb[is_sheets[0]], expected_years, filename)
            if len(is_data_first) > len(is_data):
                is_data = is_data_first

    if bs_sheets:
        bs_data = parse_bs(wb[bs_sheets[0]], expected_years, filename)

    cf_worksheets = []
    for sname in cf_op_sheets + cf_fin_sheets:
        cf_worksheets.append(wb[sname])
    if cf_worksheets:
        cf_data = parse_cf(cf_worksheets, expected_years, filename)

    wb.close()
    return is_data, bs_data, cf_data


def main():
    master = {}

    for filename, expected_years in FILES:
        print(f"\n{'='*60}")
        print(f"Processing: {filename}")
        print(f"  Expected years: {expected_years}")

        is_data, bs_data, cf_data = parse_file(filename, expected_years)

        is_count = len(is_data)
        bs_count = len(bs_data)
        cf_count = len(cf_data)
        print(f"  IS: {is_count} datapoints, BS: {bs_count} datapoints, CF: {cf_count} datapoints")

        is_years = sorted(set(yr for yr, _ in is_data.keys())) if is_data else []
        bs_years = sorted(set(yr for yr, _ in bs_data.keys())) if bs_data else []
        cf_years = sorted(set(yr for yr, _ in cf_data.keys())) if cf_data else []
        print(f"  IS years: {is_years}")
        print(f"  BS years: {bs_years}")
        print(f"  CF years: {cf_years}")

        is_metrics = sorted(set(m for _, m in is_data.keys())) if is_data else []
        bs_metrics = sorted(set(m for _, m in bs_data.keys())) if bs_data else []
        cf_metrics = sorted(set(m for _, m in cf_data.keys())) if cf_data else []
        print(f"  IS metrics: {is_metrics}")
        print(f"  BS metrics: {bs_metrics}")
        print(f"  CF metrics: {cf_metrics}")

        # Sample values for sanity check
        for yr in is_years[:1]:
            rev = is_data.get((yr, 'revenue'), is_data.get((yr, 'total_revenue')))
            ni = is_data.get((yr, 'net_income'))
            print(f"  Sanity: {yr} revenue={rev}, net_income={ni}")

        for (yr, metric), val in is_data.items():
            master[('is', yr, metric)] = val
        for (yr, metric), val in bs_data.items():
            master[('bs', yr, metric)] = val
        for (yr, metric), val in cf_data.items():
            master[('cf', yr, metric)] = val

    # Summary
    print(f"\n{'='*60}")
    print("MASTER DATA SUMMARY")
    stmt_counts = defaultdict(lambda: defaultdict(int))
    for (stmt, yr, metric), val in master.items():
        stmt_counts[stmt][yr] += 1

    for stmt in ['is', 'bs', 'cf']:
        if stmt not in stmt_counts:
            continue
        years = sorted(stmt_counts[stmt].keys())
        print(f"\n  {stmt.upper()}: {len(years)} years ({min(years)}-{max(years)})")
        for yr in years:
            print(f"    {yr}: {stmt_counts[stmt][yr]} metrics")

    total = len(master)
    print(f"\n  TOTAL: {total} datapoints to upsert")

    # Upsert
    print(f"\n{'='*60}")
    print("UPSERTING TO POSTGRESQL...")

    conn = psycopg2.connect("postgresql://vertex:VxwQLqiNmaZKblDN7pg@localhost:5432/vertex_db")
    cur = conn.cursor()

    upserted = 0
    errors = 0

    for (stmt, yr, metric), val in sorted(master.items()):
        try:
            val_full = val * 1e6
            cur.execute("""
                INSERT INTO stress_v2.historical_data (version_id, statement, year, metric, value, source)
                VALUES (%s, %s, %s, %s, %s, %s)
                ON CONFLICT (version_id, statement, year, metric) DO UPDATE
                SET value = EXCLUDED.value, source = EXCLUDED.source, updated_at = NOW()
            """, (VID, stmt, yr, metric, val_full, 'fs_excel'))
            upserted += 1
        except Exception as e:
            print(f"  ERROR: {stmt}/{yr}/{metric}: {e}")
            conn.rollback()
            errors += 1

    conn.commit()

    cur.execute("""
        SELECT statement, count(*), count(DISTINCT year), count(DISTINCT metric)
        FROM stress_v2.historical_data
        WHERE version_id = %s AND source = 'fs_excel'
        GROUP BY statement ORDER BY statement
    """, (VID,))
    print(f"\nVERIFICATION (source='fs_excel'):")
    for row in cur.fetchall():
        print(f"  {row[0].upper()}: {row[1]} rows, {row[2]} years, {row[3]} unique metrics")

    cur.execute("""
        SELECT statement, count(*), count(DISTINCT year), count(DISTINCT metric)
        FROM stress_v2.historical_data
        WHERE version_id = %s
        GROUP BY statement ORDER BY statement
    """, (VID,))
    print(f"\nTOTAL (all sources):")
    for row in cur.fetchall():
        print(f"  {row[0].upper()}: {row[1]} rows, {row[2]} years, {row[3]} unique metrics")

    conn.close()
    print(f"\nDONE: {upserted} upserted, {errors} errors")


if __name__ == '__main__':
    main()
